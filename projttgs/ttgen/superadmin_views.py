"""Super Admin module.

Self-contained feature:
  * Inbuilt login validated against credentials stored in .env
    (``SUPERADMIN_USERS`` parsed by settings into ``settings.SUPERADMIN_USERS``).
    There is no self sign-up for super admins.
  * A dedicated, real-data analytics dashboard on its own URL.

IMPORTANT: This module never touches the scheduling / generation algorithm.
It only *reads* timetable data that already exists in the database and turns
it into university-wide analytics. Every number below is computed live from
real records (no hardcoded analytics values).
"""

from __future__ import annotations

import hmac
import io
import logging
import re
import csv
from collections import defaultdict
from datetime import datetime
from functools import wraps

from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.http import HttpResponse, JsonResponse, Http404
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from django.db.models import Max, Count, Q

from .models import (
    AdminTeacher,
    CoordinatorAppointment,
    Department,
    Instructor,
    MeetingTime,
    Room,
    ScheduledSlot,
    SavedTimetable,
    Section,
    TeacherSection,
)

logger = logging.getLogger(__name__)


def _brand_from_email():
    from email.utils import formataddr

    sender = (getattr(settings, "DEFAULT_FROM_EMAIL", "") or settings.EMAIL_HOST_USER or "").strip()
    if not sender:
        return ""
    if "<" in sender and ">" in sender:
        return sender
    return formataddr(("SmartScheduler", sender))

SESSION_FLAG = "is_superadmin"
SESSION_EMAIL = "superadmin_email"
SESSION_OWNER = "superadmin_owner_id"
SA_IMPERSONATE = "sa_impersonate_uid"

WORKING_DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
_DAY_ORDER = {d: i for i, d in enumerate(WORKING_DAYS)}
# Lunch is slot 5 (1-indexed, 9 slots/day). Used to keep multi-hour lectures
# from bleeding across the lunch break when expanding their continuation slots.
LUNCH_SLOT = 5
_SEM_RE = re.compile(r"(\d+)\s*(?:st|nd|rd|th)?\s*sem", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------
def _superadmin_users():
    return getattr(settings, "SUPERADMIN_USERS", {}) or {}


def _check_credentials(email, password):
    """Constant-time validation against the .env credential list."""
    email = (email or "").strip().lower()
    password = password or ""
    expected = _superadmin_users().get(email)
    if not expected:
        # Still burn a comparison to reduce timing signal.
        hmac.compare_digest(password, password)
        return False
    return hmac.compare_digest(password, expected)


def _disable_response_cache(response):
    """Prevent browser back-forward cache from resurfacing protected pages."""
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0, private"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def superadmin_required(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.session.get(SESSION_FLAG):
            messages.info(request, "Please sign in as Super Admin to continue.")
            return redirect("superadmin_login")
        exempt_views = {"superadmin_choose_user", "superadmin_select_user", "superadmin_logout"}
        selected_owner_id = request.session.get(SESSION_OWNER)
        if view_func.__name__ not in exempt_views:
            if not selected_owner_id:
                messages.info(request, "Choose a user account first to load that account's data.")
                return redirect("superadmin_choose_user")
            if not get_user_model().objects.filter(id=selected_owner_id).exists():
                request.session.pop(SESSION_OWNER, None)
                request.session.modified = True
                messages.info(request, "Select a valid user account to continue.")
                return redirect("superadmin_choose_user")
        response = view_func(request, *args, **kwargs)
        return _disable_response_cache(response)

    return _wrapped


def superadmin_login(request):
    if request.session.get(SESSION_FLAG):
        return redirect("superadmin_choose_user")

    if request.method == "POST":
        email = (request.POST.get("email") or "").strip()
        password = request.POST.get("password") or ""
        if not _superadmin_users():
            messages.error(
                request,
                "Super Admin login is not configured. Set SUPERADMIN_USERS in .env.",
            )
        elif _check_credentials(email, password):
            request.session[SESSION_FLAG] = True
            request.session[SESSION_EMAIL] = email.strip().lower()
            request.session.pop(SESSION_OWNER, None)
            request.session.modified = True
            return redirect("superadmin_choose_user")
        else:
            messages.error(request, "Invalid email or password.")

    return _disable_response_cache(render(request, "superadmin_login.html"))


def superadmin_logout(request):
    request.session.pop(SESSION_FLAG, None)
    request.session.pop(SESSION_EMAIL, None)
    request.session.pop(SESSION_OWNER, None)
    request.session.pop(SA_IMPERSONATE, None)
    request.session.modified = True
    messages.success(request, "Signed out of Super Admin.")
    return _disable_response_cache(redirect("home"))


def _all_user_accounts(query=""):
    users = get_user_model().objects.all().order_by("username")
    query = (query or "").strip()
    if query:
        users = users.filter(Q(username__icontains=query) | Q(email__icontains=query))
    timetable_counts = dict(
        SavedTimetable.objects.values("user").annotate(c=Count("id")).values_list("user", "c")
    )
    accounts = []
    for user in users:
        full_name = (user.get_full_name() or "").strip()
        username = user.username or ""
        label = username or full_name or (user.email or "").split("@")[0] or f"Account #{user.id}"
        accounts.append({
            "id": user.id,
            "label": label,
            "display_name": full_name,
            "email": user.email or "",
            "username": username,
            "timetables": timetable_counts.get(user.id, 0),
        })
    return accounts


def _selected_owner_account(request):
    owner_id = request.session.get(SESSION_OWNER)
    if not owner_id:
        return None
    for account in _all_user_accounts():
        if account["id"] == owner_id:
            return account
    return None


@superadmin_required
def superadmin_choose_user(request):
    q = (request.GET.get("q") or "").strip()
    accounts = _all_user_accounts(q)
    current = _selected_owner_account(request)
    return render(
        request,
        "superadmin_choose_user.html",
        {
            "superadmin_email": request.session.get(SESSION_EMAIL, ""),
            "accounts": accounts,
            "account_total": len(accounts),
            "account_query": q,
            "selected_owner": current,
        },
    )


@superadmin_required
@require_POST
def superadmin_select_user(request):
    owner_id = (request.POST.get("owner_id") or "").strip()
    if not owner_id.isdigit():
        messages.error(request, "Choose a valid user account.")
        return redirect("superadmin_choose_user")

    user = get_user_model().objects.filter(id=int(owner_id)).first()
    if user is None:
        messages.error(request, "Selected user account was not found.")
        return redirect("superadmin_choose_user")

    request.session[SESSION_OWNER] = user.id
    request.session.modified = True
    messages.success(request, f"Now viewing data for {user.username or user.email or f'Account #{user.id}'}.")
    return redirect("superadmin_dashboard")


# ---------------------------------------------------------------------------
# Analytics helpers (all read-only, real DB data)
# ---------------------------------------------------------------------------
def _semester_of(section_id):
    match = _SEM_RE.search(section_id or "")
    if not match:
        return ""
    number = match.group(1)
    if number in {"11", "12", "13"}:
        suffix = "th"
    else:
        suffix = {"1": "st", "2": "nd", "3": "rd"}.get(number[-1], "th")
    return f"{number}{suffix} Sem"


def _slot_universe():
    """Distinct (day, time) cells that the institution actually uses."""
    pairs = set(MeetingTime.objects.values_list("day", "time").distinct())
    universe = {(d, str(t)) for d, t in pairs if d}
    if not universe:
        universe = {(d, str(t)) for d in WORKING_DAYS for t in range(1, 9)}
    return universe


def _active_timetable_ids(owners=None):
    """Latest saved timetable per (user, department) = current live picture.

    When ``owners`` is given, only timetables owned by those accounts are
    considered, so analytics reflect a single chosen account instead of the
    whole platform.
    """
    qs = SavedTimetable.objects.all()
    if owners:
        qs = qs.filter(user_id__in=owners)
    rows = qs.values("user", "department").annotate(mx=Max("id"))
    return [row["mx"] for row in rows if row["mx"]]


def _active_owner_ids():
    """User IDs that own at least one *non-empty* saved timetable.

    Master data (departments, teachers, rooms, sections) is stored per user.
    The same institution can exist under multiple accounts (seeded twice),
    which would double every count. Scoping analytics to the accounts that
    actually own scheduled timetables removes that duplication. Falls back to
    "all users" when nothing is scheduled yet, preserving the empty-DB view.
    """
    return set(
        SavedTimetable.objects.filter(slots__isnull=False)
        .values_list("user_id", flat=True)
        .distinct()
    )


def _owner_accounts(owners):
    """Selectable account list (id + readable label) for the account filter.

    One row per account that owns scheduled timetables, so the super admin can
    view analytics for a single coordinator account instead of every account
    combined (which inflates teacher loads, room counts, etc.).
    """
    if not owners:
        return []
    users = (
        get_user_model()
        .objects.filter(id__in=owners)
        .order_by("username")
    )
    accounts = []
    for u in users:
        full_name = (u.get_full_name() or "").strip()
        label = full_name or u.username or (u.email or "").split("@")[0] or f"Account #{u.id}"
        accounts.append({"id": u.id, "label": label, "email": u.email or ""})
    return accounts


def _request_owners(request):
    """Resolve the owner account selected in the super-admin session."""
    selected_owner_id = request.session.get(SESSION_OWNER)
    if selected_owner_id and get_user_model().objects.filter(id=selected_owner_id).exists():
        return {selected_owner_id}, str(selected_owner_id)
    return set(), ""


def _scope_to_owners(qs, owners):
    """Restrict a user-scoped queryset to the active owners (no-op if none)."""
    return qs.filter(user_id__in=owners) if owners else qs.all()


def _pct(used, total):
    if not total:
        return 0.0
    return round((used / total) * 100.0, 1)


def _theory_continuation_cells(slot):
    """Extra (day, time) cells a multi-hour theory lecture occupies.

    Theory continuation slots are not stored on ScheduledSlot — the duration
    lives on the Subject. A 2-hour lecture saved at slot N also occupies slot
    N+1, so analytics/grids must expand it to count every hour. Lunch slots are
    skipped so a lecture never bleeds across the lunch break.
    """
    try:
        duration = int(getattr(slot.subject, "duration", 1) or 1)
    except (TypeError, ValueError):
        duration = 1
    if duration <= 1:
        return []
    mt = slot.meeting_time
    if not mt or not str(mt.time).isdigit():
        return []
    day = mt.day
    start = int(mt.time)
    extra = []
    nxt = start + 1
    while len(extra) < duration - 1 and nxt <= 9:
        if nxt != int(LUNCH_SLOT):
            extra.append((day, str(nxt)))
        nxt += 1
    return extra


def _slot_cells(slot):
    """All unique timetable cells occupied by one scheduled slot.

    Labs already persist every occupied meeting time in ``lab_slots`` including
    the start hour, while theory continuation hours must be expanded from the
    subject duration. This helper normalises both paths so workload is counted
    by occupied hour, not by duplicated section assignments.
    """
    if slot.is_lab:
        cells = [(mt.day, str(mt.time)) for mt in slot.lab_slots.all()]
        if not cells and slot.meeting_time:
            cells = [(slot.meeting_time.day, str(slot.meeting_time.time))]
    else:
        cells = [(slot.meeting_time.day, str(slot.meeting_time.time))]
        cells += _theory_continuation_cells(slot)
    return list(dict.fromkeys(c for c in cells if c[0]))


def _is_lunch_cell(cell):
    return str(cell[1]) == str(LUNCH_SLOT)


def _effective_room_capacity(universe, occupied_cells):
    """Lunch slot does not reduce utilization unless that room is actually used then.

    Slot 5 is treated as lunch by default, so an empty lunch slot should not make
    a fully booked room look under-utilized. If a prefilled class occupies lunch,
    that occupied lunch cell is counted as real capacity for that room.
    """
    non_lunch = sum(1 for cell in universe if not _is_lunch_cell(cell))
    lunch_used = sum(1 for cell in occupied_cells if _is_lunch_cell(cell))
    return non_lunch + lunch_used


def _format_cell_block(cells):
    """Human-readable label for one contiguous scheduled block."""
    if not cells:
        return "—"
    ordered = sorted(cells, key=lambda item: (_DAY_ORDER.get(item[0], 99), int(item[1]) if str(item[1]).isdigit() else 99))
    day = ordered[0][0]
    nums = [int(time) for _day, time in ordered if str(time).isdigit()]
    if nums:
        start = nums[0]
        end = nums[-1]
        return f"{day} {start}" if start == end else f"{day} {start}-{end}"
    return f"{day} {ordered[0][1]}"


def _active_slots(owners=None):
    ids = _active_timetable_ids(owners)
    if not ids:
        return ScheduledSlot.objects.none()
    return (
        ScheduledSlot.objects.filter(timetable_id__in=ids)
        .select_related(
            "section",
            "section__department",
            "subject",
            "instructor",
            "second_instructor",
            "room",
            "room__department",
            "meeting_time",
        )
        .prefetch_related("lab_slots")
    )


def _build_analytics(request):
    """Compute the full analytics context from real timetable records."""
    dept_filter = (request.GET.get("dept") or "all").strip()
    sem_filter = (request.GET.get("sem") or "all").strip()
    search = (request.GET.get("q") or "").strip()

    universe = _slot_universe()
    base_room_capacity = sum(1 for cell in universe if not _is_lunch_cell(cell)) or 1

    # Accounts that actually own scheduled timetables. Master data is per-user
    # and may be duplicated across accounts. By default we scope everything to
    # every active owner (deduping by identity), but the super admin can pick a
    # single account via ?account=<id> to see that account's data on its own
    # instead of every account combined (which inflates loads / counts).
    all_owners = _active_owner_ids()
    owners, account_filter = _request_owners(request)
    accounts = _owner_accounts(all_owners)

    # Master counts (whole institution, independent of current filter).
    # The same institution can exist under several accounts, so the same
    # teacher / department / room may be stored many times. Count DISTINCT
    # real-world entities (by identity, not by DB row) so totals reflect
    # reality instead of being inflated by duplicate seed data.
    def _norm(value):
        return (value or "").strip().lower()

    teacher_rows = _scope_to_owners(Instructor.objects, owners).values_list("uid", "name")
    total_teachers = len({(_norm(uid), _norm(name)) for uid, name in teacher_rows})

    dept_rows = _scope_to_owners(Department.objects, owners).values_list("name", "code")
    total_departments = len({(_norm(name), _norm(code)) for name, code in dept_rows})

    room_rows = list(
        _scope_to_owners(Room.objects, owners).values_list(
            "r_number", "room_type", "department__name"
        )
    )
    total_rooms = len({(_norm(rn), _norm(dn)) for rn, rt, dn in room_rows})
    total_labs = len({(_norm(rn), _norm(dn)) for rn, rt, dn in room_rows if rt == "Lab"})

    section_rows = _scope_to_owners(Section.objects, owners).values_list("section_id", flat=True)
    total_sections = len({_norm(sid) for sid in section_rows})

    # Total registered accounts on the platform.
    total_users = get_user_model().objects.count()


    # Filter dropdown sources (dynamic, from real data).
    departments = list(
        _scope_to_owners(Department.objects, owners)
        .order_by("name")
        .values("id", "name", "code")
    )
    semesters = sorted(
        {
            _semester_of(sid)
            for sid in _scope_to_owners(Section.objects, owners).values_list(
                "section_id", flat=True
            )
            if _semester_of(sid)
        },
        key=lambda s: int(re.match(r"(\d+)", s).group(1)),
    )

    slots = _active_slots(owners)

    # Accumulators.
    room_cells = defaultdict(set)            # room_id -> {(day,time)}
    dept_room_cells = defaultdict(set)       # dept_id -> {(room_id,day,time)}
    heat = defaultdict(int)                  # (day,time) -> occupancy
    teacher_load = defaultdict(set)          # instructor_id -> {(day,time)} taught
    teacher_name = {}
    room_meta = {}                           # room_id -> {number,dept,is_lab}
    dept_meta = {}                           # dept_id -> {name,code}
    section_scheduled = set()
    dept_scheduled = set()
    teacher_scheduled = set()

    room_slot_sections = defaultdict(set)    # (room_id,day,time) -> {section_id}
    teacher_slot_sections = defaultdict(set)  # (instr_id,day,time) -> {section_id}
    section_slot_subjects = defaultdict(set)  # (section_id,day,time) -> {subject_id}

    preview_rows = []

    for slot in slots:
        section = slot.section
        if not section:
            continue
        sec_id = section.section_id
        dept = section.department
        dept_id = dept.id if dept else None
        sem = _semester_of(sec_id)

        # Apply filters.
        if dept_filter != "all" and str(dept_id) != dept_filter:
            continue
        if sem_filter != "all" and sem != sem_filter:
            continue
        if search:
            blob = " ".join(
                str(x).lower()
                for x in (
                    sec_id,
                    getattr(slot.subject, "subject_name", ""),
                    getattr(slot.instructor, "name", ""),
                    getattr(slot.room, "r_number", ""),
                )
            )
            if search.lower() not in blob:
                continue

        room = slot.room
        room_id = room.id if room else None
        if room_id is not None and room_id not in room_meta:
            room_meta[room_id] = {
                "number": room.r_number,
                "dept": room.department.name if room.department_id else "—",
                "dept_id": room.department_id,
                "is_lab": room.room_type == "Lab",
            }
        if dept_id is not None and dept_id not in dept_meta:
            dept_meta[dept_id] = {"name": dept.name, "code": dept.code}

        instr = slot.instructor
        instr_id = instr.id if instr else None
        if instr_id is not None:
            teacher_name[instr_id] = instr.name
            teacher_scheduled.add(instr_id)
        second = slot.second_instructor
        second_id = second.id if second else None
        if second_id is not None:
            teacher_name[second_id] = second.name

        section_scheduled.add(sec_id)
        if dept_id is not None:
            dept_scheduled.add(dept_id)

        # Count unique occupied hours so parallel sections in the same hour do
        # not inflate a teacher's workload beyond real teaching time.
        cells = _slot_cells(slot)

        for day, time in cells:
            if room_id is not None:
                room_cells[room_id].add((day, time))
                room_slot_sections[(room_id, day, time)].add(sec_id)
                if room.department_id:
                    dept_room_cells[room.department_id].add((room_id, day, time))
            heat[(day, time)] += 1
            if instr_id is not None:
                teacher_load[instr_id].add((day, time))
                teacher_slot_sections[(instr_id, day, time)].add(sec_id)
            if second_id is not None:
                teacher_load[second_id].add((day, time))
            section_slot_subjects[(sec_id, day, time)].add(slot.subject_id)

        preview_rows.append(
            {
                "section": sec_id,
                "subject": getattr(slot.subject, "subject_name", "—"),
                "teacher": getattr(instr, "name", "—"),
                "room": getattr(room, "r_number", "—"),
                "day": slot.meeting_time.day,
                "slot": str(slot.meeting_time.time),
                "type": "Lab" if slot.is_lab else "Theory",
                "dept": dept.name if dept else "—",
            }
        )

    # Scope room set for utilization denominators.
    scope_rooms = _scope_to_owners(Room.objects, owners)
    if dept_filter != "all":
        scope_rooms = scope_rooms.filter(department_id=dept_filter)
    scope_room_list = list(scope_rooms.values("id", "r_number", "room_type", "department__name"))
    scope_room_count = len(scope_room_list) or 1

    # Overall utilization.
    used_cells = sum(len(v) for v in room_cells.values())
    room_capacity_map = {
        room["id"]: _effective_room_capacity(universe, room_cells.get(room["id"], set()))
        for room in scope_room_list
    }

    available_cells = sum(room_capacity_map.values()) or 1
    overall_util = _pct(used_cells, available_cells)

    # Conflicts.
    room_conflicts = sum(1 for v in room_slot_sections.values() if len(v) > 1)
    teacher_conflicts = sum(1 for v in teacher_slot_sections.values() if len(v) > 1)
    section_conflicts = sum(1 for v in section_slot_subjects.values() if len(v) > 1)
    total_conflicts = room_conflicts + teacher_conflicts + section_conflicts

    # Department-wise utilization.
    dept_util = []
    dept_qs = _scope_to_owners(Department.objects, owners)
    if dept_filter != "all":
        dept_qs = dept_qs.filter(id=dept_filter)
    for dept in dept_qs.order_by("name"):
        dept_room_ids = list(
            _scope_to_owners(Room.objects.filter(department=dept), owners).values_list("id", flat=True)
        )
        d_rooms = len(dept_room_ids)
        used = len(dept_room_cells.get(dept.id, set()))
        denom = sum(room_capacity_map.get(room_id, base_room_capacity) for room_id in dept_room_ids)
        dept_util.append(
            {
                "id": dept.id,
                "name": dept.name,
                "code": dept.code,
                "util": _pct(used, denom),
                "used": used,
                "rooms": d_rooms,
                "scheduled": dept.id in dept_scheduled,
            }
        )
    dept_util.sort(key=lambda d: d["util"], reverse=True)

    # Room utilization list (includes idle rooms with 0%).
    room_util = []
    for room in scope_room_list:
        used = len(room_cells.get(room["id"], set()))
        capacity = room_capacity_map.get(room["id"], base_room_capacity)
        room_util.append(
            {
                "id": room["id"],
                "number": room["r_number"],
                "type": room["room_type"],
                "dept": room["department__name"] or "—",
                "used": used,
                "capacity": capacity,
                "util": _pct(used, capacity),
            }
        )
    room_util.sort(key=lambda r: r["util"], reverse=True)
    top_rooms = room_util[:10]
    least_rooms = sorted(room_util, key=lambda r: r["util"])[:5]
    idle_rooms = [r for r in room_util if r["used"] == 0]

    # Teacher workload (all in-scope teachers; 0 for unscheduled).
    teacher_qs = _scope_to_owners(Instructor.objects, owners)
    if dept_filter != "all":
        # Teachers who teach at least one in-scope section.
        teacher_qs = teacher_qs.filter(id__in=list(teacher_load.keys()) or [0])
    teacher_workload = []
    seen_teacher = {}
    for instr in teacher_qs:
        load = len(teacher_load.get(instr.id, set()))
        key = (_norm(instr.uid), _norm(instr.name))
        entry = {
            "id": instr.id,
            "name": instr.name,
            "designation": instr.designation,
            "load": load,
            "max": instr.max_workload,
            "util": _pct(load, instr.max_workload),
        }
        existing = seen_teacher.get(key)
        if existing is None:
            seen_teacher[key] = entry
            teacher_workload.append(entry)
        elif load > existing["load"]:
            # Keep the scheduled instance (with real load) for this person.
            existing.update(entry)
    teacher_workload.sort(key=lambda t: t["load"], reverse=True)
    top_teachers = teacher_workload[:5]

    # Workload distribution buckets.
    buckets = {"0-10": 0, "10-15": 0, "15-20": 0, "20+": 0}
    for t in teacher_workload:
        load = t["load"]
        if load < 10:
            buckets["0-10"] += 1
        elif load < 15:
            buckets["10-15"] += 1
        elif load < 20:
            buckets["15-20"] += 1
        else:
            buckets["20+"] += 1

    # Resource utilisation overview (rooms vs labs vs all).
    lab_used = sum(
        len(room_cells.get(r["id"], set()))
        for r in scope_room_list
        if r["room_type"] == "Lab"
    )
    lab_count_scope = sum(1 for r in scope_room_list if r["room_type"] == "Lab") or 1
    lecture_rooms = [r for r in scope_room_list if r["room_type"] != "Lab"]
    lecture_used = sum(len(room_cells.get(r["id"], set())) for r in lecture_rooms)
    lecture_count_scope = len(lecture_rooms) or 1
    resource_overview = {
        "all": {"used": used_cells, "total": available_cells, "util": overall_util},
        "labs": {
            "used": lab_used,
            "total": sum(room_capacity_map.get(r["id"], base_room_capacity) for r in scope_room_list if r["room_type"] == "Lab") or 1,
            "util": _pct(lab_used, sum(room_capacity_map.get(r["id"], base_room_capacity) for r in scope_room_list if r["room_type"] == "Lab") or 1),
        },
        "rooms": {
            "used": lecture_used,
            "total": sum(room_capacity_map.get(r["id"], base_room_capacity) for r in lecture_rooms) or 1,
            "util": _pct(lecture_used, sum(room_capacity_map.get(r["id"], base_room_capacity) for r in lecture_rooms) or 1),
        },
    }

    # Heatmap matrix (days x sorted time slots).
    times = sorted({t for _, t in universe}, key=lambda x: int(x) if x.isdigit() else x)
    days = sorted({d for d, _ in universe}, key=lambda d: _DAY_ORDER.get(d, 99))
    max_heat = max(heat.values()) if heat else 0
    heatmap = {
        "days": days,
        "times": times,
        "max": max_heat,
        "rows": [
            {"day": d, "cells": [heat.get((d, t), 0) for t in times]} for d in days
        ],
    }

    # AI insights (dynamic, derived from the numbers above).
    insights = []
    if dept_util:
        best = dept_util[0]
        insights.append(
            f"{best['name']} has the highest resource utilisation at {best['util']}%."
        )
        worst = min((d for d in dept_util if d["rooms"]), key=lambda d: d["util"], default=None)
        if worst and worst["util"] < 40:
            insights.append(
                f"{worst['name']} is under-utilised ({worst['util']}%) — capacity to absorb more classes."
            )
    if idle_rooms:
        insights.append(
            f"{len(idle_rooms)} room(s) are completely idle in the current schedule."
        )
    if total_conflicts:
        insights.append(
            f"{total_conflicts} scheduling conflict(s) detected — review room/teacher overlaps."
        )
    else:
        insights.append("No hard conflicts detected in the active timetables.")
    if top_teachers and top_teachers[0]["load"]:
        busiest = top_teachers[0]
        insights.append(
            f"{busiest['name']} carries the heaviest load ({busiest['load']} hrs / {busiest['max']} max)."
        )
    if heatmap["max"]:
        peak_idx = max(
            range(len(times)),
            key=lambda i: max((heat.get((d, times[i]), 0) for d in days), default=0),
        )
        insights.append(f"Slot {times[peak_idx]} is the busiest time across the campus.")

    # Recent activity — prefer the real per-user audit log, fall back to
    # saved timetables when no activity has been recorded yet.
    from .models import ActivityLog

    recent = []
    _action_labels = dict(ActivityLog.ACTIONS)
    for ev in ActivityLog.objects.all().order_by("-created_at")[:10]:
        who = ev.username or ev.email or "Someone"
        bits = [b for b in [ev.summary or _action_labels.get(ev.action, ""), ev.detail] if b]
        recent.append({
            "who": who,
            "action": ev.action,
            "text": " · ".join(bits) or _action_labels.get(ev.action, "Activity"),
            "when": timezone.localtime(ev.created_at).strftime("%d %b %Y, %H:%M")
            if ev.created_at else "",
        })
    if not recent:
        for st in SavedTimetable.objects.select_related("department").order_by("-created_at")[:8]:
            recent.append({
                "who": "",
                "action": "publish" if st.is_published else "save",
                "text": ("Timetable for " + (st.department.name if st.department_id else "All Departments")
                         + (" published" if st.is_published else " saved")),
                "when": timezone.localtime(st.created_at).strftime("%d %b %Y, %H:%M")
                if st.created_at else "",
            })

    # Quick timetable preview ordering.
    preview_rows.sort(key=lambda r: (r["section"], _DAY_ORDER.get(r["day"], 99), r["slot"]))

    avg_workload = round(
        sum(t["load"] for t in teacher_workload) / len(teacher_workload), 1
    ) if teacher_workload else 0

    kpis = {
        "departments_active": len({_norm(dept_meta[d]["name"]) for d in dept_scheduled}) or total_departments,
        "teachers_active": len({_norm(teacher_name[t]) for t in teacher_scheduled}) or total_teachers,
        "sections_scheduled": len(section_scheduled),
        "rooms_total": total_rooms,
        "labs_total": total_labs,
        "classes_scheduled": len(preview_rows),
        "overall_util": overall_util,
        "conflicts": total_conflicts,
        "users_total": total_users,
    }

    return {
        "kpis": kpis,
        "totals": {
            "departments": total_departments,
            "teachers": total_teachers,
            "rooms": total_rooms,
            "labs": total_labs,
            "sections": total_sections,
            "users": total_users,
        },
        "departments": departments,
        "semesters": semesters,
        "dept_filter": dept_filter,
        "sem_filter": sem_filter,
        "search": search,
        "accounts": accounts,
        "account_filter": account_filter,
        "dept_util": dept_util,
        "room_util": room_util,
        "top_rooms": top_rooms,
        "least_rooms": least_rooms,
        "idle_rooms_count": len(idle_rooms),
        "teacher_workload": teacher_workload,
        "top_teachers": top_teachers,
        "workload_buckets": buckets,
        "resource_overview": resource_overview,
        "heatmap": heatmap,
        "insights": insights,
        "recent": recent,
        "preview": preview_rows,
        "avg_workload": avg_workload,
        "conflict_breakdown": {
            "room": room_conflicts,
            "teacher": teacher_conflicts,
            "section": section_conflicts,
        },
    }


# ---------------------------------------------------------------------------
# Dashboard + drilldown + exports
# ---------------------------------------------------------------------------
def _chart_data(ctx):
    """Chart-ready payloads shared by every super-admin page."""
    return {
        "deptUtil": {
            "labels": [d["name"] for d in ctx["dept_util"]],
            "values": [d["util"] for d in ctx["dept_util"]],
        },
        "workload": {
            "labels": list(ctx["workload_buckets"].keys()),
            "values": list(ctx["workload_buckets"].values()),
        },
        "resource": ctx["resource_overview"],
        "heatmap": ctx["heatmap"],
        "topRooms": {
            "labels": [r["number"] for r in ctx["top_rooms"]],
            "values": [r["util"] for r in ctx["top_rooms"]],
        },
        "leastRooms": {
            "labels": [r["number"] for r in ctx["least_rooms"]],
            "values": [r["util"] for r in ctx["least_rooms"]],
        },
        "topTeachers": {
            "labels": [t["name"] for t in ctx["top_teachers"]],
            "values": [t["load"] for t in ctx["top_teachers"]],
        },
    }


def _page_ctx(request, page):
    """Build the full analytics context for any super-admin page."""
    ctx = _build_analytics(request)
    ctx["superadmin_email"] = request.session.get(SESSION_EMAIL, "")
    ctx["selected_owner"] = _selected_owner_account(request)
    ctx["page"] = page
    ctx["chart_data"] = _chart_data(ctx)
    return ctx


def _basic_page_ctx(request, page):
    """Minimal shell context for super-admin pages that do not need analytics."""
    return {
        "superadmin_email": request.session.get(SESSION_EMAIL, ""),
        "selected_owner": _selected_owner_account(request),
        "page": page,
    }


def _selected_owner_guard_or_404(request, owner_id):
    selected_owner_id = request.session.get(SESSION_OWNER)
    if selected_owner_id and owner_id == selected_owner_id:
        return
    raise Http404("This record is outside the selected user scope")


def _csv_cell(row, *names):
    lowered = {(str(k or "").strip().lower()): v for k, v in row.items()}
    for name in names:
        value = lowered.get(name.lower())
        if value is not None:
            return str(value).strip()
    return ""


def _department_name_from_code(code):
    dept_code = (code or "").strip()
    if not dept_code:
        return ""
    match = Department.objects.filter(code__iexact=dept_code).order_by("id").values_list("name", flat=True).first()
    return (match or "").strip()


def _import_admin_teacher_csv(uploaded_file):
    raw = uploaded_file.read().decode("utf-8-sig", errors="ignore")
    rows = csv.DictReader(io.StringIO(raw))
    created = 0
    updated = 0
    skipped = 0
    warnings = []

    with transaction.atomic():
        for idx, row in enumerate(rows, start=2):
            try:
                name = _csv_cell(row, "name", "teacher name", "teacher_name", "faculty name", "faculty_name")
                email = _csv_cell(row, "email", "mail", "email id", "email_id")
                uid = _csv_cell(row, "uid", "teacher_id", "teacher id", "teacher code", "teacher_code", "faculty uid", "faculty_uid", "code")
                department_name = _csv_cell(row, "department", "department name", "department_name", "branch")
                department_code = _csv_cell(row, "department code", "department_code", "dept code", "dept_code")
                contact_number = _csv_cell(row, "contact", "contact number", "contact_number", "phone", "mobile")
                designation = _csv_cell(row, "designation") or "Associate Professor"

                if not department_name and department_code:
                    department_name = _department_name_from_code(department_code)

                if not name and not email and not uid:
                    skipped += 1
                    warnings.append(f"Row {idx}: skipped because name, email and teacher code were all empty.")
                    continue

                if designation not in dict(AdminTeacher.DESIGNATION_CHOICES):
                    designation = "Associate Professor"

                lookup = None
                if uid:
                    lookup = AdminTeacher.objects.filter(uid__iexact=uid).first()
                if lookup is None and email:
                    lookup = AdminTeacher.objects.filter(email__iexact=email).first()
                if lookup is None and name:
                    lookup = AdminTeacher.objects.filter(name__iexact=name, department_name__iexact=department_name).first()

                payload = {
                    "name": name or (lookup.name if lookup else ""),
                    "email": email or (lookup.email if lookup else ""),
                    "uid": uid or (lookup.uid if lookup else ""),
                    "contact_number": contact_number or (lookup.contact_number if lookup else ""),
                    "designation": designation,
                    "department_name": department_name or (lookup.department_name if lookup else ""),
                    "department_code": department_code or (lookup.department_code if lookup else ""),
                    "is_active": True,
                }

                if lookup:
                    for field, value in payload.items():
                        setattr(lookup, field, value)
                    lookup.save()
                    updated += 1
                else:
                    AdminTeacher.objects.create(**payload)
                    created += 1
            except Exception as exc:
                skipped += 1
                warnings.append(f"Row {idx}: not uploaded ({exc}).")

    return created, updated, skipped, warnings


@superadmin_required
def superadmin_dashboard(request):
    return render(request, "superadmin_dashboard.html", _page_ctx(request, "overview"))


@superadmin_required
def superadmin_resource(request):
    return render(request, "superadmin_resource.html", _page_ctx(request, "resource"))


@superadmin_required
def superadmin_teachers(request):
    if request.method == "POST":
        action = (request.POST.get("action") or "").strip()
        if action == "delete_all_central_teachers":
            deleted_count, _ = AdminTeacher.objects.all().delete()
            messages.success(request, f"Deleted {deleted_count} central teacher record(s).")
            return redirect("superadmin_teachers")

        if request.FILES.get("csv_file"):
            try:
                created, updated, skipped, warnings = _import_admin_teacher_csv(request.FILES["csv_file"])
                messages.success(request, f"Central teachers imported. Created: {created}, updated: {updated}, skipped: {skipped}.")
                if warnings:
                    preview = " ".join(warnings[:5])
                    if len(warnings) > 5:
                        preview += f" And {len(warnings) - 5} more warning(s)."
                    messages.warning(request, preview)
            except Exception:
                logger.exception("Central teacher CSV import failed")
                messages.error(request, "Could not import the teacher CSV. Check the file headers and try again.")
            return redirect("superadmin_teachers")

    ctx = _page_ctx(request, "teachers")
    central_teachers = list(AdminTeacher.objects.order_by("name", "uid", "id"))
    ctx["central_teachers"] = central_teachers
    ctx["central_teacher_total"] = AdminTeacher.objects.count()
    return render(request, "superadmin_teachers.html", ctx)


@superadmin_required
def superadmin_teacher_edit(request, teacher_id):
    teacher = get_object_or_404(AdminTeacher, id=teacher_id)

    if request.method == "POST":
        teacher.name = (request.POST.get("name") or "").strip()
        teacher.uid = (request.POST.get("uid") or "").strip()
        teacher.email = (request.POST.get("email") or "").strip()
        teacher.contact_number = (request.POST.get("contact_number") or "").strip()
        designation = (request.POST.get("designation") or "").strip()
        if designation in dict(AdminTeacher.DESIGNATION_CHOICES):
            teacher.designation = designation
        teacher.department_code = (request.POST.get("department_code") or "").strip()
        teacher.department_name = (request.POST.get("department_name") or "").strip()
        if not teacher.department_name and teacher.department_code:
            teacher.department_name = _department_name_from_code(teacher.department_code)
        teacher.is_active = (request.POST.get("is_active") or "on") == "on"

        if not teacher.name:
            messages.error(request, "Teacher name is required.")
        else:
            teacher.save()
            messages.success(request, f"Updated central teacher {teacher.name}.")
            return redirect("superadmin_teachers")

    ctx = _basic_page_ctx(request, "teachers")
    ctx["target_teacher"] = teacher
    ctx["designation_choices"] = AdminTeacher.DESIGNATION_CHOICES
    return render(request, "superadmin_teacher_edit.html", ctx)


@superadmin_required
def superadmin_teacher_delete(request, teacher_id):
    teacher = get_object_or_404(AdminTeacher, id=teacher_id)

    if request.method == "POST":
        label = teacher.name or teacher.uid or f"Teacher #{teacher.id}"
        teacher.delete()
        messages.success(request, f"Deleted central teacher {label}.")
        return redirect("superadmin_teachers")

    ctx = _basic_page_ctx(request, "teachers")
    ctx["target_teacher"] = teacher
    return render(request, "superadmin_teacher_confirm_delete.html", ctx)


@superadmin_required
def superadmin_depts(request):
    return render(request, "superadmin_depts.html", _page_ctx(request, "depts"))


@superadmin_required
def superadmin_slots(request):
    return render(request, "superadmin_slots.html", _page_ctx(request, "slots"))


@superadmin_required
def superadmin_explorer(request):
    return render(request, "superadmin_explorer.html", _page_ctx(request, "explorer"))


@superadmin_required
def superadmin_saved_page(request):
    return render(request, "superadmin_saved.html", _page_ctx(request, "saved"))


@superadmin_required
def superadmin_preview(request):
    return render(request, "superadmin_preview.html", _page_ctx(request, "preview"))


@superadmin_required
def superadmin_activity(request):
    from .models import ActivityLog, UserSession

    ctx = _page_ctx(request, "activity")

    user_query = (request.GET.get("user") or "").strip()

    logs = ActivityLog.objects.all()
    if user_query:
        logs = logs.filter(
            Q(username__icontains=user_query) | Q(email__icontains=user_query)
        )

    action_labels = dict(ActivityLog.ACTIONS)

    def _ukey(email, username):
        return (email or username or "").lower()

    # ---- Detailed event feed (most recent first) ----
    feed = []
    for ev in logs.order_by("-created_at")[:250]:
        feed.append({
            "user": ev.username or ev.email or "Unknown",
            "email": ev.email,
            "action": ev.action,
            "action_label": action_labels.get(ev.action, ev.action.title()),
            "summary": ev.summary or action_labels.get(ev.action, ""),
            "detail": ev.detail,
            "method": ev.method,
            "ip": ev.ip,
            "when": timezone.localtime(ev.created_at).strftime("%d %b %Y, %H:%M:%S"),
        })

    # ---- Per-user rollup (counts + sessions + active time) ----
    counts = (
        ActivityLog.objects.values("username", "email")
        .annotate(n=Count("id"), last=Max("created_at"))
    )
    cards = {}
    for row in counts:
        key = _ukey(row["email"], row["username"])
        cards.setdefault(key, {
            "user": row["username"] or row["email"] or "Unknown",
            "email": row["email"],
            "total": 0, "by_action": {}, "last_active": None,
            "sessions": 0, "active_seconds": 0, "last_login": None,
            "events": [],
        })
        cards[key]["user"] = row["username"] or cards[key]["user"]
        cards[key]["total"] += row["n"]
        if row["last"] and (cards[key]["last_active"] is None or row["last"] > cards[key]["last_active"]):
            cards[key]["last_active"] = row["last"]

    # action breakdown per user
    for row in ActivityLog.objects.values("username", "email", "action").annotate(n=Count("id")):
        key = _ukey(row["email"], row["username"])
        if key in cards:
            cards[key]["by_action"][row["action"]] = (
                cards[key]["by_action"].get(row["action"], 0) + row["n"]
            )

    # Per-user recent "main" activities (skip pure heartbeats/other).
    MAIN_ACTIONS = {
        "login", "logout", "generate", "save", "delete", "export",
        "move", "park", "restore", "add", "edit", "substitute",
        "publish", "unpublish",
    }
    for ev in (
        ActivityLog.objects.filter(action__in=MAIN_ACTIONS)
        .order_by("-created_at")[:2000]
    ):
        key = _ukey(ev.email, ev.username)
        card = cards.get(key)
        if card is None or len(card["events"]) >= 15:
            continue
        card["events"].append({
            "action": ev.action,
            "label": action_labels.get(ev.action, ev.action.title()),
            "summary": ev.summary or action_labels.get(ev.action, ""),
            "detail": ev.detail,
            "when": timezone.localtime(ev.created_at).strftime("%d %b %Y, %H:%M"),
        })

    # sessions + active time per user
    for s in UserSession.objects.all():
        key = _ukey(s.email, s.username)
        if not key:
            continue
        card = cards.setdefault(key, {
            "user": s.username or s.email or "Unknown",
            "email": s.email,
            "total": 0, "by_action": {}, "last_active": None,
            "sessions": 0, "active_seconds": 0, "last_login": None,
            "events": [],
        })
        card["sessions"] += 1
        card["active_seconds"] += s.duration_seconds
        if card["last_login"] is None or s.login_at > card["last_login"]:
            card["last_login"] = s.login_at

    user_cards = []
    for idx, c in enumerate(cards.values()):
        if user_query:
            hay = f"{c['user']} {c['email']}".lower()
            if user_query.lower() not in hay:
                continue
        user_cards.append({
            "uid": idx,
            "user": c["user"],
            "email": c["email"],
            "total": c["total"],
            "sessions": c["sessions"],
            "active_label": _fmt_duration(c["active_seconds"]),
            "by_action": [
                {"action": a, "label": action_labels.get(a, a.title()), "n": n}
                for a, n in sorted(c["by_action"].items(), key=lambda kv: -kv[1])
            ],
            "events": c["events"],
            "last_active": timezone.localtime(c["last_active"]).strftime("%d %b %Y, %H:%M")
            if c["last_active"] else "—",
            "last_login": timezone.localtime(c["last_login"]).strftime("%d %b %Y, %H:%M")
            if c["last_login"] else "—",
        })
    user_cards.sort(key=lambda c: c["total"], reverse=True)

    ctx["feed"] = feed
    ctx["user_cards"] = user_cards
    ctx["activity_total"] = ActivityLog.objects.count()
    ctx["user_query"] = user_query
    return render(request, "superadmin_activity.html", ctx)


def _fmt_duration(seconds):
    seconds = int(seconds or 0)
    if seconds < 60:
        return f"{seconds}s"
    mins, sec = divmod(seconds, 60)
    if mins < 60:
        return f"{mins}m {sec}s"
    hrs, mins = divmod(mins, 60)
    if hrs < 24:
        return f"{hrs}h {mins}m"
    days, hrs = divmod(hrs, 24)
    return f"{days}d {hrs}h"


# ---------------------------------------------------------------------------
# Appoint coordinators (search/add a teacher, assign a role, email them)
# ---------------------------------------------------------------------------

APPOINT_ROLES = [
    "Timetable Coordinator",
    "University Timetable Coordinator",
]

_EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

# Logo embedded inline in emails (referenced as ``cid:smartscheduler_logo``).
LOGO_CID = "smartscheduler_logo"


def _attach_logo(msg):
    """Embed the real SmartScheduler logo into an email as an inline image.

    Uses a Content-ID attachment so the HTML can reference ``cid:...`` — this
    renders reliably in Gmail/Outlook (unlike SVG, which they strip). Silently
    skips if the logo file can't be read so email sending never breaks.
    """
    from email.mime.image import MIMEImage
    from pathlib import Path

    for base in settings.STATICFILES_DIRS or []:
        for fname in ("logo_email.png", "logo.jpeg"):
            logo_path = Path(base) / "img" / fname
            if logo_path.exists():
                try:
                    img = MIMEImage(logo_path.read_bytes())
                    img.add_header("Content-ID", f"<{LOGO_CID}>")
                    img.add_header("Content-Disposition", "inline", filename=fname)
                    msg.attach(img)
                except Exception:
                    logger.warning("Could not attach email logo", exc_info=True)
                return

# Role that is allowed to view university-wide statistics (super-admin analytics).
ANALYTICS_ROLE = "University Timetable Coordinator"


def _viewer_credential():
    """Pick a super-admin credential to share with a University Timetable
    Coordinator so they can sign in and view university-wide statistics.

    Prefers a ``dean`` account if one exists in ``SUPERADMIN_USERS`` (a
    secondary analytics login), otherwise falls back to the first configured
    account. Returns ``(email, password)`` or ``None`` if none configured.
    """
    users = getattr(settings, "SUPERADMIN_USERS", {}) or {}
    if not users:
        return None
    for email, pwd in users.items():
        if "dean" in email.lower():
            return email, pwd
    email = next(iter(users))
    return email, users[email]


def _appoint_email(*, name, role, dept_name, message, appointer, creds=None, login_url=""):
    """Return (subject, text_body, html_body) for an appointment email.

    The HTML uses inline styles only (best compatibility across mail clients)
    and the new brand mark/gradient used on the app's loading screen.

    ``creds`` is an optional ``(email, password)`` tuple. When provided (for a
    University Timetable Coordinator) the super-admin analytics login details
    are included so the coordinator can view university-wide statistics.
    """
    safe_name = name or "Faculty Member"
    dept_line = f" for <strong>{dept_name}</strong>" if dept_name else ""
    note_html = ""
    if message:
        note_html = f"""
            <tr><td style="padding:18px 28px 0;">
              <div style="background:#0f1a2c;border:1px solid rgba(148,163,184,.18);
                          border-radius:12px;padding:16px 18px;color:#cbd5e1;font-size:14px;line-height:1.6;">
                <div style="color:#64748b;font-size:11px;letter-spacing:.08em;text-transform:uppercase;margin-bottom:8px;">
                  Message from the administrator</div>
                {message.replace(chr(10), '<br>')}
              </div>
            </td></tr>"""

    creds_html = ""
    creds_text = ""
    if creds:
        c_email, c_pwd = creds
        link_html = (
            f'<a href="{login_url}" style="color:#38bdf8;">{login_url}</a>'
            if login_url else "your SmartScheduler super-admin URL"
        )
        creds_html = f"""
            <tr><td style="padding:18px 28px 0;">
              <div style="background:linear-gradient(160deg,#10261c,#0c1f17);border:1px solid rgba(34,197,94,.32);
                          border-radius:12px;padding:18px 20px;">
                <div style="color:#86efac;font-size:11px;letter-spacing:.08em;text-transform:uppercase;margin-bottom:10px;">
                  Statistics dashboard access</div>
                <div style="color:#94a3b8;font-size:13px;line-height:1.6;margin-bottom:12px;">
                  As a University Timetable Coordinator you can view university-wide statistics.
                  Sign in to the analytics dashboard with the credentials below:
                </div>
                <table role="presentation" cellpadding="0" cellspacing="0" style="width:100%;">
                  <tr><td style="color:#64748b;font-size:12px;padding:3px 0;width:90px;">Login URL</td>
                      <td style="color:#e2e8f0;font-size:13px;padding:3px 0;">{link_html}</td></tr>
                  <tr><td style="color:#64748b;font-size:12px;padding:3px 0;">Email</td>
                      <td style="color:#e2e8f0;font-size:13px;padding:3px 0;font-family:monospace;">{c_email}</td></tr>
                  <tr><td style="color:#64748b;font-size:12px;padding:3px 0;">Password</td>
                      <td style="color:#e2e8f0;font-size:13px;padding:3px 0;font-family:monospace;">{c_pwd}</td></tr>
                </table>
                <div style="color:#64748b;font-size:11.5px;line-height:1.5;margin-top:12px;">
                  Please keep these credentials confidential and do not share them.
                </div>
              </div>
            </td></tr>"""
        creds_text = (
            "\nStatistics dashboard access (University Timetable Coordinator):\n"
            f"  Login URL : {login_url or 'your SmartScheduler super-admin URL'}\n"
            f"  Email     : {c_email}\n"
            f"  Password  : {c_pwd}\n"
            "  Please keep these credentials confidential.\n"
        )

    subject = f"You have been appointed as {role} — SmartScheduler"

    text_body = (
        f"Hello {safe_name},\n\n"
        "J.C. Bose University of Science and Technology, YMCA (Formerly YMCA UST)\n\n"
        "SmartScheduler by the SIH Winner Innovation Team\n\n"
        f"You have been appointed as {role}"
        + (f" for {dept_name}" if dept_name else "")
        + ".\n\n"
        + (f"Message from the administrator:\n{message}\n\n" if message else "")
        + "You can sign in to SmartScheduler to manage timetable activities for your role.\n"
        + creds_text
        + f"\nAppointed by: {appointer}\n"
        "— SmartScheduler\nSIH Winner Innovation Team"
    )

    html_body = f"""\
<!DOCTYPE html><html><body style="margin:0;background:#0b1220;
     font-family:'Segoe UI',system-ui,-apple-system,sans-serif;">
  <table role="presentation" width="100%" cellpadding="0" cellspacing="0"
         style="background:#0b1220;padding:28px 12px;">
    <tr><td align="center">
      <table role="presentation" width="560" cellpadding="0" cellspacing="0"
             style="max-width:560px;width:100%;background:#111c30;border:1px solid rgba(148,163,184,.16);
                    border-radius:18px;overflow:hidden;">
        <tr><td style="background:linear-gradient(135deg,#38bdf8,#0ea5e9 55%,#0c3557);padding:26px 28px;">
          <table role="presentation" cellpadding="0" cellspacing="0"><tr>
            <td style="vertical-align:middle;">
              <div style="width:46px;height:46px;border-radius:13px;background:rgba(8,20,36,.55);
                          text-align:center;line-height:46px;">
                <img src="cid:smartscheduler_logo" width="34" height="41" alt="SmartScheduler"
                     style="display:inline-block;width:34px;height:41px;vertical-align:middle;">
              </div>
            </td>
            <td style="padding-left:14px;vertical-align:middle;">
              <div style="color:#ffffff;font-size:19px;font-weight:700;">SmartScheduler</div>
              <div style="color:rgba(255,255,255,.82);font-size:12px;">University Timetable Platform</div>
                            <div style="color:#fff7ed;font-size:11px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;margin-top:6px;">SIH Winner Innovation Team</div>
            </td>
          </tr></table>
        </td></tr>
        <tr><td style="background:#ffffff;padding:14px 28px;border-bottom:1px solid rgba(148,163,184,.16);">
          <table role="presentation" cellpadding="0" cellspacing="0"><tr>
            <td style="vertical-align:middle;">
              <img src="https://upload.wikimedia.org/wikipedia/en/a/ae/J.C._Bose_University_of_Science_and_Technology%2C_YMCA_logo.png"
                   width="40" height="40" alt="J.C. Bose University"
                   style="display:block;width:40px;height:40px;object-fit:contain;">
            </td>
            <td style="padding-left:12px;vertical-align:middle;">
              <div style="color:#e11d2f;font-size:13.5px;font-weight:700;line-height:1.35;">
                J.C. Bose University of Science and Technology, YMCA (Formerly YMCA UST)</div>
            </td>
          </tr></table>
        </td></tr>
        <tr><td style="padding:26px 28px 4px;">
          <div style="color:#e2e8f0;font-size:17px;font-weight:600;margin-bottom:6px;">
            Hello {safe_name},</div>
          <div style="color:#94a3b8;font-size:14px;line-height:1.65;">
            You have been appointed as the role below on SmartScheduler{dept_line}.
          </div>
        </td></tr>
        <tr><td style="padding:18px 28px 0;">
          <div style="background:linear-gradient(160deg,#0f1a2c,#0d1626);border:1px solid rgba(56,189,248,.28);
                      border-radius:12px;padding:18px 20px;">
            <div style="color:#64748b;font-size:11px;letter-spacing:.08em;text-transform:uppercase;">Role</div>
            <div style="color:#38bdf8;font-size:20px;font-weight:700;margin-top:4px;">{role}</div>
            {f'<div style="color:#94a3b8;font-size:13px;margin-top:8px;">Department · {dept_name}</div>' if dept_name else ''}
          </div>
        </td></tr>
        {note_html}
        {creds_html}
        <tr><td style="padding:22px 28px 6px;">
          <div style="color:#94a3b8;font-size:13.5px;line-height:1.65;">
            Please sign in to SmartScheduler to begin managing timetable activities for your role.
          </div>
        </td></tr>
        <tr><td style="padding:8px 28px 26px;">
          <div style="border-top:1px solid rgba(148,163,184,.16);padding-top:16px;
                      color:#64748b;font-size:12px;line-height:1.6;">
            Appointed by <span style="color:#cbd5e1;">{appointer}</span><br>
                        This is an automated message from SmartScheduler.<br><span style="color:#f59e0b;font-weight:700;">SIH Winner Innovation Team</span>
          </div>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body></html>"""

    return subject, text_body, html_body


def _superadmin_appoint_send(request):
    name = (request.POST.get("name") or "").strip()
    email = (request.POST.get("email") or "").strip()
    role = (request.POST.get("role") or "").strip()
    dept_name = (request.POST.get("department") or "").strip()
    message = (request.POST.get("message") or "").strip()

    if not name or not email or role not in APPOINT_ROLES:
        messages.error(request, "Please provide the teacher's name, email and a valid role.")
        return redirect("superadmin_appoint")
    if not _EMAIL_RE.match(email):
        messages.error(request, "That email address doesn't look valid.")
        return redirect("superadmin_appoint")
    if not (settings.EMAIL_HOST_USER and settings.EMAIL_HOST_PASSWORD):
        messages.error(
            request,
            "Email is not configured. Add EMAIL_HOST_USER and EMAIL_HOST_PASSWORD in .env, then restart the server.",
        )
        return redirect("superadmin_appoint")

    appointer = request.session.get(SESSION_EMAIL, "SmartScheduler Super Admin")

    # Only a University Timetable Coordinator receives super-admin analytics
    # login details so they can view university-wide statistics.
    creds = None
    login_url = ""
    if role == ANALYTICS_ROLE:
        creds = _viewer_credential()
        login_url = request.build_absolute_uri(reverse("superadmin_login"))

    subject, text_body, html_body = _appoint_email(
        name=name, role=role, dept_name=dept_name, message=message,
        appointer=appointer, creds=creds, login_url=login_url,
    )
    sender = _brand_from_email()
    email_ok = False
    try:
        from email.utils import formataddr
        msg = EmailMultiAlternatives(
            subject=subject,
            body=text_body,
            from_email=sender,
            to=[formataddr((name, email))],
            reply_to=[appointer],
        )
        msg.attach_alternative(html_body, "text/html")
        _attach_logo(msg)
        msg.send(fail_silently=False)
        email_ok = True
        messages.success(request, f"Appointment email sent to {name} ({email}).")
    except Exception as exc:
        logger.exception("Appointment email failed")
        messages.error(request, f"Couldn't send the email: {exc}")

    # Persist the appointment so it can be tracked and listed role-wise.
    # An (email, role) pair is unique; re-appointing updates the record.
    try:
        CoordinatorAppointment.objects.update_or_create(
            email=email.lower(),
            role=role,
            defaults={
                "name": name,
                "department": dept_name,
                "message": message,
                "appointed_by": appointer,
                "analytics_access": role == ANALYTICS_ROLE,
                "email_sent": email_ok,
            },
        )
    except Exception:
        logger.exception("Could not save coordinator appointment")

    return redirect("superadmin_appoint")


@superadmin_required
def superadmin_appoint(request):
    if request.method == "POST":
        return _superadmin_appoint_send(request)

    ctx = _page_ctx(request, "appoint")
    owners = _active_owner_ids()
    rows = list(
        _scope_to_owners(Instructor.objects, owners)
        .order_by("name")
        .values("id", "uid", "name", "email", "designation")
    )
    # De-duplicate master data that may be repeated across owner accounts.
    seen, teachers = set(), []
    for t in rows:
        key = (t["name"].strip().lower(), (t["email"] or "").strip().lower())
        if key in seen:
            continue
        seen.add(key)
        teachers.append(t)
    ctx["appoint_teachers"] = teachers
    ctx["roles"] = APPOINT_ROLES

    # Existing appointments, grouped role-wise for the table below the form.
    appts = list(CoordinatorAppointment.objects.all().order_by("role", "name"))
    grouped = []
    for role_name in APPOINT_ROLES:
        members = [a for a in appts if a.role == role_name]
        grouped.append({"role": role_name, "count": len(members), "members": members})
    ctx["appointment_groups"] = grouped
    ctx["appointment_total"] = len(appts)
    return render(request, "superadmin_appoint.html", ctx)


@superadmin_required
def superadmin_appoint_delete(request, aid):
    """Confirmation page (GET) + remove an appointed coordinator (POST)."""
    try:
        appt = CoordinatorAppointment.objects.get(id=aid)
    except CoordinatorAppointment.DoesNotExist:
        raise Http404("Appointment not found")

    if request.method == "POST":
        label = appt.name
        appt.delete()
        messages.success(request, f"Removed {label} from the {appt.role} role.")
        return redirect("superadmin_appoint")

    ctx = _page_ctx(request, "appoint")
    ctx["target_appointment"] = {
        "id": appt.id,
        "name": appt.name,
        "email": appt.email,
        "role": appt.role,
        "department": appt.department or "University-wide",
        "analytics_access": appt.analytics_access,
        "created": timezone.localtime(appt.created_at).strftime("%d %b %Y, %H:%M")
        if appt.created_at
        else "—",
    }
    return render(request, "superadmin_appoint_confirm_delete.html", ctx)


@superadmin_required
def superadmin_drilldown(request):
    """University -> Department -> Section -> Teacher -> Subject drill-down (JSON)."""
    level = (request.GET.get("level") or "university").strip()
    key = (request.GET.get("id") or "").strip()
    owners, _account_filter = _request_owners(request)
    slots = _active_slots(owners)

    def _row(slot):
        return {
            "section": slot.section.section_id if slot.section else "—",
            "subject": getattr(slot.subject, "subject_name", "—"),
            "teacher": getattr(slot.instructor, "name", "—"),
            "room": getattr(slot.room, "r_number", "—"),
            "day": slot.meeting_time.day,
            "slot": str(slot.meeting_time.time),
            "type": "Lab" if slot.is_lab else "Theory",
        }

    if level == "university":
        agg = defaultdict(lambda: {"sections": set(), "teachers": set(), "classes": 0})
        for s in slots:
            dept = s.section.department if s.section else None
            if not dept:
                continue
            a = agg[(dept.id, dept.name, dept.code)]
            a["sections"].add(s.section.section_id)
            if s.instructor_id:
                a["teachers"].add(s.instructor_id)
            a["classes"] += 1
        data = [
            {
                "id": k[0],
                "name": k[1],
                "code": k[2],
                "sections": len(v["sections"]),
                "teachers": len(v["teachers"]),
                "classes": v["classes"],
            }
            for k, v in sorted(agg.items(), key=lambda x: x[0][1])
        ]
        return JsonResponse({"level": "university", "items": data})

    if level == "department":
        agg = defaultdict(lambda: {"teachers": set(), "classes": 0, "sem": ""})
        for s in slots:
            if not s.section or str(s.section.department_id) != key:
                continue
            a = agg[(s.section.id, s.section.section_id)]
            a["sem"] = _semester_of(s.section.section_id)
            if s.instructor_id:
                a["teachers"].add(s.instructor_id)
            a["classes"] += 1
        data = [
            {
                "id": k[0],
                "name": k[1],
                "sem": v["sem"],
                "teachers": len(v["teachers"]),
                "classes": v["classes"],
            }
            for k, v in sorted(agg.items(), key=lambda x: x[0][1])
        ]
        return JsonResponse({"level": "department", "items": data})

    if level == "section":
        rows = [_row(s) for s in slots if s.section and str(s.section_id) == key]
        rows.sort(key=lambda r: (_DAY_ORDER.get(r["day"], 99), r["slot"]))
        return JsonResponse({"level": "section", "items": rows})

    if level == "teacher":
        rows = [_row(s) for s in slots if str(s.instructor_id) == key]
        rows.sort(key=lambda r: (_DAY_ORDER.get(r["day"], 99), r["slot"]))
        return JsonResponse({"level": "teacher", "items": rows})

    if level == "subject":
        rows = [_row(s) for s in slots if str(s.subject_id) == key]
        rows.sort(key=lambda r: (r["section"], _DAY_ORDER.get(r["day"], 99), r["slot"]))
        return JsonResponse({"level": "subject", "items": rows})

    return JsonResponse({"level": level, "items": []})


@superadmin_required
def superadmin_export_excel(request):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except Exception:
        messages.error(request, "Excel export library is not available.")
        return redirect("superadmin_dashboard")

    ctx = _build_analytics(request)
    wb = Workbook()

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2563EB")

    def _style_header(ws):
        for cell in ws[1]:
            cell.font = head_font
            cell.fill = head_fill

    ws = wb.active
    ws.title = "KPIs"
    ws.append(["Metric", "Value"])
    for label, value in [
        ("Departments Active", ctx["kpis"]["departments_active"]),
        ("Teachers Active", ctx["kpis"]["teachers_active"]),
        ("Sections Scheduled", ctx["kpis"]["sections_scheduled"]),
        ("Rooms Total", ctx["kpis"]["rooms_total"]),
        ("Labs Total", ctx["kpis"]["labs_total"]),
        ("Classes Scheduled", ctx["kpis"]["classes_scheduled"]),
        ("Overall Utilization %", ctx["kpis"]["overall_util"]),
        ("Conflicts Detected", ctx["kpis"]["conflicts"]),
    ]:
        ws.append([label, value])
    _style_header(ws)

    ws2 = wb.create_sheet("Department Utilization")
    ws2.append(["Department", "Code", "Utilization %", "Used Cells", "Rooms"])
    for d in ctx["dept_util"]:
        ws2.append([d["name"], d["code"], d["util"], d["used"], d["rooms"]])
    _style_header(ws2)

    ws3 = wb.create_sheet("Room Utilization")
    ws3.append(["Room", "Type", "Department", "Used", "Capacity", "Utilization %"])
    for r in ctx["room_util"]:
        ws3.append([r["number"], r["type"], r["dept"], r["used"], r["capacity"], r["util"]])
    _style_header(ws3)

    ws4 = wb.create_sheet("Teacher Workload")
    ws4.append(["Teacher", "Designation", "Load (hrs)", "Max", "Utilization %"])
    for t in ctx["teacher_workload"]:
        ws4.append([t["name"], t["designation"], t["load"], t["max"], t["util"]])
    _style_header(ws4)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    resp = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="superadmin_analytics_{stamp}.xlsx"'
    return resp


@superadmin_required
def superadmin_export_pdf(request):
    ctx = _build_analytics(request)
    ctx["generated_at"] = timezone.localtime(timezone.now()).strftime("%d %b %Y, %H:%M")
    html = render(request, "superadmin_report.html", ctx).content.decode("utf-8")
    try:
        from xhtml2pdf import pisa
    except Exception:
        messages.error(request, "PDF export library is not available.")
        return redirect("superadmin_dashboard")

    buffer = io.BytesIO()
    result = pisa.CreatePDF(io.StringIO(html), dest=buffer)
    if result.err:
        messages.error(request, "Could not render the PDF report.")
        return redirect("superadmin_dashboard")
    buffer.seek(0)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    resp = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    resp["Content-Disposition"] = f'attachment; filename="superadmin_report_{stamp}.pdf"'
    return resp


# ---------------------------------------------------------------------------
# Resource Utilization — room analytics by utilization bucket (real data)
# ---------------------------------------------------------------------------
def _sorted_times(universe):
    times = sorted({t for _, t in universe}, key=lambda x: int(x) if x.isdigit() else 99)
    days = sorted({d for d, _ in universe}, key=lambda d: _DAY_ORDER.get(d, 99))
    return days, times


def _room_occupancy(owners=None):
    """room_id -> {(day,time): {section, subject, teacher, is_lab}} from active slots."""
    occ = defaultdict(dict)
    for slot in _active_slots(owners):
        room = slot.room
        if not room:
            continue
        cells = [(slot.meeting_time.day, str(slot.meeting_time.time))]
        if slot.is_lab:
            cells += [(mt.day, str(mt.time)) for mt in slot.lab_slots.all()]
        info = {
            "section": slot.section.section_id if slot.section else "—",
            "subject": getattr(slot.subject, "subject_name", "—"),
            "teacher": getattr(slot.instructor, "name", "—"),
            "is_lab": slot.is_lab,
        }
        for day, time in cells:
            if day:
                occ[room.id][(day, time)] = info
    return occ


def _bucket_of(util):
    if util > 70:
        return "gt70"
    if util >= 50:
        return "50-70"
    if util >= 20:
        return "20-50"
    return "ideal"


_BUCKET_LABELS = {
    "gt70": "Above 70% (High load)",
    "50-70": "50% – 70% (Healthy)",
    "20-50": "20% – 50% (Moderate)",
    "ideal": "Ideal (< 20%, plenty free)",
    "all": "All Rooms",
}


@superadmin_required
def superadmin_room_analytics(request):
    """Detailed room analytics filtered by utilization bucket (real DB data)."""
    bucket = (request.GET.get("bucket") or "all").strip()
    dept_filter = (request.GET.get("dept") or "all").strip()
    room_id = (request.GET.get("room_id") or "").strip()

    universe = _slot_universe()
    base_room_capacity = sum(1 for cell in universe if not _is_lunch_cell(cell)) or 1
    days, times = _sorted_times(universe)
    owners, _account_filter = _request_owners(request)
    occ = _room_occupancy(owners)

    rooms_qs = _scope_to_owners(
        Room.objects.select_related("department"), owners
    ).order_by("r_number")
    if dept_filter != "all":
        rooms_qs = rooms_qs.filter(department_id=dept_filter)
    if room_id.isdigit():
        rooms_qs = rooms_qs.filter(id=int(room_id))

    rooms = []
    bucket_counts = {"gt70": 0, "50-70": 0, "20-50": 0, "ideal": 0}
    for room in rooms_qs:
        cells = occ.get(room.id, {})
        used = len(cells)
        capacity = _effective_room_capacity(universe, set(cells.keys())) or base_room_capacity
        util = _pct(used, capacity)
        b = _bucket_of(util)
        bucket_counts[b] += 1
        if bucket != "all" and b != bucket:
            continue

        free_slots = [
            {"day": d, "slot": t} for (d, t) in sorted(
                {cell for cell in universe if not _is_lunch_cell(cell)} - set(cells.keys()), key=lambda c: (_DAY_ORDER.get(c[0], 99), int(c[1]) if c[1].isdigit() else 99)
            )
        ]
        grid = []
        for d in days:
            row = {"day": d, "cells": []}
            for t in times:
                cell = cells.get((d, t))
                row["cells"].append(
                    {
                        "free": cell is None,
                        "section": cell["section"] if cell else "",
                        "subject": cell["subject"] if cell else "",
                        "is_lab": cell["is_lab"] if cell else False,
                    }
                )
            grid.append(row)

        rooms.append(
            {
                "id": room.id,
                "number": room.r_number,
                "type": room.room_type,
                "lab_category": room.lab_category or "—",
                "department": room.department.name if room.department_id else "—",
                "capacity": room.seating_capacity,
                "used": used,
                "total": capacity,
                "util": util,
                "free_count": max(capacity - used, 0),
                "free_slots": free_slots,
                "grid": grid,
                "bucket": b,
            }
        )

    rooms.sort(key=lambda r: r["util"], reverse=True)
    class_count = sum(r["used"] for r in rooms)
    avg_util = round(sum(r["util"] for r in rooms) / len(rooms), 1) if rooms else 0.0

    return JsonResponse(
        {
            "bucket": bucket,
            "bucket_label": _BUCKET_LABELS.get(bucket, "Rooms"),
            "single_room": room_id.isdigit(),
            "bucket_counts": bucket_counts,
            "days": days,
            "times": times,
            "summary": {
                "rooms": len(rooms),
                "classes": class_count,
                "avg_util": avg_util,
            },
            "rooms": rooms,
        }
    )


# ---------------------------------------------------------------------------
# Teacher analytics — full profile + schedule (real data)
# ---------------------------------------------------------------------------
@superadmin_required
def superadmin_teacher_detail(request):
    """Complete profile + day/time schedule + workload stats for one teacher."""
    tid = (request.GET.get("id") or "").strip()
    try:
        instr = Instructor.objects.get(id=tid)
    except (Instructor.DoesNotExist, ValueError):
        raise Http404("Teacher not found")

    universe = _slot_universe()
    days, times = _sorted_times(universe)

    schedule_map = {}
    grid_map = {}
    departments = set()
    subjects = set()
    sections = set()
    lab_count = 0
    theory_count = 0
    per_day = defaultdict(int)

    owners, _account_filter = _request_owners(request)
    teacher_slots = _active_slots(owners).filter(
        Q(instructor_id=instr.id) | Q(second_instructor_id=instr.id)
    )
    for slot in teacher_slots:
        sec = slot.section
        if sec and sec.department_id:
            departments.add(sec.department.name)
        subjects.add(getattr(slot.subject, "subject_name", "—"))
        if sec:
            sections.add(sec.section_id)

        cells = _slot_cells(slot)
        if slot.is_lab:
            lab_count += 1
        else:
            theory_count += 1

        for day, time in cells:
            entry = schedule_map.setdefault(
                (day, time),
                {
                    "day": day,
                    "slot": time,
                    "sections": set(),
                    "subjects": set(),
                    "rooms": set(),
                    "types": set(),
                },
            )
            entry["sections"].add(sec.section_id if sec else "—")
            entry["subjects"].add(getattr(slot.subject, "subject_name", "—"))
            entry["rooms"].add(getattr(slot.room, "r_number", "—"))
            entry["types"].add("Lab" if slot.is_lab else "Theory")

    schedule = []
    for (day, time), entry in schedule_map.items():
        sections_list = sorted(entry["sections"])
        subjects_list = sorted(entry["subjects"])
        rooms_list = sorted(entry["rooms"])
        types_list = sorted(entry["types"])
        merged = {
            "day": day,
            "slot": time,
            "section": ", ".join(sections_list),
            "subject": ", ".join(subjects_list),
            "room": ", ".join(rooms_list),
            "type": "Lab" if types_list == ["Lab"] else "Theory" if types_list == ["Theory"] else "Mixed",
        }
        schedule.append(merged)
        grid_map[(day, time)] = merged
        per_day[day] += 1

    schedule.sort(key=lambda e: (_DAY_ORDER.get(e["day"], 99), int(e["slot"]) if e["slot"].isdigit() else 99))
    load = len(schedule)
    grid = [
        {
            "day": d,
            "cells": [grid_map.get((d, t)) for t in times],
        }
        for d in days
    ]
    busiest_day = max(per_day.items(), key=lambda x: x[1])[0] if per_day else "—"

    # Department from sections taught (fallback to TeacherSection mappings).
    if not departments:
        for ts in TeacherSection.objects.filter(instructor=instr).select_related("section__department"):
            if ts.section and ts.section.department_id:
                departments.add(ts.section.department.name)

    return JsonResponse(
        {
            "profile": {
                "id": instr.id,
                "uid": instr.uid,
                "name": instr.name,
                "designation": instr.designation,
                "email": instr.email,
                "contact": instr.contact_number,
                "department": ", ".join(sorted(departments)) or "—",
            },
            "stats": {
                "load": load,
                "max": instr.max_workload,
                "util": _pct(load, instr.max_workload),
                "sections": len(sections),
                "subjects": len(subjects),
                "labs": lab_count,
                "theory": theory_count,
                "busiest_day": busiest_day,
            },
            "subjects": sorted(subjects),
            "days": days,
            "times": times,
            "grid": grid,
            "schedule": schedule,
            "per_day": [{"day": d, "count": per_day.get(d, 0)} for d in days],
        }
    )


# ---------------------------------------------------------------------------
# Teacher workload analysis — selected user's active timetable breakdown
# ---------------------------------------------------------------------------
def _saved_timetable_label(st):
    if st.user:
        owner = (st.user.email or st.user.username or f"User {st.user_id}").strip()
    else:
        owner = f"User {st.user_id}"
    dept = st.department.name if st.department else "All Departments"
    when = st.created_at.strftime("%d %b %Y") if st.created_at else ""
    label = f"{owner} · {dept}"
    if when:
        label += f" · {when}"
    return owner, label


def _norm_identity(value):
    return re.sub(r"\s+", " ", (value or "").strip().lower())


@superadmin_required
def superadmin_teacher_workload(request):
    """Detailed workload breakdown for one teacher in the selected user's
    active saved timetable set.

    Teacher identity is matched by uid/name inside the session-selected owner
    so duplicate teacher rows across different accounts never leak in.
    """
    raw_id = (request.GET.get("id") or "").strip()
    try:
        clicked = Instructor.objects.get(id=raw_id)
    except (Instructor.DoesNotExist, ValueError):
        raise Http404("Teacher not found")

    target_uid = _norm_identity(clicked.uid)
    target_name = _norm_identity(clicked.name)
    owners, _account_filter = _request_owners(request)
    selected_owner = _selected_owner_account(request)

    response = {
        "teacher": {
            "id": clicked.id,
            "name": clicked.name,
            "uid": clicked.uid,
            "designation": clicked.designation,
        },
        "analysis": None,
    }

    active_ids = _active_timetable_ids(owners)
    if not active_ids:
        response["analysis"] = {"found": False, "departments": []}
        return JsonResponse(response)

    # Instructors in the selected owner scope that match the clicked teacher.
    matched_ids = []
    matched_instructors = []
    for instr in Instructor.objects.filter(user_id__in=owners).select_related("department"):
        if (_norm_identity(instr.uid) and _norm_identity(instr.uid) == target_uid) or (
            _norm_identity(instr.name) and _norm_identity(instr.name) == target_name
        ):
            matched_ids.append(instr.id)
            matched_instructors.append(instr)

    if not matched_ids:
        response["analysis"] = {"found": False, "departments": []}
        return JsonResponse(response)

    slots = (
        ScheduledSlot.objects.filter(timetable_id__in=active_ids)
        .select_related(
            "section", "section__department", "subject", "instructor", "second_instructor"
        )
        .prefetch_related("lab_slots")
    )

    # Home department: stored Instructor.department first, else busiest dept.
    home_code = ""
    home_name = ""
    for instr in matched_instructors:
        if instr.department_id:
            home_code = (instr.department.code or "").strip()
            home_name = (instr.department.name or "").strip() or home_code
            break

    # dept_code -> aggregate, plus subject-level breakdown rows.
    dept_map = {}
    total_cells = set()
    lecture_cells = set()
    lab_cells = set()
    home_cells = set()
    cross_cells = set()
    detail_map = {}

    for slot in slots:
        if slot.instructor_id not in matched_ids and slot.second_instructor_id not in matched_ids:
            continue
        sec = slot.section
        dept = sec.department if sec else None
        code = (getattr(dept, "code", "") or "").strip() or "—"
        name = (getattr(dept, "name", "") or "").strip() or code
        cells = _slot_cells(slot)
        duration = len(cells)
        if duration < 1:
            duration = 1
        section_name = sec.section_id if sec else "—"
        subject_name = getattr(slot.subject, "subject_name", "") or "—"
        block_label = _format_cell_block(cells)
        is_home = bool(home_code) and code.lower() == home_code.lower()
        kind = "Lab" if slot.is_lab else "Lecture"

        bucket = dept_map.setdefault(
            code,
            {
                "code": code,
                "name": name,
                "cells": set(),
                "lecture_cells": set(),
                "lab_cells": set(),
                "subjects": set(),
            },
        )
        bucket["cells"].update(cells)
        if slot.is_lab:
            bucket["lab_cells"].update(cells)
            lab_cells.update(cells)
        else:
            bucket["lecture_cells"].update(cells)
            lecture_cells.update(cells)
        bucket["subjects"].add(subject_name)
        total_cells.update(cells)
        if is_home:
            home_cells.update(cells)
        else:
            cross_cells.update(cells)

        detail_key = (code, name, subject_name, section_name, kind, duration)
        detail = detail_map.setdefault(
            detail_key,
            {
                "department_code": code,
                "department_name": name,
                "subject": subject_name,
                "section": section_name,
                "type": kind,
                "duration": duration,
                "weekly_hours": 0,
                "occurrences": 0,
                "blocks": [],
                "is_home": is_home,
            },
        )
        detail["weekly_hours"] += len(cells)
        detail["occurrences"] += 1
        detail["blocks"].append(block_label)

    if not home_code and dept_map:
        top = max(dept_map.values(), key=lambda d: len(d["cells"]))
        home_code = top["code"]
        home_name = top["name"]

    departments = []
    for bucket in dept_map.values():
        departments.append(
            {
                "code": bucket["code"],
                "name": bucket["name"],
                "hours": len(bucket["cells"]),
                "lecture_hours": len(bucket["lecture_cells"]),
                "lab_hours": len(bucket["lab_cells"]),
                "subjects": sorted(bucket["subjects"]),
                "is_home": bool(home_code) and bucket["code"].lower() == home_code.lower(),
            }
        )
    # Home department first, then by hours desc.
    departments.sort(key=lambda d: (not d["is_home"], -d["hours"]))

    details = list(detail_map.values())
    for item in details:
        item["blocks"] = sorted(item["blocks"], key=lambda value: (_DAY_ORDER.get(value.split()[0], 99), value))
    details.sort(key=lambda item: (not item["is_home"], item["department_name"], item["subject"], item["section"], item["type"]))

    source_label = "Selected user active saved timetable"
    if len(active_ids) > 1:
        source_label = f"Selected user active saved timetables ({len(active_ids)})"
    if selected_owner:
        source_label += f" · {(selected_owner.get('label') or selected_owner.get('username') or selected_owner.get('email') or '').strip()}"

    response["analysis"] = {
        "source_label": source_label,
        "home_department": home_name or "—",
        "home_code": home_code,
        "total_hours": len(total_cells),
        "lecture_hours": len(lecture_cells),
        "lab_hours": len(lab_cells),
        "home_hours": len(home_cells),
        "cross_hours": len(cross_cells),
        "home_lecture_hours": sum(d["lecture_hours"] for d in departments if d["is_home"]),
        "home_lab_hours": sum(d["lab_hours"] for d in departments if d["is_home"]),
        "cross_departments": sum(1 for d in departments if not d["is_home"] and d["hours"]),
        "max_workload": matched_instructors[0].max_workload if matched_instructors else clicked.max_workload,
        "departments": departments,
        "details": details,
        "found": bool(matched_ids),
    }
    return JsonResponse(response)


# ---------------------------------------------------------------------------
# Saved timetable management — list / detail / drag-drop edit (real data)
# ---------------------------------------------------------------------------
@superadmin_required
def superadmin_saved_list(request):
    """Every saved timetable across all users (admin oversight)."""
    items = []
    owners, _account_filter = _request_owners(request)
    qs = (
        SavedTimetable.objects.select_related("department", "user")
        .filter(user_id__in=owners)
        .order_by("-created_at")
    )
    for st in qs:
        slots = st.slots.all()
        section_count = len({s.section_id for s in slots})
        items.append(
            {
                "id": st.id,
                "label": st.department.name if st.department_id else "All Departments",
                "owner": getattr(st.user, "username", "—") if st.user_id else "—",
                "slots": len(slots),
                "sections": section_count,
                "published": st.is_published,
                "created": timezone.localtime(st.created_at).strftime("%d %b %Y, %H:%M")
                if st.created_at
                else "",
            }
        )
    return JsonResponse({"items": items})


# ---------------------------------------------------------------------------
# User management — list accounts, delete a user (with confirmation) and all
# of their data. Deleting a user cascades to every owner-scoped table
# (departments, teachers, rooms, sections, subjects, saved timetables, …)
# via the FK on_delete=CASCADE already defined on those models. This never
# touches the scheduling / generation algorithm.
# ---------------------------------------------------------------------------
def _user_data_summary(user):
    """Counts of the records that would be removed with this user."""
    return {
        "departments": Department.objects.filter(user=user).count(),
        "teachers": Instructor.objects.filter(user=user).count(),
        "rooms": Room.objects.filter(user=user).count(),
        "sections": Section.objects.filter(user=user).count(),
        "timetables": SavedTimetable.objects.filter(user=user).count(),
    }


@superadmin_required
def superadmin_users(request):
    """List every registered account with a snapshot of their data."""
    User = get_user_model()
    q = (request.GET.get("q") or "").strip()

    users_qs = User.objects.all().order_by("-date_joined")
    if q:
        users_qs = users_qs.filter(Q(username__icontains=q) | Q(email__icontains=q))

    dept_counts = dict(Department.objects.values_list("user_id").annotate(c=Count("id")))
    teacher_counts = dict(Instructor.objects.values_list("user_id").annotate(c=Count("id")))
    room_counts = dict(Room.objects.values_list("user_id").annotate(c=Count("id")))
    section_counts = dict(Section.objects.values_list("user_id").annotate(c=Count("id")))
    tt_counts = dict(SavedTimetable.objects.values_list("user_id").annotate(c=Count("id")))

    rows = []
    for u in users_qs:
        rows.append(
            {
                "id": u.id,
                "username": u.get_username(),
                "email": getattr(u, "email", "") or "—",
                "is_staff": u.is_staff,
                "is_superuser": u.is_superuser,
                "joined": timezone.localtime(u.date_joined).strftime("%d %b %Y")
                if getattr(u, "date_joined", None)
                else "—",
                "departments": dept_counts.get(u.id, 0),
                "teachers": teacher_counts.get(u.id, 0),
                "rooms": room_counts.get(u.id, 0),
                "sections": section_counts.get(u.id, 0),
                "timetables": tt_counts.get(u.id, 0),
            }
        )

    ctx = _basic_page_ctx(request, "users")
    ctx["user_rows"] = rows
    ctx["user_total"] = len(rows)
    ctx["user_query"] = q
    return render(request, "superadmin_users.html", ctx)


@superadmin_required
def superadmin_user_delete(request, uid):
    """Confirmation page (GET) + hard delete of a user and all data (POST)."""
    User = get_user_model()
    try:
        target = User.objects.get(id=uid)
    except User.DoesNotExist:
        raise Http404("User not found")

    # Never allow deleting a superuser through this panel (safety guard).
    if target.is_superuser:
        messages.error(request, "Superuser accounts cannot be deleted from here.")
        return redirect("superadmin_users")

    if request.method == "POST":
        label = target.get_username()
        with transaction.atomic():
            target.delete()
        messages.success(request, f"Deleted user “{label}” and all of their data.")
        return redirect("superadmin_users")

    ctx = _basic_page_ctx(request, "users")
    ctx["target_user"] = {
        "id": target.id,
        "username": target.get_username(),
        "email": getattr(target, "email", "") or "—",
        "joined": timezone.localtime(target.date_joined).strftime("%d %b %Y, %H:%M")
        if getattr(target, "date_joined", None)
        else "—",
    }
    ctx["delete_summary"] = _user_data_summary(target)
    return render(request, "superadmin_user_confirm_delete.html", ctx)


# ---------------------------------------------------------------------------
# Saved timetable deletion — confirmation page (GET) + hard delete (POST).
# Removes the SavedTimetable and its scheduled slots only; master data and
# the generation algorithm are untouched.
# ---------------------------------------------------------------------------
@superadmin_required
def superadmin_saved_delete(request, tid):
    try:
        st = SavedTimetable.objects.select_related("department", "user").get(id=tid)
    except SavedTimetable.DoesNotExist:
        raise Http404("Timetable not found")
    _selected_owner_guard_or_404(request, st.user_id)

    if request.method == "POST":
        label = st.department.name if st.department_id else "All Departments"
        with transaction.atomic():
            st.delete()
        messages.success(request, f"Deleted saved timetable ({label}).")
        return redirect("superadmin_saved_page")

    ctx = _page_ctx(request, "saved")
    ctx["target_timetable"] = {
        "id": st.id,
        "label": st.department.name if st.department_id else "All Departments",
        "owner": getattr(st.user, "username", "—") if st.user_id else "—",
        "published": st.is_published,
        "slots": st.slots.count(),
        "sections": len({s.section_id for s in st.slots.all()}),
        "created": timezone.localtime(st.created_at).strftime("%d %b %Y, %H:%M")
        if st.created_at
        else "—",
    }
    return render(request, "superadmin_saved_confirm_delete.html", ctx)


@superadmin_required
def superadmin_saved_detail(request, tid):
    """Section-wise editable grid for one saved timetable."""
    try:
        st = SavedTimetable.objects.get(id=tid)
    except SavedTimetable.DoesNotExist:
        raise Http404("Timetable not found")
    _selected_owner_guard_or_404(request, st.user_id)

    universe = _slot_universe()
    days, times = _sorted_times(universe)

    slots = (
        st.slots.select_related("section", "subject", "instructor", "room", "meeting_time")
        .prefetch_related("lab_slots")
        .order_by("section__section_id", "meeting_time__day", "meeting_time__time")
    )

    sections = {}
    for slot in slots:
        sec = slot.section
        if not sec:
            continue
        key = sec.section_id
        if key not in sections:
            sections[key] = {
                "section": key,
                "department": sec.department.name if sec.department_id else "—",
                "cells": {},
            }
        cell = {
            "slot_id": slot.id,
            "subject": getattr(slot.subject, "subject_name", "—"),
            "teacher": getattr(slot.instructor, "name", "—"),
            "room": getattr(slot.room, "r_number", "—"),
            "is_lab": slot.is_lab,
            "day": slot.meeting_time.day,
            "time": str(slot.meeting_time.time),
        }
        sections[key]["cells"][f"{slot.meeting_time.day}|{slot.meeting_time.time}"] = cell

    section_list = []
    for data in sections.values():
        grid = []
        for d in days:
            row = {"day": d, "cells": []}
            for t in times:
                row["cells"].append(data["cells"].get(f"{d}|{t}"))
            grid.append(row)
        section_list.append(
            {"section": data["section"], "department": data["department"], "grid": grid}
        )
    section_list.sort(key=lambda s: s["section"])

    return JsonResponse(
        {
            "id": st.id,
            "label": st.department.name if st.department_id else "All Departments",
            "published": st.is_published,
            "days": days,
            "times": times,
            "sections": section_list,
        }
    )


@superadmin_required
@require_POST
def superadmin_move_slot(request):
    """Drag-drop edit: move a saved ScheduledSlot to a new day/time cell.

    This edits *saved* timetable data only — it does NOT touch the
    scheduling/generation algorithm.
    """
    slot_id = (request.POST.get("slot_id") or "").strip()
    new_day = (request.POST.get("day") or "").strip()
    new_time = (request.POST.get("time") or "").strip()

    try:
        slot = ScheduledSlot.objects.select_related("timetable", "section").get(id=slot_id)
    except (ScheduledSlot.DoesNotExist, ValueError):
        return JsonResponse({"ok": False, "message": "Slot not found."}, status=404)
    _selected_owner_guard_or_404(request, getattr(slot.timetable, "user_id", None))

    owner_id = getattr(slot.timetable, "user_id", None)
    target = MeetingTime.objects.filter(day=new_day, time=new_time)
    if owner_id:
        target = target.filter(user_id=owner_id)
    target_mt = target.first()
    if not target_mt:
        return JsonResponse(
            {"ok": False, "message": f"No time slot configured for {new_day} slot {new_time}."},
            status=400,
        )

    # Prevent double-booking the same section cell (theory).
    clash = (
        ScheduledSlot.objects.filter(
            timetable=slot.timetable,
            section=slot.section,
            meeting_time=target_mt,
            is_lab=False,
        )
        .exclude(id=slot.id)
        .exists()
    )
    if clash and not slot.is_lab:
        return JsonResponse(
            {"ok": False, "message": "That section already has a class in this slot."},
            status=409,
        )

    with transaction.atomic():
        slot.meeting_time = target_mt
        slot.save(update_fields=["meeting_time"])
        if slot.is_lab:
            slot.lab_slots.set([target_mt])

    return JsonResponse(
        {
            "ok": True,
            "message": "Slot moved.",
            "slot_id": slot.id,
            "day": new_day,
            "time": new_time,
        }
    )


# ---------------------------------------------------------------------------
# Open a saved timetable in the EXISTING full-featured page (with full edit).
#
# The existing coordinator page (``saved_timetable`` view + saved_timetable.html)
# and all its edit endpoints are owner-scoped (``request.user``). To let a Super
# Admin use that same page with full editing, we temporarily impersonate the
# timetable's owner for the duration via :class:`SuperAdminImpersonationMiddleware`.
# Impersonation only ever applies to non-``/superadmin/`` paths, so the Super
# Admin's own analytics pages always run as themselves.
# ---------------------------------------------------------------------------
@superadmin_required
def superadmin_open_saved(request, tid):
    try:
        st = SavedTimetable.objects.get(id=tid)
    except SavedTimetable.DoesNotExist:
        raise Http404("Timetable not found")
    _selected_owner_guard_or_404(request, st.user_id)
    request.session[SA_IMPERSONATE] = st.user_id
    request.session.modified = True
    return redirect("saved_timetable", tid=tid)


@superadmin_required
def superadmin_stop_impersonate(request):
    request.session.pop(SA_IMPERSONATE, None)
    request.session.modified = True
    return redirect("superadmin_saved_page")


class SuperAdminImpersonationMiddleware:
    """When a Super Admin opens a saved timetable, run owner-scoped pages as
    that owner so the existing full-featured editor works end-to-end.

    Active only when:
      * the session is an authenticated Super Admin, AND
      * an impersonation target (owner uid) is stored in the session, AND
      * the request path is NOT under ``/superadmin/``.

    A small floating banner is injected so the admin can exit back to the
    Super Admin area at any time. This never touches the scheduling algorithm.
    """

    EXIT_PATH = "/superadmin/stop-impersonate/"

    def __init__(self, get_response):
        self.get_response = get_response

    def __call__(self, request):
        impersonated_user = None
        if (
            request.session.get(SESSION_FLAG)
            and request.session.get(SA_IMPERSONATE)
            and not request.path.startswith("/superadmin/")
        ):
            from django.contrib.auth import get_user_model

            User = get_user_model()
            try:
                impersonated_user = User.objects.get(id=request.session[SA_IMPERSONATE])
                request.user = impersonated_user
            except User.DoesNotExist:
                request.session.pop(SA_IMPERSONATE, None)
                request.session.modified = True

        response = self.get_response(request)

        if request.path.startswith("/superadmin/") or impersonated_user is not None:
            _disable_response_cache(response)

        if impersonated_user is not None:
            self._inject_banner(response, impersonated_user)
        return response

    def _inject_banner(self, response, user):
        try:
            content_type = response.get("Content-Type", "")
            if "text/html" not in content_type or not getattr(response, "content", None):
                return
            html = response.content.decode("utf-8")
            if "</body>" not in html:
                return
            label = getattr(user, "username", "") or getattr(user, "email", "") or "owner"
            banner = (
                '<div style="position:fixed;left:50%;bottom:18px;transform:translateX(-50%);'
                "z-index:99999;background:#0f1a2c;border:1px solid rgba(139,92,246,.5);"
                "color:#e2e8f0;padding:10px 16px;border-radius:12px;font-size:13px;"
                "font-family:system-ui,sans-serif;box-shadow:0 18px 50px rgba(0,0,0,.5);"
                'display:flex;align-items:center;gap:12px">'
                "<span>🛡️ Super Admin — editing <b>" + label + "</b>'s timetable</span>"
                '<a href="' + self.EXIT_PATH + '" style="background:linear-gradient(135deg,'
                "#3b82f6,#8b5cf6);color:#fff;padding:6px 12px;border-radius:8px;"
                'text-decoration:none;font-weight:600">Exit to Super Admin</a></div>'
            )
            response.content = html.replace("</body>", banner + "</body>", 1).encode("utf-8")
            if response.get("Content-Length") is not None:
                response["Content-Length"] = len(response.content)
        except Exception:
            # Never break the page over a cosmetic banner.
            pass
