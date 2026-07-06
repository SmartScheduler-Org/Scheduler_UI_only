from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from .forms import *
from .models import *
from account.models import Profile, TeacherOnboarding
from django.core.mail import send_mail, EmailMessage
from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.views.generic import View
import logging
import random as rnd
from django.contrib import messages
import os
import json
import requests
from django.http import HttpResponseForbidden, Http404
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.conf import settings
from django.utils import timezone
from django.db import transaction
from django.db.models import Q
import copy
import csv
import math
import re
from datetime import date
from smtplib import SMTPAuthenticationError
from itertools import combinations
from collections import defaultdict
from .models import MeetingTime
from .models import DAYS_OF_WEEK, TIME_SLOTS
from .forms import DepartmentForm
from ttgen.utils import section_sort_key


import random as rnd

logger = logging.getLogger(__name__)

LOGO_CID = "smartscheduler_logo"
SIH_WINNER_TEXT = "SmartScheduler | SIH Winner Innovation Team"
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")


def _brand_from_email():
    from email.utils import formataddr

    sender = (getattr(settings, "DEFAULT_FROM_EMAIL", "") or settings.EMAIL_HOST_USER or "").strip()
    if not sender:
        return ""
    if "<" in sender and ">" in sender:
        return sender
    return formataddr(("SmartScheduler", sender))


# ──────────────────────────────────────────────────────────────────────────
# Live generation log capture
# Mirrors everything the GA prints to the terminal into a per-user, in-memory
# ring buffer so the hosted loading screen can show the same logs the developer
# sees locally. A thread-aware tee on sys.stdout routes prints made by the
# generation request to that user's buffer while still echoing to the terminal.
# ──────────────────────────────────────────────────────────────────────────
import sys as _gen_sys
import threading as _gen_threading
import time as _gen_time

_GEN_LOG_LOCK = _gen_threading.Lock()
_GEN_LOG_BUFFERS = {}          # user_id -> {"lines": [...], "done": bool, "ts": float}
_GEN_LOG_THREAD_MAP = {}       # thread_id -> user_id
_GEN_LOG_MAX_LINES = 4000
_GEN_LOG_TEE_INSTALLED = False


class _GenLogTee:
    """Wraps the real stdout: echoes everything, and additionally appends lines
    to the buffer of the user whose generation runs on the current thread."""

    def __init__(self, real_stream):
        self._real = real_stream
        self._partials = {}  # thread_id -> leftover text without trailing newline

    def write(self, text):
        try:
            self._real.write(text)
        except Exception:
            pass
        try:
            tid = _gen_threading.get_ident()
            user_id = _GEN_LOG_THREAD_MAP.get(tid)
            if user_id is None or not text:
                return
            buf = self._partials.get(tid, "") + text
            lines = buf.split("\n")
            self._partials[tid] = lines.pop()  # last item is incomplete remainder
            if not lines:
                return
            with _GEN_LOG_LOCK:
                entry = _GEN_LOG_BUFFERS.get(user_id)
                if entry is None:
                    return
                store = entry["lines"]
                for ln in lines:
                    store.append(ln)
                if len(store) > _GEN_LOG_MAX_LINES:
                    del store[: len(store) - _GEN_LOG_MAX_LINES]
        except Exception:
            pass

    def flush(self):
        try:
            self._real.flush()
        except Exception:
            pass

    def __getattr__(self, name):
        return getattr(self._real, name)


def _gen_log_install_tee():
    global _GEN_LOG_TEE_INSTALLED
    if _GEN_LOG_TEE_INSTALLED:
        return
    try:
        _gen_sys.stdout = _GenLogTee(_gen_sys.stdout)
        _GEN_LOG_TEE_INSTALLED = True
    except Exception:
        _GEN_LOG_TEE_INSTALLED = False


def _gen_log_start(user_id):
    """Begin capturing stdout for this user's generation on the current thread."""
    if user_id is None:
        return
    _gen_log_install_tee()
    with _GEN_LOG_LOCK:
        _GEN_LOG_BUFFERS[user_id] = {
            "lines": ["[SmartScheduler] Generation engine starting…"],
            "done": False,
            "ts": _gen_time.time(),
        }
    _GEN_LOG_THREAD_MAP[_gen_threading.get_ident()] = user_id


def _gen_log_finish(user_id, message=None):
    """Stop routing this thread's stdout and mark the buffer complete."""
    _GEN_LOG_THREAD_MAP.pop(_gen_threading.get_ident(), None)
    if user_id is None:
        return
    with _GEN_LOG_LOCK:
        entry = _GEN_LOG_BUFFERS.get(user_id)
        if entry is not None:
            if message:
                entry["lines"].append(message)
            entry["done"] = True


@login_required
def generation_logs(request):
    """Return new generation log lines for the current user (incremental poll)."""
    user_id = request.user.id
    try:
        since = int(request.GET.get("since", "0"))
    except (TypeError, ValueError):
        since = 0
    with _GEN_LOG_LOCK:
        entry = _GEN_LOG_BUFFERS.get(user_id)
        if entry is None:
            return JsonResponse({"lines": [], "total": 0, "done": False, "active": False})
        total = len(entry["lines"])
        if since < 0:
            since = 0
        new_lines = entry["lines"][since:] if since < total else []
        return JsonResponse({
            "lines": new_lines,
            "total": total,
            "done": bool(entry["done"]),
            "active": True,
        })


def _delete_all_step_entries(request, queryset, redirect_name, entity_label):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    total = queryset.count()
    if total == 0:
        messages.info(request, f"No {entity_label} found to delete.")
        return redirect(redirect_name)

    queryset.delete()
    reset_global_schedule_cache(request.user.id)
    messages.success(request, f"Deleted {total} {entity_label}.")
    return redirect(redirect_name)


def _get_step_edit_object(request, model, query_param="edit", post_param="edit_id", **filters):
    raw_id = request.POST.get(post_param) if request.method == "POST" else request.GET.get(query_param)
    if not raw_id:
        return None
    try:
        object_id = int(raw_id)
    except (TypeError, ValueError):
        return None
    return model.objects.filter(pk=object_id, **filters).first()


def _get_required_step_edit_object(request, model, post_param="edit_id", **filters):
    raw_id = request.POST.get(post_param)
    if not raw_id:
        return None
    try:
        object_id = int(raw_id)
    except (TypeError, ValueError):
        raise Http404("Invalid edit item")
    obj = model.objects.filter(pk=object_id, **filters).first()
    if obj is None:
        raise Http404("Edit item not found")
    return obj


# ---------------- PER-USER STATE ----------------
_USER_STATE = {}  # {user_id: {"classes": ..., "labs": ..., "schedules": [...], "view_mode": ..., "data": ...}}


def _get_user_state(user_id):
    """Get or create per-user in-memory state."""
    if user_id not in _USER_STATE:
        _USER_STATE[user_id] = {
            "classes": None,
            "labs": None,
            "schedules": [],
            "view_mode": None,
            "data": None,
            "generated_parking_items": [],
            "generated_slot_room_reservations": {},
            "generated_parking_next_id": 1,
            "generated_manual_slot_next_id": 1,
            "generated_edit_index": None,
            "prefill_mode": False,
            "prefill_section_ids": [],
            "prefill_locked_classes": [],
            "prefill_locked_labs": [],
        }
    return _USER_STATE[user_id]


def _reset_generated_drag_state(state, current_index=None):
    state["generated_parking_items"] = []
    state["generated_slot_room_reservations"] = {}
    state["generated_parking_next_id"] = 1
    state["generated_manual_slot_next_id"] = 1
    state["generated_edit_index"] = current_index


def _ensure_manual_prefill_slot_uid(state, item_obj):
    uid = str(getattr(item_obj, "manual_slot_uid", "") or "").strip()
    if uid:
        return uid
    next_id = int(state.get("generated_manual_slot_next_id") or 1)
    uid = f"manual-{next_id}"
    state["generated_manual_slot_next_id"] = next_id + 1
    setattr(item_obj, "manual_slot_uid", uid)
    return uid


def ensure_manual_prefill_slot_uids(state):
    for item_obj in list(state.get("classes") or []):
        if getattr(item_obj, "manual_entry", False):
            _ensure_manual_prefill_slot_uid(state, item_obj)
    for item_obj in list(state.get("labs") or []):
        if getattr(item_obj, "manual_entry", False):
            _ensure_manual_prefill_slot_uid(state, item_obj)
    for parked in list(state.get("generated_parking_items") or []):
        item_obj = parked.get("item")
        if item_obj is not None and getattr(item_obj, "manual_entry", False):
            parked["manual_slot_uid"] = _ensure_manual_prefill_slot_uid(state, item_obj)


def _next_generated_parking_id(state):
    next_id = state.get("generated_parking_next_id", 1)
    state["generated_parking_next_id"] = next_id + 1
    return next_id


def _decorate_generated_tables_with_parking(tables, state):
    parking_by_section = defaultdict(list)
    for item in state.get("generated_parking_items", []):
        parking_by_section[item["section_id"]].append(item)

    reserved_rooms = state.get("generated_slot_room_reservations", {})
    for table in tables:
        section_id = table["section"].section_id
        table["parking_items"] = parking_by_section.get(section_id, [])
        for row in table.get("rows", []):
            for cell in row.get("cells", []):
                key = (section_id, row["day"], str(cell.get("slot_number")))
                reservation = reserved_rooms.get(key)
                cell["reserved_room"] = reservation["room"] if reservation else None
                cell["has_room_context"] = bool(reservation)
    return tables


class ManualPrefillSubject:
    def __init__(self, label, duration=1):
        clean_label = (label or "Manual Subject").strip() or "Manual Subject"
        self.pk = None
        self.id = None
        self.subject_number = clean_label
        self.subject_name = clean_label
        self.room_required = "Lecture Hall"
        self.required_lab_category = ""
        self.specific_rooms = ""
        self.classes_per_week = 0
        self.duration = max(1, min(int(duration or 1), 8))

    def __str__(self):
        return self.subject_name


def _resolve_manual_prefill_subject(section_obj, subject_text, duration):
    subject_text = (subject_text or "").strip()
    if not subject_text:
        return ManualPrefillSubject("Manual Subject", duration=duration)
    matched = section_obj.allowed_subjects.filter(subject_number__iexact=subject_text).first()
    if matched is None:
        matched = section_obj.allowed_subjects.filter(subject_name__iexact=subject_text).first()
    if matched is not None:
        return matched
    return ManualPrefillSubject(subject_text, duration=duration)


def _parse_prefill_duration(value):
    try:
        duration = int(value)
    except (TypeError, ValueError):
        duration = 1
    return max(1, min(duration, 8))


def _prefill_teacher_uid(teacher):
    if teacher is None:
        return ""
    return str(getattr(teacher, "uid", "") or getattr(teacher, "pk", "") or "")


def _prefill_room_number(room):
    if room is None:
        return ""
    return str(getattr(room, "r_number", "") or getattr(room, "pk", "") or "")


def _prefill_subject_text(subject):
    if subject is None:
        return "Manual Subject"
    return str(getattr(subject, "subject_name", "") or getattr(subject, "subject_number", "") or subject or "Manual Subject")


def _serialize_prefill_class(cls):
    meeting_times = list(getattr(cls, "meeting_times", None) or [])
    if not meeting_times and getattr(cls, "meeting_time", None):
        meeting_times = [cls.meeting_time]
    first_time = meeting_times[0] if meeting_times else None
    return {
        "type": "class",
        "section_id": str(getattr(cls, "section", "") or ""),
        "subject_text": _prefill_subject_text(getattr(cls, "subject", None)),
        "teacher_uid": _prefill_teacher_uid(getattr(cls, "instructor", None)),
        "co_teacher_uids": [_prefill_teacher_uid(teacher) for teacher in list(getattr(cls, "co_instructors", []) or []) if teacher],
        "room_number": _prefill_room_number(getattr(cls, "room", None)),
        "day": getattr(first_time, "day", "") if first_time else "",
        "start_slot": str(getattr(first_time, "time", "") if first_time else ""),
        "duration": max(1, min(int(getattr(cls, "duration", None) or len(meeting_times) or 1), 8)),
        "manual_entry": bool(getattr(cls, "manual_entry", False)),
        "manual_slot_uid": str(getattr(cls, "manual_slot_uid", "") or ""),
        "prefill_locked": bool(getattr(cls, "prefill_locked", False)),
    }


def _serialize_prefill_lab(lab):
    meeting_times = list(getattr(lab, "meeting_times", None) or [])
    first_time = meeting_times[0] if meeting_times else None
    return {
        "type": "lab",
        "section_id": str(getattr(lab, "section", "") or ""),
        "subject_text": _prefill_subject_text(getattr(lab, "subject", None)),
        "teacher_uid": _prefill_teacher_uid(getattr(lab, "instructor", None)),
        "second_teacher_uid": _prefill_teacher_uid(getattr(lab, "second_instructor", None)),
        "co_teacher_uids": [_prefill_teacher_uid(teacher) for teacher in list(getattr(lab, "co_instructors", []) or []) if teacher],
        "room_number": _prefill_room_number(getattr(lab, "room", None)),
        "day": getattr(first_time, "day", "") if first_time else "",
        "start_slot": str(getattr(first_time, "time", "") if first_time else ""),
        "duration": max(1, min(int(getattr(lab, "duration", None) or len(meeting_times) or LAB_DURATION), 8)),
        "manual_entry": bool(getattr(lab, "manual_entry", False)),
        "manual_slot_uid": str(getattr(lab, "manual_slot_uid", "") or ""),
        "prefill_locked": bool(getattr(lab, "prefill_locked", False)),
        "batch": int(getattr(lab, "batch", 1) or 1),
        "total_batches": int(getattr(lab, "total_batches", 1) or 1),
    }


def save_prefill_session_snapshot(request, state=None):
    if not request or not getattr(request, "user", None) or not request.user.is_authenticated:
        return None
    state = state or _get_user_state(request.user.id)
    if not state.get("prefill_mode"):
        return None
    ensure_manual_prefill_slot_uids(state)
    snapshot = {
        "section_ids": list(state.get("prefill_section_ids") or request.session.get("prefill_section_ids") or []),
        "classes": [_serialize_prefill_class(cls) for cls in list(state.get("classes") or [])],
        "labs": [_serialize_prefill_lab(lab) for lab in list(state.get("labs") or [])],
        "parking_items": [],
    }
    for parked in list(state.get("generated_parking_items") or []):
        item_obj = parked.get("item")
        if item_obj is None:
            continue
        entry = _serialize_prefill_lab(item_obj) if parked.get("move_type") == "lab" else _serialize_prefill_class(item_obj)
        entry.update({
            "move_type": parked.get("move_type") or entry["type"],
            "section_id": str(parked.get("section_id") or entry.get("section_id") or ""),
            "slot_span": max(1, min(int(parked.get("slot_span") or entry.get("duration") or 1), 8)),
        })
        snapshot["parking_items"].append(entry)
    request.session["prefill_saved_slots"] = snapshot
    request.session.modified = True
    return snapshot


_PREFILL_CSV_COLUMNS = [
    "record_type",
    "section_id",
    "move_type",
    "subject_text",
    "teacher_uid",
    "second_teacher_uid",
    "co_teacher_uids",
    "room_number",
    "day",
    "start_slot",
    "duration",
    "manual_entry",
    "manual_slot_uid",
    "prefill_locked",
    "batch",
    "total_batches",
    "parking",
    "slot_span",
]


def _prefill_csv_bool(value):
    return "true" if bool(value) else "false"


def _parse_prefill_csv_bool(value):
    return str(value or "").strip().lower() in {"1", "true", "yes", "on"}


def _prefill_csv_row_from_entry(entry, parking=False):
    move_type = str(entry.get("move_type") or entry.get("type") or "class")
    return {
        "record_type": "slot",
        "section_id": str(entry.get("section_id") or ""),
        "move_type": move_type,
        "subject_text": str(entry.get("subject_text") or ""),
        "teacher_uid": str(entry.get("teacher_uid") or ""),
        "second_teacher_uid": str(entry.get("second_teacher_uid") or ""),
        "co_teacher_uids": ";".join([str(uid or "").strip() for uid in list(entry.get("co_teacher_uids") or []) if str(uid or "").strip()]),
        "room_number": str(entry.get("room_number") or ""),
        "day": str(entry.get("day") or ""),
        "start_slot": str(entry.get("start_slot") or ""),
        "duration": str(entry.get("duration") or ""),
        "manual_entry": _prefill_csv_bool(entry.get("manual_entry", False)),
        "manual_slot_uid": str(entry.get("manual_slot_uid") or ""),
        "prefill_locked": _prefill_csv_bool(entry.get("prefill_locked", False)),
        "batch": str(entry.get("batch") or ""),
        "total_batches": str(entry.get("total_batches") or ""),
        "parking": _prefill_csv_bool(parking),
        "slot_span": str(entry.get("slot_span") or entry.get("duration") or ""),
    }


def _build_prefill_csv_snapshot(decoded_lines):
    reader = csv.DictReader(decoded_lines)
    required_columns = {"record_type", "section_id"}
    if reader.fieldnames is None or not required_columns.issubset({str(name or "").strip() for name in reader.fieldnames}):
        raise ValueError("CSV header is invalid. Please use a downloaded prefill CSV file.")

    snapshot = {
        "section_ids": [],
        "classes": [],
        "labs": [],
        "parking_items": [],
    }
    seen_sections = set()

    for row_number, row in enumerate(reader, start=2):
        record_type = str((row or {}).get("record_type") or "").strip().lower()
        section_id = str((row or {}).get("section_id") or "").strip()
        if not record_type and not section_id:
            continue
        if not section_id:
            raise ValueError(f"Row {row_number}: section_id is required.")
        if section_id not in seen_sections:
            seen_sections.add(section_id)
            snapshot["section_ids"].append(section_id)

        if record_type == "section":
            continue
        if record_type != "slot":
            raise ValueError(f"Row {row_number}: unsupported record_type '{record_type}'.")

        move_type = str((row or {}).get("move_type") or "class").strip().lower()
        if move_type not in {"class", "lab"}:
            raise ValueError(f"Row {row_number}: move_type must be 'class' or 'lab'.")

        entry = {
            "type": move_type,
            "section_id": section_id,
            "subject_text": str((row or {}).get("subject_text") or "").strip(),
            "teacher_uid": str((row or {}).get("teacher_uid") or "").strip(),
            "second_teacher_uid": str((row or {}).get("second_teacher_uid") or "").strip(),
            "co_teacher_uids": [token.strip() for token in str((row or {}).get("co_teacher_uids") or "").split(";") if token.strip()],
            "room_number": str((row or {}).get("room_number") or "").strip(),
            "day": str((row or {}).get("day") or "").strip(),
            "start_slot": str((row or {}).get("start_slot") or "").strip(),
            "duration": str((row or {}).get("duration") or "").strip(),
            "manual_entry": _parse_prefill_csv_bool((row or {}).get("manual_entry")),
            "manual_slot_uid": str((row or {}).get("manual_slot_uid") or "").strip(),
            "prefill_locked": _parse_prefill_csv_bool((row or {}).get("prefill_locked")),
        }
        if move_type == "lab":
            entry["batch"] = int(str((row or {}).get("batch") or "1").strip() or 1)
            entry["total_batches"] = int(str((row or {}).get("total_batches") or "1").strip() or 1)

        if not entry["subject_text"]:
            raise ValueError(f"Row {row_number}: subject_text is required.")
        if not entry["teacher_uid"]:
            raise ValueError(f"Row {row_number}: teacher_uid is required.")

        if _parse_prefill_csv_bool((row or {}).get("parking")):
            entry["move_type"] = move_type
            entry["slot_span"] = max(1, min(int(str((row or {}).get("slot_span") or entry.get("duration") or "1").strip() or 1), 8))
            snapshot["parking_items"].append(entry)
        elif move_type == "lab":
            snapshot["labs"].append(entry)
        else:
            snapshot["classes"].append(entry)

    return snapshot


def _activate_prefill_snapshot(request, snapshot):
    section_ids = [str(section_id) for section_id in list((snapshot or {}).get("section_ids") or []) if str(section_id or "").strip()]
    if not section_ids:
        return False
    state = _get_user_state(request.user.id)
    state["classes"] = []
    state["labs"] = []
    state["schedules"] = [{"classes": [], "labs": [], "stats": {}, "reco_block": {}}]
    state["prefill_mode"] = True
    state["prefill_section_ids"] = section_ids
    state["prefill_locked_classes"] = []
    state["prefill_locked_labs"] = []
    _reset_generated_drag_state(state, current_index=1)
    request.session["current_index"] = 1
    request.session["prefill_mode"] = True
    request.session["prefill_section_ids"] = section_ids
    request.session["prefill_saved_slots"] = snapshot
    request.session.modified = True
    restore_prefill_session_snapshot(request, state, section_ids)
    return True


def _clear_active_saved_prefill(request):
    request.session.pop("active_saved_prefill_id", None)
    request.session.modified = True


def _set_active_saved_prefill(request, prefill):
    request.session["active_saved_prefill_id"] = prefill.pk
    request.session.modified = True


def _get_active_saved_prefill(request):
    prefill_id = request.session.get("active_saved_prefill_id")
    if not prefill_id:
        return None
    return SavedPrefill.objects.filter(pk=prefill_id, user=request.user).first()


def _combine_prefill_snapshots(prefills):
    combined = {
        "section_ids": [],
        "classes": [],
        "labs": [],
        "parking_items": [],
    }
    seen_sections = set()
    for prefill in prefills:
        snapshot = prefill.snapshot or {}
        for section_id in list(snapshot.get("section_ids") or []):
            section_id = str(section_id or "").strip()
            if not section_id or section_id in seen_sections:
                continue
            seen_sections.add(section_id)
            combined["section_ids"].append(section_id)
        combined["classes"].extend(list(snapshot.get("classes") or []))
        combined["labs"].extend(list(snapshot.get("labs") or []))
        combined["parking_items"].extend(list(snapshot.get("parking_items") or []))
    return combined


def _prefill_restore_teacher(uid, user):
    uid = str(uid or "").strip()
    if not uid:
        return None
    qs = Instructor.objects.filter(user=user)
    if uid.isdigit():
        return qs.filter(Q(uid__iexact=uid) | Q(pk=int(uid))).first()
    return qs.filter(uid__iexact=uid).first()


def _prefill_restore_room(room_number, user):
    room_number = str(room_number or "").strip()
    if not room_number:
        return None
    qs = Room.objects.filter(user=user)
    if room_number.isdigit():
        return qs.filter(Q(r_number__iexact=room_number) | Q(pk=int(room_number))).first()
    return qs.filter(r_number__iexact=room_number).first()


def _prefill_restore_meeting_times(entry, user):
    day = entry.get("day")
    try:
        start_slot = int(entry.get("start_slot"))
    except (TypeError, ValueError):
        return []
    duration = _parse_prefill_duration(entry.get("duration"))
    times = []
    for offset in range(duration):
        meeting_time = get_meeting_time(day, start_slot + offset, user=user)
        if meeting_time is None:
            return []
        times.append(meeting_time)
    return times


def _prefill_restore_class(entry, user):
    ClassImpl = globals().get("Class")
    if ClassImpl is None:
        return None
    try:
        section_obj = Section.objects.get(section_id=entry.get("section_id"), user=user)
    except Section.DoesNotExist:
        return None
    teacher = _prefill_restore_teacher(entry.get("teacher_uid"), user)
    room_number = str(entry.get("room_number") or "").strip()
    room = _prefill_restore_room(room_number, user)
    if teacher is None:
        return None
    missing_room_label = ""
    if room is None:
        if not room_number:
            return None
        missing_room_label = room_number
    duration = _parse_prefill_duration(entry.get("duration"))
    subject = _resolve_manual_prefill_subject(section_obj, entry.get("subject_text"), duration)
    cls = ClassImpl(_next_in_memory_class_id(), section_obj.department, section_obj.section_id, subject)
    cls.set_instructor(teacher)
    co_teachers = [_prefill_restore_teacher(uid, user) for uid in list(entry.get("co_teacher_uids") or [])]
    if hasattr(cls, "set_co_instructors"):
        cls.set_co_instructors([teacher_obj for teacher_obj in co_teachers if teacher_obj])
    cls.set_room(room)
    if missing_room_label:
        cls.room_label = missing_room_label
        cls.missing_room = True
    times = _prefill_restore_meeting_times(entry, user)
    if times:
        cls.set_meetingTime(times[0])
        cls.meeting_times = times
    cls.duration = duration
    cls.manual_entry = bool(entry.get("manual_entry", True))
    cls.manual_slot_uid = str(entry.get("manual_slot_uid") or "")
    cls.prefill_locked = bool(entry.get("prefill_locked", bool(times)))
    return cls


def _prefill_restore_lab(entry, user):
    LabImpl = globals().get("Lab")
    if LabImpl is None:
        return None
    try:
        section_obj = Section.objects.get(section_id=entry.get("section_id"), user=user)
    except Section.DoesNotExist:
        return None
    teacher = _prefill_restore_teacher(entry.get("teacher_uid"), user)
    room_number = str(entry.get("room_number") or "").strip()
    room = _prefill_restore_room(room_number, user)
    if teacher is None:
        return None
    missing_room_label = ""
    if room is None:
        if not room_number:
            return None
        missing_room_label = room_number
    duration = _parse_prefill_duration(entry.get("duration"))
    subject = _resolve_manual_prefill_subject(section_obj, entry.get("subject_text"), duration)
    lab = LabImpl(_next_in_memory_class_id(), section_obj.department, section_obj.section_id, subject, entry.get("batch", 1), entry.get("total_batches", 1))
    lab.set_instructor(teacher)
    second_teacher = _prefill_restore_teacher(entry.get("second_teacher_uid"), user)
    if second_teacher and hasattr(lab, "set_second_instructor"):
        lab.set_second_instructor(second_teacher)
    co_teachers = [_prefill_restore_teacher(uid, user) for uid in list(entry.get("co_teacher_uids") or [])]
    if hasattr(lab, "set_co_instructors"):
        lab.set_co_instructors([teacher_obj for teacher_obj in co_teachers if teacher_obj])
    lab.set_room(room)
    if missing_room_label:
        lab.room_label = missing_room_label
        lab.missing_room = True
    times = _prefill_restore_meeting_times(entry, user)
    if times:
        lab.set_meetingTimes(times)
    lab.duration = duration
    lab.manual_entry = bool(entry.get("manual_entry", True))
    lab.manual_slot_uid = str(entry.get("manual_slot_uid") or "")
    lab.prefill_locked = bool(entry.get("prefill_locked", bool(times)))
    return lab


def restore_prefill_session_snapshot(request, state, section_ids):
    snapshot = request.session.get("prefill_saved_slots") or {}
    saved_sections = set(snapshot.get("section_ids") or [])
    allowed_sections = set(section_ids or [])
    if saved_sections and saved_sections != allowed_sections:
        return False
    if state.get("classes") or state.get("labs") or state.get("generated_parking_items"):
        return False
    classes = []
    labs = []
    for entry in list(snapshot.get("classes") or []):
        if entry.get("section_id") not in allowed_sections:
            continue
        restored = _prefill_restore_class(entry, request.user)
        if restored is not None:
            classes.append(restored)
    for entry in list(snapshot.get("labs") or []):
        if entry.get("section_id") not in allowed_sections:
            continue
        restored = _prefill_restore_lab(entry, request.user)
        if restored is not None:
            labs.append(restored)
    state["classes"] = classes
    state["labs"] = labs
    state["prefill_locked_classes"] = list(classes)
    state["prefill_locked_labs"] = list(labs)
    state["generated_parking_items"] = []
    state["generated_parking_next_id"] = 1
    build_parking_item = globals().get("_generated_build_parking_item")
    if build_parking_item:
        for entry in list(snapshot.get("parking_items") or []):
            if entry.get("section_id") not in allowed_sections:
                continue
            move_type = entry.get("move_type") or entry.get("type") or "class"
            restored = _prefill_restore_lab(entry, request.user) if move_type == "lab" else _prefill_restore_class(entry, request.user)
            if restored is not None:
                state["generated_parking_items"].append(build_parking_item(state, entry.get("section_id"), move_type, restored))
    return bool(classes or labs or state.get("generated_parking_items"))


# Legacy aliases so private modules loaded via exec() can reference them.
# These are updated by show_timetable / generate to point at the current user's data.
GLOBAL_CLASSES = None
GLOBAL_LABS = None
GLOBAL_GENERATED_SCHEDULES = []
CURRENT_VIEW_MODE = None

POPULATION_SIZE = 3
USE_PSO_REFINEMENT = True
NUMB_OF_ELITE_SCHEDULES = 1
TOURNAMENT_SELECTION_SIZE = 2
MUTATION_RATE = 0.05

LAB_DURATION = 4
VALID_LAB_START_SLOTS = ["1", "6"]
LUNCH_SLOT = "5"


def get_valid_start_slots(duration):
    """Return valid start slots for a given duration, avoiding lunch (slot 5)."""
    if duration > 4:
        usable_slots = [str(slot) for slot in range(1, 10) if str(slot) != LUNCH_SLOT]
        return [
            slot
            for index, slot in enumerate(usable_slots)
            if len(usable_slots[index:index + duration]) == duration
        ]
    pre_lunch = [str(s) for s in range(1, 5) if s + duration - 1 <= 4]
    post_lunch = [str(s) for s in range(6, 10) if s + duration - 1 <= 9]
    return pre_lunch + post_lunch


def _csv_issue(issues, row_number, detail):
    issues.append(f"Row {row_number}: {detail}")


def _emit_csv_issues(request, issues):
    for issue in issues:
        messages.warning(request, issue)




# placeholder for private generation algorithm loaded outside this repo
data = None
GENERATOR_RULES_AVAILABLE = False
GENERATOR_ALGO_AVAILABLE = False
GENERATOR_RUNTIME_AVAILABLE = False
_PRIVATE_GENERATOR_MTIMES = {
    "rules": None,
    "algo": None,
    "runtime": None,
}


def _private_file_path(env_var, filename):
    configured_path = os.environ.get(env_var)
    if configured_path:
        return os.path.expanduser(configured_path)
    configured_dir = os.environ.get("TTGEN_PRIVATE_DIR")
    if configured_dir:
        return os.path.join(os.path.expanduser(configured_dir), filename)
    return os.path.join(os.path.expanduser("~"), ".ttgen_private", filename)


def _load_private_generator_rules():
    global GENERATOR_RULES_AVAILABLE

    rules_path = _private_file_path("TTGEN_RULES_PATH", "views_other_rules.py")
    if not os.path.exists(rules_path):
        return

    current_mtime = os.path.getmtime(rules_path)
    if _PRIVATE_GENERATOR_MTIMES["rules"] == current_mtime:
        return

    with open(rules_path, "r", encoding="utf-8") as rules_file:
        code = compile(rules_file.read(), rules_path, "exec")

    exec(code, globals())
    GENERATOR_RULES_AVAILABLE = True
    _PRIVATE_GENERATOR_MTIMES["rules"] = current_mtime


def _load_private_generator_algo():
    global GENERATOR_ALGO_AVAILABLE

    algo_path = _private_file_path("TTGEN_ALGO_PATH", "views_other_algorithm.py")
    if not os.path.exists(algo_path):
        return

    current_mtime = os.path.getmtime(algo_path)
    if _PRIVATE_GENERATOR_MTIMES["algo"] == current_mtime:
        return

    with open(algo_path, "r", encoding="utf-8") as algo_file:
        code = compile(algo_file.read(), algo_path, "exec")

    exec(code, globals())
    GENERATOR_ALGO_AVAILABLE = True
    _PRIVATE_GENERATOR_MTIMES["algo"] = current_mtime


def _load_private_generator_runtime():
    global GENERATOR_RUNTIME_AVAILABLE

    runtime_path = _private_file_path("TTGEN_RUNTIME_PATH", "views_other_runtime.py")
    if not os.path.exists(runtime_path):
        return

    current_mtime = os.path.getmtime(runtime_path)
    if _PRIVATE_GENERATOR_MTIMES["runtime"] == current_mtime:
        return

    with open(runtime_path, "r", encoding="utf-8") as runtime_file:
        code = compile(runtime_file.read(), runtime_path, "exec")

    exec(code, globals())
    GENERATOR_RUNTIME_AVAILABLE = True
    _PRIVATE_GENERATOR_MTIMES["runtime"] = current_mtime


def ensure_private_generator_loaded():
    _load_private_generator_rules()
    _load_private_generator_algo()
    _load_private_generator_runtime()


_load_private_generator_rules()
_load_private_generator_algo()
_load_private_generator_runtime()

def ensure_cs_department(user=None):
    if user:
        department, _ = Department.objects.get_or_create(
            user=user,
            code="CS",
            defaults={"name": "Computer Science"},
        )
    else:
        department, _ = Department.objects.get_or_create(
            code="CS",
            defaults={"name": "Computer Science"},
        )
    return department




SECTION_SEMESTER_PATTERN = re.compile(r"(\d+(?:st|nd|rd|th)\s+sem)", re.IGNORECASE)


def get_section_signature(section_id):
    raw = (section_id or "").strip().lower()
    match = SECTION_SEMESTER_PATTERN.search(raw)
    semester = match.group(1).lower() if match else ""
    prefix = raw[:match.start()] if match else raw
    tokens = []
    for token in prefix.split():
        normalized = re.sub(r"[^a-z]+", "", token)
        if normalized:
            tokens.append(normalized)
    return semester, tuple(tokens)


def clone_section_subjects_from_similar(section):
    if section.allowed_subjects.exists():
        return None

    semester, tokens = get_section_signature(section.section_id)
    if not semester or not tokens:
        return None

    best_match = None
    best_score = None

    candidates = (
        Section.objects.filter(department=section.department, user=section.user)
        .exclude(pk=section.pk)
        .prefetch_related("allowed_subjects")
    )
    for candidate in candidates:
        if not candidate.allowed_subjects.exists():
            continue
        candidate_semester, candidate_tokens = get_section_signature(candidate.section_id)
        if candidate_semester != semester:
            continue
        overlap = len(set(tokens) & set(candidate_tokens))
        if overlap <= 0:
            continue
        score = (overlap, candidate.allowed_subjects.count(), -len(candidate.section_id))
        if best_score is None or score > best_score:
            best_score = score
            best_match = candidate

    if not best_match:
        return None

    subjects_to_copy = list(best_match.allowed_subjects.all())
    if not subjects_to_copy:
        return None

    section.allowed_subjects.add(*subjects_to_copy)
    return best_match





# BASIC NAVIGATION VIEWS
def index(request): return render(request, 'index.html')
def about(request): return render(request, 'aboutus.html')
def live_demo(request): return render(request, 'live_demo.html')
def services(request): return render(request, 'services.html')
def help(request): return render(request, 'help.html')
def terms(request): return render(request, 'terms.html')
def privacy(request): return render(request, 'privacy.html')
def role(request):
    if request.user.is_authenticated:
        profile, _ = Profile.objects.get_or_create(user=request.user)
        role_value = (profile.role or "").strip().lower()
        if role_value:
            # Already has a role and should be sent to that role's page.
            if role_value == 'hod':
                return redirect('admindash')
            elif role_value == 'teacher':
                return redirect('teacher_dashboard')
            elif role_value == 'dean':
                return redirect('teachertimetable')
    return render(request, 'role.html')

@login_required
def admindash_role_set(request):
    """Set role to HOD and redirect to dashboard."""
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if profile.role and profile.role != 'hod':
        return render(request, 'role_locked.html', {'current_role': profile.get_role_display()})
    profile.role = 'hod'
    profile.save()
    return redirect('admindash')


@login_required
def teacher_role_set(request):
    """Set role to Teacher and redirect."""
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if profile.role and profile.role != 'teacher':
        return render(request, 'role_locked.html', {'current_role': profile.get_role_display()})
    profile.role = 'teacher'
    profile.save()
    return redirect('teacher_dashboard')


@login_required
def dean_role_set(request):
    """Set role to Dean and redirect."""
    profile, _ = Profile.objects.get_or_create(user=request.user)
    if profile.role and profile.role != 'dean':
        return render(request, 'role_locked.html', {'current_role': profile.get_role_display()})
    profile.role = 'dean'
    profile.save()
    return redirect('teachertimetable')


def teacherlogin(request):
    """Teacher login by Name + Teacher Code + Password (all three must match)."""
    from django.contrib.auth import login as auth_login

    if request.user.is_authenticated:
        profile, _ = Profile.objects.get_or_create(user=request.user)
        if (profile.role or "").lower() == "teacher" and (profile.linked_instructor_id or getattr(profile, "linked_admin_teacher_id", None)):
            return redirect("teacher_dashboard")

    if request.method == "POST":
        name = (request.POST.get("name") or "").strip()
        code = (request.POST.get("code") or "").strip()
        password = request.POST.get("password") or ""

        if not name or not code or not password:
            messages.error(request, "Please enter your name, teacher code and password.")
            return render(request, "teacherlogin.html", _teacher_nav_context())

        authenticated_user = None
        profiles = Profile.objects.select_related("user", "linked_instructor", "linked_admin_teacher").filter(role__iexact="teacher")
        for profile in profiles:
            identity_name = ""
            identity_code = ""
            if getattr(profile, "linked_admin_teacher", None):
                identity_name = (profile.linked_admin_teacher.name or "").strip()
                identity_code = (profile.linked_admin_teacher.uid or "").strip()
            elif getattr(profile, "linked_instructor", None):
                identity_name = (profile.linked_instructor.name or "").strip()
                identity_code = (profile.linked_instructor.uid or "").strip()

            if not identity_name or not identity_code:
                continue
            if identity_name.lower() == name.lower() and identity_code.lower() == code.lower() and profile.user.check_password(password):
                authenticated_user = profile.user
                break

        if authenticated_user is None:
            messages.error(request, "Invalid name, teacher code or password.")
            return render(request, "teacherlogin.html", _teacher_nav_context())

        auth_login(request, authenticated_user, backend="django.contrib.auth.backends.ModelBackend")
        return redirect("teacher_dashboard")

    return render(request, "teacherlogin.html", _teacher_nav_context())

def deanlogin(request): return render(request, 'deanlogin.html')
def teachertimetable(request): return render(request, 'teachertimetable.html')


def _get_teacher_profile_or_locked_response(user):
    profile, _ = Profile.objects.get_or_create(user=user)
    role_value = (profile.role or "").strip().lower()
    if role_value and role_value != "teacher":
        return profile, profile.get_role_display()
    return profile, ""


def _get_active_teacher_timetable(profile):
    if not profile.active_timetable_id:
        return None
    return SavedTimetable.objects.select_related("user").filter(
        id=profile.active_timetable_id,
        is_published=True,
    ).first()


def _connect_teacher_timetable(profile, code):
    timetable = SavedTimetable.objects.select_related("user").filter(
        is_published=True,
        publish_code=code,
    ).first()
    if not timetable:
        return None, "No published timetable found with that code."

    update_fields = ["active_timetable"]
    profile.active_timetable = timetable
    resolved_instructor = _resolve_profile_instructor(profile, timetable.user)
    if resolved_instructor and profile.linked_instructor_id != resolved_instructor.id:
        profile.linked_instructor = resolved_instructor
        update_fields.append("linked_instructor")
    elif profile.linked_instructor and profile.linked_instructor.user_id != timetable.user_id:
        profile.linked_instructor = None
        update_fields.append("linked_instructor")
    profile.save(update_fields=update_fields)
    return timetable, None


def _ensure_teacher_role(profile):
    if not profile.role:
        profile.role = "teacher"
        profile.save(update_fields=["role"])


def _get_teacher_onboarding(user):
    return getattr(user, "teacher_onboarding", None)


def _teacher_onboarding_redirect_response(request):
    return None


def _user_can_manage_teacher_onboarding(user):
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    profile, _ = Profile.objects.get_or_create(user=user)
    return (profile.role or "").strip().lower() == "hod"


def _teacher_nav_context():
    return {
        "hide_live_demo": True,
        "hide_nav_cta": True,
        "show_teacher_profile_icon": True,
    }


def _format_local_datetime(value, fmt):
    if not value:
        return ""
    return timezone.localtime(value).strftime(fmt)


def _resolve_teacher_dashboard_context(request, profile):
    active_timetable = _get_active_teacher_timetable(profile)
    onboarding = _get_teacher_onboarding(request.user)
    sync_fields = []

    if profile.active_timetable_id and active_timetable is None:
        profile.active_timetable = None
        sync_fields.append("active_timetable")

    resolved_instructor = _resolve_profile_instructor(profile, active_timetable.user if active_timetable else None)
    if resolved_instructor and profile.linked_instructor_id != resolved_instructor.id:
        profile.linked_instructor = resolved_instructor
        sync_fields.append("linked_instructor")
    elif profile.linked_instructor_id and getattr(profile, "linked_admin_teacher_id", None) and resolved_instructor is None:
        profile.linked_instructor = None
        sync_fields.append("linked_instructor")
    elif (
        profile.linked_instructor_id
        and active_timetable
        and profile.linked_instructor.user_id != active_timetable.user_id
        and getattr(profile, "linked_admin_teacher_id", None) is None
    ):
        profile.linked_instructor = None
        sync_fields.append("linked_instructor")

    if sync_fields:
        profile.save(update_fields=sync_fields)

    linked_instructor = profile.linked_instructor
    linked_admin_teacher = getattr(profile, "linked_admin_teacher", None)
    teacher_subjects = []
    my_teacher_table = None
    teacher_workload = {"lectures": 0, "labs": 0, "shared_labs": 0, "total": 0}
    published_section_count = 0
    published_teacher_count = 0

    if linked_instructor:
        teacher_subjects = list(
            Subject.objects.filter(
                user=linked_instructor.user,
                instructors=linked_instructor,
            ).order_by("subject_name", "subject_number").distinct()
        )

    if active_timetable:
        classes, labs = _rebuild_classes_and_labs_from_saved(active_timetable)
        section_tables = build_section_tables(classes, labs, user=active_timetable.user)
        teacher_tables = build_teacher_tables(classes, labs, user=active_timetable.user)
        teacher_workloads = _compute_teacher_workloads(classes, labs)
        published_section_count = len(section_tables)
        published_teacher_count = len(teacher_tables)
        if linked_instructor:
            my_teacher_table = next(
                (
                    table for table in teacher_tables
                    if table["teacher"].id == linked_instructor.id
                ),
                None,
            )
            teacher_workload = teacher_workloads.get(linked_instructor, teacher_workload)

    schedule_rows = _teacher_schedule_rows(my_teacher_table)
    rooms_used = sorted({row["room"] for row in schedule_rows if row.get("room") and row["room"] != "—"})

    teacher_name = ""
    teacher_code = ""
    teacher_email = ""
    teacher_department = "—"
    teacher_department_code = ""
    if linked_admin_teacher:
        teacher_name = linked_admin_teacher.name or ""
        teacher_code = linked_admin_teacher.uid or ""
        teacher_email = linked_admin_teacher.email or ""
        teacher_department = linked_admin_teacher.department_name or teacher_department
        teacher_department_code = linked_admin_teacher.department_code or ""
    if linked_instructor and not linked_admin_teacher:
        dept = linked_instructor.department
        teacher_name = linked_instructor.name or teacher_name
        teacher_code = linked_instructor.uid or teacher_code
        teacher_email = linked_instructor.email or teacher_email
        teacher_department = dept.name if dept else teacher_department
        teacher_department_code = dept.code if dept else teacher_department_code
    elif linked_instructor and linked_admin_teacher:
        dept = linked_instructor.department
        if dept and teacher_department == "—":
            teacher_department = dept.name or teacher_department
            teacher_department_code = dept.code or teacher_department_code

    display_name = (
        onboarding.full_name
        if onboarding
        else request.user.get_full_name().strip() or request.user.username
    )
    if teacher_name:
        display_name = teacher_name

    role_label = profile.get_role_display() or "Teacher"
    designation = (
        linked_admin_teacher.designation
        if linked_admin_teacher
        else linked_instructor.designation if linked_instructor
        else onboarding.designation if onboarding else "Teacher"
    )
    profile_card_state = (
        f"Linked to {linked_instructor.uid}"
        if linked_instructor else
        f"Central teacher {linked_admin_teacher.uid or linked_admin_teacher.name}"
        if linked_admin_teacher else
        "Add contact and faculty UID"
    )
    published_card_state = (
        f"Code {active_timetable.publish_code}"
        if active_timetable else
        "Connect publish code"
    )
    if my_teacher_table:
        timetable_card_state = "Teacher timetable ready"
    elif active_timetable and linked_instructor:
        timetable_card_state = "No classes assigned yet"
    elif active_timetable:
        timetable_card_state = "Link faculty UID first"
    else:
        timetable_card_state = "Connect published timetable first"

    context = {
        "teacher_profile": profile,
        "teacher_onboarding": onboarding,
        "active_timetable": active_timetable,
        "linked_instructor": linked_instructor,
        "linked_admin_teacher": linked_admin_teacher,
        "teacher_subjects": teacher_subjects,
        "my_teacher_table": my_teacher_table,
        "teacher_workload": teacher_workload,
        "schedule_rows": schedule_rows,
        "rooms_used": rooms_used,
        "rooms_count": len(rooms_used),
        "class_count": len(schedule_rows),
        "published_section_count": published_section_count,
        "published_teacher_count": published_teacher_count,
        "teacher_display_name": display_name,
        "teacher_role_label": role_label,
        "teacher_designation": designation,
        "teacher_code": teacher_code,
        "teacher_email": teacher_email or (request.user.email or ""),
        "teacher_department": teacher_department,
        "teacher_department_code": teacher_department_code,
        "faculty_uid_value": linked_instructor.uid if linked_instructor else teacher_code,
        "profile_email_value": teacher_email or (request.user.email or ""),
        "slot_labels": SLOT_LABELS,
        "profile_card_state": profile_card_state,
        "published_card_state": published_card_state,
        "timetable_card_state": timetable_card_state,
        "teacher_subject_count": len(teacher_subjects),
    }
    context.update(_teacher_nav_context())
    return context


# ===========================================================================
# NEW TEACHER PORTAL — credential based (register / login / dashboard)
# Teachers register by selecting their department + name (auto-filled teacher
# code & email), then set their own password. Login matches Name + Teacher
# Code + Password. This replaces the old Google + onboarding + publish-code
# flow for teachers.
# ===========================================================================
TEACHER_DEFAULT_PASSWORD = None  # teachers set their own password


# Hardcoded department list (every department in the DB, kept as-is).
TEACHER_DEPARTMENT_LIST = [
    {"name": "Business Studies", "code": "BS"},
    {"name": "Centre for Energy Studies", "code": "CES"},
    {"name": "Chemistry", "code": "CHE"},
    {"name": "Civil Engineering", "code": "CE"},
    {"name": "Communication and Media Technology", "code": "CMT"},
    {"name": "Community College of Skill Development", "code": "CCSD"},
    {"name": "Computer Science", "code": "CL"},
    {"name": "Computer Science & Applications", "code": "CSA"},
    {"name": "Computer Science & Engineering", "code": "CSE"},
    {"name": "Electrical Engineering", "code": "EL"},
    {"name": "Electronics Engineering", "code": "EE"},
    {"name": "Environmental Sciences", "code": "ES"},
    {"name": "Life Sciences", "code": "LS"},
    {"name": "Literature & Languages", "code": "LL"},
    {"name": "Management Studies", "code": "MS"},
    {"name": "Mathematics", "code": "MATH"},
    {"name": "Mechanical Engineering", "code": "MECH"},
    {"name": "Physics", "code": "PHY"},
]


def _teacher_department_catalog():
    catalog = {}
    for item in TEACHER_DEPARTMENT_LIST:
        code = (item.get("code") or "").strip()
        name = (item.get("name") or "").strip()
        if code:
            catalog[code.lower()] = {"name": name or code, "code": code}
    for name, code in Department.objects.values_list("name", "code").distinct():
        code = (code or "").strip()
        name = (name or "").strip()
        if code:
            catalog[code.lower()] = {"name": name or code, "code": code}
    return catalog


def _normalize_teacher_department(name, code):
    name = (name or "").strip()
    code = (code or "").strip()
    catalog = _teacher_department_catalog()

    if code and code.lower() in catalog:
        item = catalog[code.lower()]
        resolved_name = item.get("name") or name or code
        return {"name": resolved_name, "code": item.get("code") or code}

    if name:
        for item in catalog.values():
            if (item.get("name") or "").strip().lower() == name.lower():
                return {"name": item.get("name") or name, "code": item.get("code") or code}

    return {"name": name or code, "code": code}


def _teacher_department_options():
    """Preferred department list for teacher self-service search."""
    rows = []
    seen = set()
    for teacher in AdminTeacher.objects.filter(is_active=True).order_by("department_name", "department_code", "name"):
        normalized = _normalize_teacher_department(teacher.department_name, teacher.department_code)
        name = normalized["name"]
        code = normalized["code"]
        key = (name.lower(), code.lower())
        if key in seen or (not name and not code):
            continue
        seen.add(key)
        rows.append({"name": name or code, "code": code})
    return rows or list(TEACHER_DEPARTMENT_LIST)


def _admin_teacher_has_account(admin_teacher):
    if not admin_teacher:
        return False

    filters = Q(linked_admin_teacher=admin_teacher)
    uid = (admin_teacher.uid or "").strip()
    name = (admin_teacher.name or "").strip()
    if uid and name:
        filters |= Q(linked_instructor__uid__iexact=uid, linked_instructor__name__iexact=name)
    return Profile.objects.filter(role__iexact="teacher").filter(filters).exists()


def _admin_teacher_option_payload(admin_teacher):
    normalized = _normalize_teacher_department(admin_teacher.department_name, admin_teacher.department_code)
    return {
        "id": admin_teacher.id,
        "name": admin_teacher.name,
        "uid": admin_teacher.uid or "",
        "email": admin_teacher.email or "",
        "department": normalized["name"],
        "department_code": normalized["code"],
        "has_account": _admin_teacher_has_account(admin_teacher),
    }


def _find_instructor_for_admin_teacher(admin_teacher, timetable_user=None):
    if not admin_teacher:
        return None

    qs = Instructor.objects.select_related("department")
    if timetable_user is not None:
        qs = qs.filter(user=timetable_user)

    uid = (admin_teacher.uid or "").strip()
    email = (admin_teacher.email or "").strip()
    name = (admin_teacher.name or "").strip()
    dept_name = (admin_teacher.department_name or "").strip().lower()
    dept_code = (admin_teacher.department_code or "").strip().lower()

    filters = Q()
    if uid:
        filters |= Q(uid__iexact=uid)
    if email:
        filters |= Q(email__iexact=email)
    if name:
        filters |= Q(name__iexact=name)
    if not filters:
        return None

    candidates = list(qs.filter(filters))
    if not candidates and timetable_user is not None:
        return _find_instructor_for_admin_teacher(admin_teacher, None)

    def _score(instr):
        points = 0
        if uid and (instr.uid or "").strip().lower() == uid.lower():
            points += 5
        if email and (instr.email or "").strip().lower() == email.lower():
            points += 4
        if name and (instr.name or "").strip().lower() == name.lower():
            points += 3
        instr_dept = getattr(instr, "department", None)
        if instr_dept:
            if dept_name and (instr_dept.name or "").strip().lower() == dept_name:
                points += 2
            if dept_code and (instr_dept.code or "").strip().lower() == dept_code:
                points += 2
        return (-points, instr.id)

    if not candidates:
        return None
    candidates.sort(key=_score)
    best = candidates[0]
    best_uid = bool(uid and (best.uid or "").strip().lower() == uid.lower())
    best_email = bool(email and (best.email or "").strip().lower() == email.lower())
    best_name = bool(name and (best.name or "").strip().lower() == name.lower())
    best_dept = False
    best_dept_obj = getattr(best, "department", None)
    if best_dept_obj:
        best_dept = bool(
            (dept_name and (best_dept_obj.name or "").strip().lower() == dept_name)
            or (dept_code and (best_dept_obj.code or "").strip().lower() == dept_code)
        )

    confident_match = (
        best_email
        or (best_uid and best_name)
        or (best_uid and best_dept)
        or (best_name and best_dept)
    )
    return best if confident_match else None


def _resolve_profile_instructor(profile, timetable_user=None):
    linked_admin_teacher = getattr(profile, "linked_admin_teacher", None)
    if linked_admin_teacher:
        return _find_instructor_for_admin_teacher(linked_admin_teacher, timetable_user)
    linked_instructor = getattr(profile, "linked_instructor", None)
    if linked_instructor and (timetable_user is None or linked_instructor.user_id == timetable_user.id):
        return linked_instructor
    return linked_instructor


def _teachers_in_department(dept_name, dept_code, query=""):
    """Central teacher directory search used by teacher self-registration."""
    qs = AdminTeacher.objects.filter(is_active=True)
    normalized = _normalize_teacher_department(dept_name, dept_code)
    canonical_name = (normalized.get("name") or "").strip()
    canonical_code = (normalized.get("code") or "").strip()
    filters = Q()
    if canonical_code:
        filters &= Q(department_code__iexact=canonical_code)
    elif canonical_name:
        filters &= Q(department_name__iexact=canonical_name)
    query = (query or "").strip()
    if query:
        filters &= (Q(name__icontains=query) | Q(email__icontains=query) | Q(uid__icontains=query))
    return [_admin_teacher_option_payload(teacher) for teacher in qs.filter(filters).order_by("name", "uid", "id")]


def _matching_instructor_ids(uid, name):
    """All instructor ids sharing the same teacher code + name (case-insensitive).
    Handles institutes seeded under multiple coordinator accounts."""
    return list(
        Instructor.objects.filter(uid__iexact=(uid or "").strip(), name__iexact=(name or "").strip())
        .values_list("id", flat=True)
    )


def _resolve_teacher_timetable(linked_instructor):
    """Find the best timetable + this teacher's table for the linked instructor.

    Searches across every instructor record that shares the same teacher code
    and name (duplicates across coordinator accounts). Prefers the most recent
    *published* timetable, falling back to the most recent saved one.
    Returns (timetable, my_table, workload, instructor_used) or (None, ...).
    """
    empty = (None, None, {"lectures": 0, "labs": 0, "shared_labs": 0, "total": 0}, None)
    if not linked_instructor:
        return empty

    candidate_ids = _matching_instructor_ids(linked_instructor.uid, linked_instructor.name)
    if not candidate_ids:
        candidate_ids = [linked_instructor.id]

    slot_qs = (
        ScheduledSlot.objects.filter(
            Q(instructor_id__in=candidate_ids) | Q(second_instructor_id__in=candidate_ids)
        )
        .select_related("timetable")
        .order_by("-timetable__is_published", "-timetable__created_at")
    )
    chosen_timetable = None
    for slot in slot_qs:
        if slot.timetable and slot.timetable.is_published:
            chosen_timetable = slot.timetable
            break
    if chosen_timetable is None:
        first = slot_qs.first()
        chosen_timetable = first.timetable if first else None

    if chosen_timetable is None:
        return empty

    classes, labs = _rebuild_classes_and_labs_from_saved(chosen_timetable)
    teacher_tables = build_teacher_tables(classes, labs, user=chosen_timetable.user)
    teacher_workloads = _compute_teacher_workloads(classes, labs)

    my_table = next(
        (t for t in teacher_tables if t["teacher"].id in candidate_ids),
        None,
    )
    instructor_used = my_table["teacher"] if my_table else linked_instructor
    workload = teacher_workloads.get(instructor_used, empty[2])
    return chosen_timetable, my_table, workload, instructor_used


def _teacher_schedule_rows(my_table):
    """Flatten a teacher table into a day/time ordered teaching-schedule list."""
    if not my_table:
        return []
    rows = []
    for row in my_table.get("rows", []):
        day = row.get("day")
        for cell in row.get("cells", []):
            cell_type = cell.get("type")
            slot_no = cell.get("slot_number")
            time_label = SLOT_LABELS.get(str(slot_no), "")
            if cell_type == "class":
                for cls in cell.get("classes", []):
                    rows.append({
                        "day": day,
                        "slot": slot_no,
                        "time": time_label,
                        "subject": getattr(cls.subject, "subject_name", "—"),
                        "section": cls.section,
                        "room": getattr(cls.room, "r_number", "—"),
                        "type": "Lecture",
                    })
            elif cell_type == "lab":
                for lab in cell.get("labs", []):
                    rows.append({
                        "day": day,
                        "slot": slot_no,
                        "time": time_label,
                        "subject": f"{getattr(lab.subject, 'subject_name', '—')} (Lab)",
                        "section": lab.section,
                        "room": getattr(lab.room, "r_number", "—"),
                        "type": "Lab",
                    })
    rows.sort(key=lambda r: (DAYS.index(r["day"]) if r["day"] in DAYS else 99, int(r["slot"]) if str(r["slot"]).isdigit() else 99))
    return rows


# --- Teacher registration OTP (email identity verification) -----------------
TEACHER_OTP_TTL_SECONDS = 300          # 5 minutes
TEACHER_OTP_MAX_ATTEMPTS = 5
TEACHER_OTP_SESSION_KEY = "teacher_reg_otp"


def _generate_teacher_otp():
    import secrets
    return f"{secrets.randbelow(10000):04d}"


def _store_teacher_otp(request, instructor, code):
    import time
    request.session[TEACHER_OTP_SESSION_KEY] = {
        "instructor_id": instructor.id,
        "code": code,
        "expires_at": time.time() + TEACHER_OTP_TTL_SECONDS,
        "attempts": 0,
        "verified": False,
        "email": instructor.email or "",
    }
    request.session.modified = True


def _get_teacher_otp(request):
    return request.session.get(TEACHER_OTP_SESSION_KEY)


def _save_teacher_otp(request, data):
    request.session[TEACHER_OTP_SESSION_KEY] = data
    request.session.modified = True


def _clear_teacher_otp(request):
    request.session.pop(TEACHER_OTP_SESSION_KEY, None)
    request.session.modified = True


def _teacher_otp_verified_for(request, instructor_id):
    data = _get_teacher_otp(request)
    return bool(
        data
        and data.get("verified")
        and str(data.get("instructor_id")) == str(instructor_id)
    )


def _mask_email(email):
    email = (email or "").strip()
    if "@" not in email:
        return email
    local, domain = email.split("@", 1)
    if len(local) <= 2:
        masked_local = local[:1] + "*"
    else:
        masked_local = local[0] + "*" * (len(local) - 2) + local[-1]
    return f"{masked_local}@{domain}"


def _send_teacher_otp_email(instructor, code):
    """Send the 4-digit verification code to the teacher's registered email."""
    from django.core.mail import EmailMultiAlternatives

    name = instructor.name or "Teacher"
    dept = getattr(instructor, "department", None)
    dept_name = getattr(dept, "name", "") or getattr(instructor, "department_name", "") or "your department"
    subject = "Verify Your SmartScheduler Account"

    text_body = (
        f"Dear {name},\n\n"
        "SmartScheduler by the SIH Winner Innovation Team\n\n"
        f"We received a request to create your SmartScheduler account "
        f"using your details from {dept_name}.\n\n"
        f"Your verification code is: {code}\n\n"
        f"This code will expire in 5 minutes.\n"
        f"Please do not share this code with anyone.\n\n"
        f"If you did not initiate this request, you can safely ignore this email.\n\n"
        f"Regards,\nSmartScheduler Team\n{SIH_WINNER_TEXT}"
    )

    html_body = f"""
    <div style="margin:0;padding:24px;background:#0b1020;font-family:Segoe UI,Roboto,Arial,sans-serif;">
      <div style="max-width:520px;margin:0 auto;background:#11162b;border:1px solid #232a45;border-radius:16px;overflow:hidden;">
        <div style="padding:22px 28px;background:linear-gradient(120deg,#6366f1,#ec4899);">
                    <table role="presentation" cellpadding="0" cellspacing="0"><tr>
                        <td style="vertical-align:middle;">
                            <div style="width:46px;height:46px;border-radius:50%;background:rgba(8,20,36,.55);text-align:center;line-height:46px;overflow:hidden;">
                                <img src="cid:{LOGO_CID}" width="34" height="41" alt="SmartScheduler" style="display:inline-block;width:34px;height:41px;vertical-align:middle;">
                            </div>
                        </td>
                        <td style="padding-left:14px;vertical-align:middle;">
                            <h1 style="margin:0;color:#fff;font-size:20px;font-weight:700;">SmartScheduler</h1>
                            <p style="margin:4px 0 0;color:rgba(255,255,255,.85);font-size:13px;">Account Verification</p>
                            <p style="margin:8px 0 0;color:#fff7ed;font-size:11px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;">SIH Winner Innovation Team</p>
                        </td>
                    </tr></table>
        </div>
        <div style="padding:28px;">
          <p style="color:#e2e8f0;font-size:15px;margin:0 0 14px;">Dear <strong>{name}</strong>,</p>
          <p style="color:#94a3b8;font-size:14px;line-height:1.6;margin:0 0 18px;">
            We received a request to create your SmartScheduler account using your
            details from <strong style="color:#cbd5e1;">{dept_name}</strong>.
            Use the verification code below to continue.
          </p>
          <div style="text-align:center;margin:24px 0;">
            <div style="display:inline-block;padding:16px 30px;background:#0b1020;border:1px solid #2b3358;border-radius:14px;">
              <span style="font-size:34px;letter-spacing:10px;font-weight:800;color:#a5b4fc;">{code}</span>
            </div>
          </div>
          <p style="color:#94a3b8;font-size:13px;line-height:1.6;margin:0 0 8px;">
            This code will expire in <strong style="color:#f9a8d4;">5 minutes</strong>.
          </p>
          <p style="color:#94a3b8;font-size:13px;line-height:1.6;margin:0 0 18px;">
            Please <strong>do not share</strong> this code with anyone. If you did not
            initiate this request, you can safely ignore this email.
          </p>
                      <p style="color:#64748b;font-size:13px;margin:18px 0 0;">Regards,<br>SmartScheduler Team<br><span style="color:#f59e0b;font-weight:700;">SIH Winner Innovation Team</span></p>
        </div>
      </div>
    </div>
    """

    sender = _brand_from_email()
    msg = EmailMultiAlternatives(subject, text_body, sender, [instructor.email])
    msg.attach_alternative(html_body, "text/html")
    _attach_email_logo(msg)
    msg.send(fail_silently=False)


def _attach_email_logo(msg):
    from email.mime.image import MIMEImage
    from pathlib import Path

    for base in settings.STATICFILES_DIRS or []:
        for fname in ("logo_email.png", "logo.jpeg"):
            logo_path = Path(base) / "img" / fname
            if logo_path.exists():
                try:
                    img = MIMEImage(logo_path.read_bytes())
                    img.add_header("Content-ID", f"<{LOGO_CID}>")
                    img.add_header("Content-Disposition", "inline", filename=fname)
                    msg.attach(img)
                except Exception:
                    logger.warning("Could not attach email logo", exc_info=True)
                return


def _saved_timetable_title(saved_t):
    dept_name = getattr(getattr(saved_t, "department", None), "name", "") or "All Departments"
    created = timezone.localtime(saved_t.created_at).strftime("%d %b %Y %I:%M %p") if saved_t.created_at else ""
    return f"{dept_name} timetable · {created}" if created else f"{dept_name} timetable"


def _collect_saved_teacher_recipients(saved_t):
    classes, labs = _rebuild_classes_and_labs_from_saved(saved_t)
    recipients = {}

    def _add(teacher):
        if not teacher:
            return
        email = (getattr(teacher, "email", "") or "").strip().lower()
        if not email or not EMAIL_RE.match(email):
            return
        teacher_id = getattr(teacher, "id", None)
        dept = getattr(getattr(teacher, "department", None), "name", "") or ""
        recipients[teacher_id] = {
            "id": teacher_id,
            "name": getattr(teacher, "name", "Teacher") or "Teacher",
            "email": email,
            "department": dept,
        }

    for cls in classes:
        _add(getattr(cls, "instructor", None))
        for co_teacher in getattr(cls, "co_instructors", []) or []:
            _add(co_teacher)
    for lab in labs:
        _add(getattr(lab, "instructor", None))
        _add(getattr(lab, "second_instructor", None))
        for co_teacher in getattr(lab, "co_instructors", []) or []:
            _add(co_teacher)

    return sorted(recipients.values(), key=lambda item: ((item.get("name") or "").lower(), item.get("email") or ""))


def _collect_admin_teacher_recipients(query=""):
    qs = AdminTeacher.objects.filter(is_active=True)
    query = (query or "").strip()
    if query:
        qs = qs.filter(Q(name__icontains=query) | Q(email__icontains=query) | Q(uid__icontains=query))

    recipients = []
    seen = set()
    for teacher in qs.order_by("name", "uid", "id"):
        email = (teacher.email or "").strip().lower()
        if not email or not EMAIL_RE.match(email):
            continue
        key = ((teacher.name or "").strip().lower(), (teacher.uid or "").strip().lower(), email)
        if key in seen:
            continue
        seen.add(key)
        recipients.append({
            "id": f"admin:{teacher.id}",
            "name": teacher.name or "Teacher",
            "email": email,
            "department": teacher.department_name or "",
            "uid": teacher.uid or "",
        })
    return recipients


def _coordinator_recipients():
        qs = CoordinatorAppointment.objects.filter(role__icontains="Timetable Coordinator").order_by("role", "name")
        recipients = []
        seen = set()
        for appointment in qs:
                email = (appointment.email or "").strip().lower()
                if not email or email in seen or not EMAIL_RE.match(email):
                        continue
                seen.add(email)
                recipients.append({
                        "name": appointment.name or "Coordinator",
                        "email": email,
                        "role": appointment.role,
                        "department": appointment.department,
                })
        return recipients


def _published_timetable_email_content(*, recipient_name, publish_code, login_url, timetable_title, role_label, note=""):
        safe_name = recipient_name or "Faculty Member"
        safe_note_text = (note or "").strip()
        safe_note_html = safe_note_text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br>")
        subject = f"Published timetable access code {publish_code} — SmartScheduler"
        text_body = (
                f"Hello {safe_name},\n\n"
                "J.C. Bose University of Science and Technology, YMCA (Formerly YMCA UST)\n"
                f"{SIH_WINNER_TEXT}\n\n"
                f"A timetable has been published for you on SmartScheduler.\n"
                f"Timetable: {timetable_title}\n"
                f"Role: {role_label}\n"
                f"Publish code: {publish_code}\n\n"
                f"Use this code here: {login_url}\n"
                "This code uniquely identifies the timetable you should connect to.\n"
                + (f"\nNote:\n{safe_note_text}\n" if safe_note_text else "")
                + "\nRegards,\nSmartScheduler Team\n"
        )
        note_block = ""
        if safe_note_text:
                note_block = (
                        '<tr><td style="padding:18px 28px 0;">'
                        '<div style="background:#0f1a2c;border:1px solid rgba(148,163,184,.18);border-radius:12px;padding:16px 18px;color:#cbd5e1;font-size:14px;line-height:1.6;">'
                        '<div style="color:#64748b;font-size:11px;letter-spacing:.08em;text-transform:uppercase;margin-bottom:8px;">Additional note</div>'
                        f'{safe_note_html}</div></td></tr>'
                )
        html_body = f"""
<!DOCTYPE html><html><body style="margin:0;background:#0b1220;font-family:'Segoe UI',system-ui,-apple-system,sans-serif;">
    <table role="presentation" width="100%" cellpadding="0" cellspacing="0" style="background:#0b1220;padding:28px 12px;">
        <tr><td align="center">
            <table role="presentation" width="580" cellpadding="0" cellspacing="0" style="max-width:580px;width:100%;background:#111c30;border:1px solid rgba(148,163,184,.16);border-radius:18px;overflow:hidden;">
                <tr><td style="background:linear-gradient(135deg,#38bdf8,#0ea5e9 55%,#0c3557);padding:26px 28px;">
                    <table role="presentation" cellpadding="0" cellspacing="0"><tr>
                        <td style="vertical-align:middle;">
                            <div style="width:46px;height:46px;border-radius:13px;background:rgba(8,20,36,.55);text-align:center;line-height:46px;">
                                <img src="cid:{LOGO_CID}" width="34" height="41" alt="SmartScheduler" style="display:inline-block;width:34px;height:41px;vertical-align:middle;">
                            </div>
                        </td>
                        <td style="padding-left:14px;vertical-align:middle;">
                            <div style="color:#ffffff;font-size:19px;font-weight:700;">SmartScheduler</div>
                            <div style="color:rgba(255,255,255,.82);font-size:12px;">Published Timetable Access</div>
                            <div style="color:#fff7ed;font-size:11px;font-weight:700;letter-spacing:.08em;text-transform:uppercase;margin-top:6px;">SIH Winner Innovation Team</div>
                        </td>
                    </tr></table>
                </td></tr>
                <tr><td style="background:#ffffff;padding:14px 28px;border-bottom:1px solid rgba(148,163,184,.16);">
                    <table role="presentation" cellpadding="0" cellspacing="0"><tr>
                        <td style="vertical-align:middle;">
                            <img src="https://upload.wikimedia.org/wikipedia/en/a/ae/J.C._Bose_University_of_Science_and_Technology%2C_YMCA_logo.png" width="40" height="40" alt="J.C. Bose University" style="display:block;width:40px;height:40px;object-fit:contain;">
                        </td>
                        <td style="padding-left:12px;vertical-align:middle;">
                            <div style="color:#e11d2f;font-size:13.5px;font-weight:700;line-height:1.35;">J.C. Bose University of Science and Technology, YMCA (Formerly YMCA UST)</div>
                        </td>
                    </tr></table>
                </td></tr>
                <tr><td style="padding:26px 28px 4px;">
                    <div style="color:#e2e8f0;font-size:17px;font-weight:600;margin-bottom:6px;">Hello {safe_name},</div>
                    <div style="color:#94a3b8;font-size:14px;line-height:1.65;">A timetable has been published for you on SmartScheduler. Use the access code below to connect to the right timetable.</div>
                </td></tr>
                <tr><td style="padding:18px 28px 0;">
                    <div style="background:linear-gradient(160deg,#0f1a2c,#0d1626);border:1px solid rgba(56,189,248,.28);border-radius:12px;padding:18px 20px;">
                        <div style="color:#64748b;font-size:11px;letter-spacing:.08em;text-transform:uppercase;">Publish code</div>
                        <div style="color:#38bdf8;font-size:26px;font-weight:800;letter-spacing:.18em;margin-top:6px;">{publish_code}</div>
                        <div style="color:#94a3b8;font-size:13px;margin-top:10px;">Timetable · {timetable_title}<br>Recipient type · {role_label}</div>
                    </div>
                </td></tr>
                <tr><td style="padding:18px 28px 0;">
                    <div style="background:#0f1a2c;border:1px solid rgba(148,163,184,.18);border-radius:12px;padding:16px 18px;color:#cbd5e1;font-size:14px;line-height:1.7;">
                        Open <a href="{login_url}" style="color:#38bdf8;">{login_url}</a> and enter this publish code. The code uniquely identifies this timetable.
                    </div>
                </td></tr>
                {note_block}
                <tr><td style="padding:22px 28px 28px;">
                    <div style="border-top:1px solid rgba(148,163,184,.16);padding-top:16px;color:#64748b;font-size:12px;line-height:1.6;">This is an automated message from SmartScheduler.<br><span style="color:#f59e0b;font-weight:700;">SIH Winner Innovation Team</span></div>
                </td></tr>
            </table>
        </td></tr>
    </table>
</body></html>"""
        return subject, text_body, html_body


def _send_publish_notification_email(*, recipient_name, recipient_email, publish_code, request, role_label, timetable_title):
        from django.core.mail import EmailMultiAlternatives
        from email.utils import formataddr

        login_url = request.build_absolute_uri(reverse("teacher_published_timetable"))
        subject, text_body, html_body = _published_timetable_email_content(
                recipient_name=recipient_name,
                publish_code=publish_code,
                login_url=login_url,
                timetable_title=timetable_title,
                role_label=role_label,
        )
        sender = _brand_from_email()
        msg = EmailMultiAlternatives(
                subject=subject,
                body=text_body,
                from_email=sender,
                to=[formataddr((recipient_name, recipient_email))],
        )
        msg.attach_alternative(html_body, "text/html")
        _attach_email_logo(msg)
        msg.send(fail_silently=False)


def _send_publish_summary_email(request, saved_t, action_label, sent_count, failed_count):
        from django.core.mail import EmailMultiAlternatives

        owner_email = (getattr(request.user, "email", "") or "").strip()
        if not owner_email or not EMAIL_RE.match(owner_email):
                return
        subject, text_body, html_body = _published_timetable_email_content(
                recipient_name=getattr(request.user, "get_full_name", lambda: "")() or getattr(request.user, "username", "Owner"),
                publish_code=saved_t.publish_code,
                login_url=request.build_absolute_uri(reverse("saved_timetable_publish_notifications", args=[saved_t.id])),
                timetable_title=_saved_timetable_title(saved_t),
                role_label="Timetable Owner",
                note=f"{action_label}\nSuccessful emails: {sent_count}\nFailed emails: {failed_count}",
        )
        sender = _brand_from_email()
        msg = EmailMultiAlternatives(subject, text_body, sender, [owner_email])
        msg.attach_alternative(html_body, "text/html")
        _attach_email_logo(msg)
        msg.send(fail_silently=True)


def teacher_register(request):
    """Self-registration: pick department, search name, verify by email OTP, set password."""
    from django.contrib.auth import get_user_model, login as auth_login

    if request.user.is_authenticated:
        profile, _ = Profile.objects.get_or_create(user=request.user)
        if (profile.role or "").lower() == "teacher" and (profile.linked_instructor_id or getattr(profile, "linked_admin_teacher_id", None)):
            return redirect("teacher_dashboard")

    dept_options = _teacher_department_options()

    if request.method == "POST":
        teacher_id = (request.POST.get("teacher_id") or request.POST.get("instructor_id") or "").strip()
        password = request.POST.get("password") or ""
        confirm = request.POST.get("confirm_password") or ""

        teacher = AdminTeacher.objects.filter(id=teacher_id, is_active=True).first() if teacher_id.isdigit() else None

        errors = []
        if teacher is None:
            errors.append("Please select your name from the department list.")
        elif not _teacher_otp_verified_for(request, teacher.id):
            errors.append("Please verify your identity with the email code before creating your account.")
        if len(password) < 5:
            errors.append("Password must be at least 5 characters.")
        if password != confirm:
            errors.append("Passwords do not match.")
        if teacher is not None and _admin_teacher_has_account(teacher):
            errors.append("An account already exists for this teacher. Please log in instead.")

        if errors:
            for err in errors:
                messages.error(request, err)
        else:
            User = get_user_model()
            username = f"teacher_{teacher.id}"
            base_username = username
            suffix = 1
            while User.objects.filter(username=username).exists():
                suffix += 1
                username = f"{base_username}_{suffix}"

            new_user = User.objects.create_user(
                username=username,
                email=teacher.email or "",
                password=password,
            )
            if teacher.name:
                new_user.first_name = teacher.name[:30]
                new_user.save(update_fields=["first_name"])

            profile, _ = Profile.objects.get_or_create(user=new_user)
            profile.role = "teacher"
            profile.linked_admin_teacher = teacher
            profile.linked_instructor = _find_instructor_for_admin_teacher(teacher)
            profile.save(update_fields=["role", "linked_admin_teacher", "linked_instructor"])

            _clear_teacher_otp(request)
            auth_login(request, new_user, backend="django.contrib.auth.backends.ModelBackend")
            messages.success(request, "Account created. Welcome to your teacher dashboard.")
            return redirect("teacher_dashboard")

    context = {
        "dept_options": dept_options,
    }
    context.update(_teacher_nav_context())
    return render(request, "teacher_register.html", context)


def teacher_register_teachers(request):
    """AJAX: teachers in a department (by name+code) matching a search query."""
    dept_name = (request.GET.get("name") or "").strip()
    dept_code = (request.GET.get("code") or "").strip()
    query = (request.GET.get("q") or "").strip()
    teachers = _teachers_in_department(dept_name, dept_code, query)
    return JsonResponse({"teachers": teachers})


def teacher_register_info(request):
    """AJAX: teacher code + email for a selected instructor id."""
    teacher_id = (request.GET.get("id") or "").strip()
    if not teacher_id.isdigit():
        return JsonResponse({"ok": False}, status=400)
    teacher = AdminTeacher.objects.filter(id=teacher_id, is_active=True).first()
    if teacher is None:
        return JsonResponse({"ok": False}, status=404)
    return JsonResponse({
        "ok": True,
        "uid": teacher.uid,
        "email": teacher.email or "",
        "name": teacher.name,
        "department": teacher.department_name or "",
        "has_account": _admin_teacher_has_account(teacher),
    })


def teacher_register_send_otp(request):
    """AJAX (POST): generate a 4-digit OTP and email it to the teacher."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Invalid request method."}, status=405)

    teacher_id = (request.POST.get("teacher_id") or request.POST.get("instructor_id") or "").strip()
    teacher = AdminTeacher.objects.filter(id=teacher_id, is_active=True).first() if teacher_id.isdigit() else None
    if teacher is None:
        return JsonResponse({"ok": False, "error": "Please select your profile first."}, status=404)
    if _admin_teacher_has_account(teacher):
        return JsonResponse({"ok": False, "error": "An account already exists for this teacher. Please log in instead."}, status=400)
    if not (teacher.uid or "").strip():
        return JsonResponse({"ok": False, "error": "Teacher code is missing for this profile. Please contact your coordinator."}, status=400)
    if not (teacher.email or "").strip():
        return JsonResponse({"ok": False, "error": "No email is registered for this teacher. Please contact your coordinator."}, status=400)

    code = _generate_teacher_otp()
    _store_teacher_otp(request, teacher, code)
    try:
        _send_teacher_otp_email(teacher, code)
    except Exception:
        logger.exception("Teacher OTP email failed")
        _clear_teacher_otp(request)
        return JsonResponse({"ok": False, "error": "We could not send the verification email. Please try again."}, status=502)

    return JsonResponse({
        "ok": True,
        "masked_email": _mask_email(teacher.email),
        "expires_in": TEACHER_OTP_TTL_SECONDS,
    })


def teacher_register_verify_otp(request):
    """AJAX (POST): verify the 4-digit OTP for the selected teacher."""
    import time

    if request.method != "POST":
        return JsonResponse({"ok": False, "error": "Invalid request method."}, status=405)

    instructor_id = (request.POST.get("instructor_id") or "").strip()
    otp = (request.POST.get("otp") or "").strip()
    data = _get_teacher_otp(request)

    if not data or str(data.get("instructor_id")) != str(instructor_id):
        return JsonResponse({"ok": False, "error": "Please request a verification code first."}, status=400)

    if time.time() > data.get("expires_at", 0):
        _clear_teacher_otp(request)
        return JsonResponse({"ok": False, "error": "Your code has expired. Please resend a new code.", "expired": True}, status=400)

    if data.get("attempts", 0) >= TEACHER_OTP_MAX_ATTEMPTS:
        _clear_teacher_otp(request)
        return JsonResponse({"ok": False, "error": "Too many incorrect attempts. Please resend a new code.", "locked": True}, status=429)

    if not re.fullmatch(r"\d{4}", otp):
        data["attempts"] = data.get("attempts", 0) + 1
        _save_teacher_otp(request, data)
        return JsonResponse({"ok": False, "error": "Please enter the 4-digit code."}, status=400)

    if otp != data.get("code"):
        data["attempts"] = data.get("attempts", 0) + 1
        left = TEACHER_OTP_MAX_ATTEMPTS - data["attempts"]
        if left <= 0:
            _clear_teacher_otp(request)
            return JsonResponse({"ok": False, "error": "Too many incorrect attempts. Please resend a new code.", "locked": True}, status=429)
        _save_teacher_otp(request, data)
        return JsonResponse({"ok": False, "error": f"Incorrect code. {left} attempt(s) left.", "attempts_left": left}, status=400)

    data["verified"] = True
    _save_teacher_otp(request, data)
    return JsonResponse({"ok": True})


@login_required
def teacher_onboarding(request):
    return redirect("teacher_dashboard")


@login_required
def teacher_dashboard(request):
    profile, locked_response = _get_teacher_profile_or_locked_response(request.user)
    if locked_response:
        return render(request, 'role_locked.html', {'current_role': locked_response})

    _ensure_teacher_role(profile)

    if not (profile.linked_instructor_id or getattr(profile, "linked_admin_teacher_id", None)):
        messages.error(request, "Please register your teacher account to continue.")
        return redirect("teacher_register")
    context = _resolve_teacher_dashboard_context(request, profile)
    return render(request, "teacher_dashboard.html", context)



@login_required
def teacher_profile_page(request):
    profile, locked_response = _get_teacher_profile_or_locked_response(request.user)
    if locked_response:
        return render(request, 'role_locked.html', {'current_role': locked_response})

    _ensure_teacher_role(profile)

    redirect_response = _teacher_onboarding_redirect_response(request)
    if redirect_response:
        return redirect_response

    if request.method == "POST":
        contact_number = request.POST.get("contact_number", "").strip()
        faculty_uid = request.POST.get("faculty_uid", "").strip()
        profile.contact_number = contact_number
        update_fields = ["contact_number"]

        if faculty_uid:
            active_timetable = _get_active_teacher_timetable(profile)
            if not active_timetable:
                messages.error(request, "Connect the HOD published timetable before linking your faculty UID.")
            else:
                instructor = Instructor.objects.filter(
                    user=active_timetable.user,
                    uid__iexact=faculty_uid,
                ).first()
                if instructor is None:
                    messages.error(request, "No teacher record matched that faculty UID in the published timetable.")
                else:
                    profile.linked_instructor = instructor
                    update_fields.append("linked_instructor")
                    messages.success(request, "Your faculty profile is linked and ready for the teacher timetable page.")
        elif request.POST.get("clear_faculty_link") == "1":
            profile.linked_instructor = None
            update_fields.append("linked_instructor")
            messages.success(request, "Faculty link removed from your teacher profile.")
        else:
            messages.success(request, "Your teacher profile details were updated.")

        profile.save(update_fields=list(dict.fromkeys(update_fields)))
        return redirect("teacher_profile_page")

    context = _resolve_teacher_dashboard_context(request, profile)
    return render(request, "teacher_profile_page.html", context)


@login_required
def teacher_published_timetable(request):
    profile, locked_response = _get_teacher_profile_or_locked_response(request.user)
    if locked_response:
        return render(request, 'role_locked.html', {'current_role': locked_response})

    _ensure_teacher_role(profile)

    redirect_response = _teacher_onboarding_redirect_response(request)
    if redirect_response:
        return redirect_response

    if request.method == "POST":
        code = request.POST.get("access_code", "").strip()
        if not code:
            messages.error(request, "Please enter the publish code shared by your HOD.")
        else:
            timetable, error_message = _connect_teacher_timetable(profile, code)
            if error_message:
                messages.error(request, error_message)
            else:
                messages.success(
                    request,
                    f"HOD published timetable connected with code {timetable.publish_code}.",
                )
        return redirect("teacher_published_timetable")

    context = _resolve_teacher_dashboard_context(request, profile)
    return render(request, "teacher_published_timetable.html", context)


@login_required
def teacher_my_timetable(request):
    profile, locked_response = _get_teacher_profile_or_locked_response(request.user)
    if locked_response:
        return render(request, 'role_locked.html', {'current_role': locked_response})

    _ensure_teacher_role(profile)

    redirect_response = _teacher_onboarding_redirect_response(request)
    if redirect_response:
        return redirect_response

    context = _resolve_teacher_dashboard_context(request, profile)
    return render(request, "teacher_my_timetable.html", context)

# CONTACT FORM
def contact(request):
    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        email   = request.POST.get('email', '').strip()
        subject = request.POST.get('subject', 'No subject').strip()
        message = request.POST.get('message', '').strip()

        body = (
            f"New contact form submission from SmartScheduler\n"
            f"{'-' * 44}\n"
            f"Name    : {name}\n"
            f"Email   : {email}\n"
            f"Subject : {subject}\n"
            f"{'-' * 44}\n\n"
            f"{message}\n"
        )
        try:
            msg = EmailMessage(
                subject=f"[SmartScheduler] {subject} — from {name}",
                body=body,
                from_email=_brand_from_email(),
                to=['smartschedulertech@gmail.com'],
                reply_to=[f"{name} <{email}>"],         # Reply goes to the visitor
            )
            msg.send(fail_silently=False)
            messages.success(request, "Message sent! We'll get back to you soon.")
        except Exception:
            messages.error(request, "Couldn't send your message right now. Please try again later.")
        return redirect('contact')
    return render(request, 'contact.html')


def institute_application(request):
    if request.method == "POST":
        institute_type = request.POST.get("institute_type", "").strip()
        other_type = request.POST.get("other_type", "").strip()
        contact_name = request.POST.get("contact_name", "").strip()
        official_email = request.POST.get("official_email", "").strip()
        contact_number = request.POST.get("contact_number", "").strip()
        note = request.POST.get("note", "").strip()

        selected_type = other_type if institute_type == "Other" and other_type else institute_type

        if not settings.EMAIL_HOST_USER or not settings.EMAIL_HOST_PASSWORD:
            messages.error(
                request,
                "Email is not configured yet. Add EMAIL_HOST_USER and EMAIL_HOST_PASSWORD in your .env file, then restart the server.",
            )
            return render(request, "institute_application.html")

        body = (
            "New institute customization application\n"
            "--------------------------------------\n"
            f"Institute type : {selected_type}\n"
            f"Contact person : {contact_name}\n"
            f"Official email : {official_email}\n"
            f"Contact number : {contact_number}\n"
            "--------------------------------------\n\n"
            f"Additional note:\n{note or 'No note added.'}\n"
        )

        try:
            msg = EmailMessage(
                subject=f"[SmartScheduler] Institute application - {selected_type or 'New request'}",
                body=body,
                from_email=_brand_from_email(),
                to=["smartschedulertech@gmail.com"],
                reply_to=[f"{contact_name} <{official_email}>"] if official_email else None,
            )
            msg.send(fail_silently=False)
            return redirect("institute_application_thanks")
        except SMTPAuthenticationError:
            logger.exception("Institute application email authentication failed")
            messages.error(
                request,
                "Gmail rejected the sender email/password. Use the exact Gmail address in EMAIL_HOST_USER and a valid 16-character Gmail App Password in EMAIL_HOST_PASSWORD.",
            )
        except Exception as exc:
            logger.exception("Institute application email failed")
            error_message = "Could not send your application right now. Please check the SMTP settings and try again."
            if settings.DEBUG:
                error_message = f"{error_message} Server said: {exc}"
            messages.error(request, error_message)

    return render(request, "institute_application.html")


def institute_application_thanks(request):
    return render(request, "institute_application_thanks.html")


# ADMIN DASHBOARD
@login_required
def admindash(request):
    context = {
        'teacher_count': Instructor.objects.filter(user=request.user).count(),
        'department_count': Department.objects.filter(user=request.user).count(),
        'class_count': Section.objects.filter(user=request.user).count(),
        'teacher_onboarding_count': TeacherOnboarding.objects.count(),
    }
    return render(request, 'admindashboard.html', context)


@login_required
def teacher_onboarding_responses_page(request):
    if not _user_can_manage_teacher_onboarding(request.user):
        return HttpResponseForbidden("You do not have permission to view teacher onboarding submissions.")

    submissions = []
    for submission in TeacherOnboarding.objects.select_related("user").all():
        submissions.append({
            "id": submission.id,
            "full_name": submission.full_name,
            "username": submission.user.username,
            "designation": submission.designation,
            "joining_year": submission.joining_year,
            "email": submission.email,
            "subjects_taught": submission.subjects_taught,
            "submitted_at": _format_local_datetime(submission.submitted_at, "%d %b %Y, %I:%M %p"),
            "requires_resubmission": submission.requires_resubmission,
            "delete_url": reverse("delete_teacher_onboarding", args=[submission.id]),
            "resubmit_url": reverse("request_teacher_onboarding_resubmission", args=[submission.id]),
        })

    return render(
        request,
        "teacher_onboarding_responses.html",
        {
            "submissions_json": json.dumps(submissions),
            "total": len(submissions),
        },
    )


@login_required
def request_teacher_onboarding_resubmission(request, submission_id):
    if request.method != "POST":
        return HttpResponseForbidden("POST request required.")
    if not _user_can_manage_teacher_onboarding(request.user):
        return HttpResponseForbidden("You do not have permission to request teacher form resubmission.")

    submission = TeacherOnboarding.objects.filter(id=submission_id).first()
    if submission is None:
        messages.error(request, "Teacher form submission was not found.")
        return redirect("teacher_onboarding_responses")

    submission.requires_resubmission = True
    submission.resubmission_requested_at = timezone.now()
    submission.save(update_fields=["requires_resubmission", "resubmission_requested_at", "updated_at"])
    messages.success(request, f"Resubmission was requested for {submission.full_name}.")
    return redirect("teacher_onboarding_responses")


@login_required
def delete_teacher_onboarding(request, submission_id):
    if request.method != "POST":
        return HttpResponseForbidden("POST request required.")
    if not _user_can_manage_teacher_onboarding(request.user):
        return HttpResponseForbidden("You do not have permission to delete teacher form submissions.")

    submission = TeacherOnboarding.objects.filter(id=submission_id).first()
    if submission is None:
        messages.error(request, "Teacher form submission was not found.")
    else:
        full_name = submission.full_name
        submission.delete()
        messages.success(request, f"Teacher form for {full_name} was deleted.")
    return redirect("teacher_onboarding_responses")


@login_required
def export_teacher_onboarding_csv(request):
    if not _user_can_manage_teacher_onboarding(request.user):
        return HttpResponseForbidden("You do not have permission to export teacher onboarding submissions.")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="teacher_onboarding_submissions.csv"'
    writer = csv.writer(response)
    writer.writerow(["Username", "Full Name", "Designation", "Joining Year", "Email", "Subjects Taught", "Submitted"])
    for submission in TeacherOnboarding.objects.select_related("user").all():
        writer.writerow([
            submission.user.username,
            submission.full_name,
            submission.designation,
            submission.joining_year,
            submission.email,
            submission.subjects_taught,
            _format_local_datetime(submission.submitted_at, "%d %b %Y %H:%M"),
        ])
    return response


# Helper to reset GA cache when admin modifies models
def reset_global_schedule_cache(user_id=None):
    global data, GLOBAL_GENERATED_SCHEDULES, GLOBAL_CLASSES, GLOBAL_LABS
    if user_id is not None:
        state = _get_user_state(user_id)
        state["data"] = None
        state["schedules"] = []
        state["classes"] = None
        state["labs"] = None
    # Also clear legacy globals
    data = None
    GLOBAL_GENERATED_SCHEDULES = []
    GLOBAL_CLASSES = None
    GLOBAL_LABS = None
    # Clear the algorithm data cache for this user
    if "reset_user_data_cache" in globals() and user_id is not None:
        reset_user_data_cache(user_id)


def _runtime_unavailable_response(request):
    messages.error(request, "You are not able to access this feature right now. Please contact your administrator.")
    return redirect("generate")


if "SLOT_LABELS" not in globals():
    SLOT_LABELS = {
        "1": "9:00 - 9:55",
        "2": "9:55 - 10:50",
        "3": "10:50 - 11:45",
        "4": "11:45 - 12:40",
        "5": "12:40 - 1:35",
        "6": "1:35 - 2:30",
        "7": "2:30 - 3:25",
        "8": "3:25 - 4:20",
        "9": "4:20 - 5:15",
    }


if "DAYS" not in globals():
    DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]


if "Class" not in globals():
    class Class:
        def __init__(self, id, dept, section, subject):
            self.section_id = id
            self.department = dept
            self.subject = subject
            self.instructor = None
            self.co_instructors = []
            self.meeting_time = None
            self.meeting_times = []
            self.room = None
            self.section = section
            self.duration = getattr(subject, 'duration', 1)
            self.room_label = ""
            self.missing_room = False

        def set_instructor(self, instructor): self.instructor = instructor
        def set_co_instructors(self, instructors): self.co_instructors = list(instructors or [])
        def set_meetingTime(self, mt): self.meeting_time = mt
        def set_room(self, room): self.room = room


if "Lab" not in globals():
    class Lab:
        def __init__(self, id, dept, section, subject, batch=1, total_batches=1):
            self.section_id = id
            self.department = dept
            self.subject = subject
            self.instructor = None
            self.second_instructor = None   # shared lab: optional second teacher
            self.co_instructors = []
            self.room = None
            self.section = section
            self.duration = getattr(subject, 'duration', LAB_DURATION)
            self.meeting_times = []
            self.batch = batch
            self.total_batches = total_batches
            self.room_label = ""
            self.missing_room = False

        def set_second_instructor(self, instructor): self.second_instructor = instructor
        def set_co_instructors(self, instructors): self.co_instructors = list(instructors or [])
        def set_instructor(self, instructor): self.instructor = instructor
        def set_meetingTimes(self, mts): self.meeting_times = mts
        def set_room(self, room): self.room = room


if "build_section_tables" not in globals():
    def build_section_tables(all_classes, all_labs, user=None):
        from collections import OrderedDict

        section_map = OrderedDict()
        section_qs = Section.objects.all()
        if user is not None:
            section_qs = section_qs.filter(user=user)
        section_qs = section_qs.select_related("department").prefetch_related("allowed_subjects__instructors")

        for section in section_qs:
            section_map.setdefault(
                section.section_id,
                {"classes": [], "labs": [], "section": section, "dept": section.department},
            )

        for cls in all_classes:
            sec_key = cls.section
            if sec_key not in section_map:
                section_map[sec_key] = {"classes": [], "labs": [], "section": None, "dept": cls.department}
            section_map[sec_key]["classes"].append(cls)

        for lab in all_labs:
            sec_key = lab.section
            if sec_key not in section_map:
                section_map[sec_key] = {"classes": [], "labs": [], "section": None, "dept": lab.department}
            section_map[sec_key]["labs"].append(lab)

        lecture_rooms_qs = Room.objects.exclude(room_type="Lab")
        lab_rooms_qs = Room.objects.filter(room_type="Lab")
        if user is not None:
            lecture_rooms_qs = lecture_rooms_qs.filter(user=user)
            lab_rooms_qs = lab_rooms_qs.filter(user=user)
        lecture_rooms = list(lecture_rooms_qs.select_related("department"))
        lab_rooms = list(lab_rooms_qs.select_related("department"))
        all_rooms = lecture_rooms + lab_rooms

        def _slots_for_class(cls):
            if getattr(cls, "meeting_times", None):
                return list(cls.meeting_times)
            if getattr(cls, "meeting_time", None):
                return [cls.meeting_time]
            return []

        def _register_busy(container, key, slots):
            busy = container.setdefault(key, set())
            for mt in slots:
                busy.add((mt.day, mt.time))

        room_busy = {}
        section_busy = {}
        instructor_busy = {}

        for cls in all_classes:
            slots = _slots_for_class(cls)
            if cls.room:
                _register_busy(room_busy, cls.room.r_number, slots)
            _register_busy(section_busy, cls.section, slots)
            if cls.instructor:
                _register_busy(instructor_busy, cls.instructor.pk, slots)

        for lab in all_labs:
            slots = list(getattr(lab, "meeting_times", []) or [])
            if lab.room:
                _register_busy(room_busy, lab.room.r_number, slots)
            _register_busy(section_busy, lab.section, slots)
            if lab.instructor:
                _register_busy(instructor_busy, lab.instructor.pk, slots)
            second_instructor = getattr(lab, "second_instructor", None)
            if second_instructor:
                _register_busy(instructor_busy, second_instructor.pk, slots)

        def _iter_candidate_blocks(duration):
            valid_slots = [slot for slot in range(1, 10) if slot != int(LUNCH_SLOT)]
            for day in DAYS:
                for slot in valid_slots:
                    block = [str(slot + offset) for offset in range(duration)]
                    if any(int(value) == int(LUNCH_SLOT) for value in block):
                        continue
                    if any(int(value) not in valid_slots for value in block):
                        continue
                    yield day, block

        def _build_suggested_slots(section_obj, subject):
            if subject.room_required == "Lab":
                return []

            duration = max(1, getattr(subject, "duration", 1) or 1)
            instructors = list(subject.instructors.all())
            if not instructors or section_obj is None:
                return []

            candidate_rooms = [
                room for room in lecture_rooms
                if room.department == section_obj.department
            ] or lecture_rooms

            suggestions = []
            seen = set()
            for day, block in _iter_candidate_blocks(duration):
                if any((day, slot) in section_busy.get(section_obj.section_id, set()) for slot in block):
                    continue
                for room in candidate_rooms:
                    if any((day, slot) in room_busy.get(room.r_number, set()) for slot in block):
                        continue
                    available_teacher = next(
                        (
                            teacher for teacher in instructors
                            if not any((day, slot) in instructor_busy.get(teacher.pk, set()) for slot in block)
                        ),
                        None,
                    )
                    if available_teacher is None:
                        continue
                    key = (day, block[0], room.r_number)
                    if key in seen:
                        continue
                    seen.add(key)
                    suggestions.append(
                        {
                            "day": day,
                            "slot_label": SLOT_LABELS.get(block[0], f"Slot {block[0]}"),
                            "room": room.r_number,
                        }
                    )
                    break
                if len(suggestions) >= 3:
                    break
            return suggestions

        tables = []
        for sec_id, data in section_map.items():
            grid = {day: {} for day in DAYS}
            for cls in data["classes"]:
                if cls.meeting_time:
                    day = cls.meeting_time.day
                    slot = int(cls.meeting_time.time)
                    if day in grid:
                        grid[day].setdefault(slot, {"classes": [], "labs": []})
                        grid[day][slot]["classes"].append(cls)
            for lab in data["labs"]:
                if lab.meeting_times:
                    first_mt = lab.meeting_times[0]
                    day = first_mt.day
                    slot = int(first_mt.time)
                    if day in grid:
                        grid[day].setdefault(slot, {"classes": [], "labs": []})
                        grid[day][slot]["labs"].append(lab)

            rows = []
            for day in DAYS:
                cells = []
                skip_until = 0
                for s in range(1, 10):
                    if s <= skip_until:
                        continue
                    if s == 5:
                        cells.append({"type": "lunch", "colspan": 1, "slot_number": s})
                        continue
                    cell_data = grid[day].get(s, {"classes": [], "labs": []})
                    if cell_data["labs"]:
                        lab_span = max(
                            (len(lb.meeting_times) for lb in cell_data["labs"] if lb.meeting_times),
                            default=getattr(cell_data["labs"][0], "duration", LAB_DURATION),
                        )
                        cells.append({
                            "type": "lab",
                            "colspan": lab_span,
                            "slot_number": s,
                            "labs": cell_data["labs"],
                        })
                        skip_until = s + lab_span - 1
                    elif cell_data["classes"]:
                        cls_dur = getattr(cell_data["classes"][0], "duration", 1)
                        cells.append({
                            "type": "class",
                            "colspan": cls_dur,
                            "slot_number": s,
                            "classes": cell_data["classes"],
                        })
                        if cls_dur > 1:
                            skip_until = s + cls_dur - 1
                    else:
                        cells.append({"type": "empty", "colspan": 1, "slot_number": s})
                rows.append({"day": day, "cells": cells})

            section_obj = data.get("section")

            if section_obj is None:
                class _SectionProxy:
                    def __init__(self, sid, dept):
                        self.section_id = sid
                        self.department = dept
                section_obj = _SectionProxy(sec_id, data["dept"])
                allowed_subjects = []
            else:
                allowed_subjects = list(section_obj.allowed_subjects.all())

            class_counts = {}
            for cls in data["classes"]:
                subject_pk = getattr(cls.subject, "pk", None)
                class_counts[subject_pk] = class_counts.get(subject_pk, 0) + 1

            lab_counts = {}
            for lab in data["labs"]:
                subject_pk = getattr(lab.subject, "pk", None)
                if not lab.meeting_times:
                    continue
                first_mt = lab.meeting_times[0]
                lab_counts.setdefault(subject_pk, set()).add((first_mt.day, first_mt.time))

            subject_counts = []
            missed_labs = []
            total_missing_classes = 0

            for subject in sorted(
                allowed_subjects,
                key=lambda item: (0 if item.room_required == "Lab" else 1, item.subject_number, item.subject_name),
            ):
                required = max(0, getattr(subject, "classes_per_week", 0) or 0)
                is_lab = subject.room_required == "Lab"
                count = len(lab_counts.get(subject.pk, set())) if is_lab else class_counts.get(subject.pk, 0)
                count = min(count, required) if required else count
                missing = max(required - count, 0)
                total_missing_classes += missing

                reason = ""
                suggested_slots = []
                instructors = list(subject.instructors.all())
                if missing:
                    if not instructors:
                        reason = "Teacher not mapped"
                    elif is_lab:
                        specific_room_tokens = [
                            token for token in normalize_specific_rooms(getattr(subject, "specific_rooms", "")).split(";")
                            if token
                        ]
                        specific_rooms = [
                            room for room in all_rooms
                            if room.r_number in specific_room_tokens
                        ]
                        required_categories = normalize_lab_categories(getattr(subject, "required_lab_category", ""))
                        matching_rooms = [
                            room for room in lab_rooms
                            if not required_categories or lab_category_matches(room.lab_category, required_categories)
                        ]
                        if specific_room_tokens:
                            reason = "No conflict-free lab slot available" if specific_rooms else "Assigned specific room unavailable"
                        elif required_categories and not matching_rooms:
                            reason = "Required lab category unavailable"
                        else:
                            reason = "No conflict-free lab slot available"
                    else:
                        reason = "No conflict-free lecture slot available"
                        suggested_slots = _build_suggested_slots(section_obj, subject)

                entry = {
                    "name": subject.subject_name,
                    "count": count,
                    "required": required,
                    "missing": missing,
                    "unfulfilled": missing,
                    "is_lab": is_lab,
                    "reason": reason,
                    "suggested_slots": suggested_slots,
                }
                subject_counts.append(entry)
                if is_lab and missing:
                    missed_labs.append(
                        {
                            "name": subject.subject_name,
                            "missing": missing,
                            "reason": reason,
                        }
                    )

            tables.append({
                "section": section_obj,
                "rows": rows,
                "subject_counts": subject_counts,
                "missed_labs": missed_labs,
                "total_missing_classes": total_missing_classes,
            })
        return tables


if "build_teacher_tables" not in globals():
    def build_teacher_tables(all_classes, all_labs, user=None):
        from collections import OrderedDict

        def _lab_slot_count(lab):
            return len(lab.meeting_times) if lab.meeting_times else getattr(lab, 'duration', LAB_DURATION)

        def _format_load(value):
            return int(value) if isinstance(value, float) and value.is_integer() else value

        def _teacher_workload(items):
            lectures = sum(getattr(cls, 'duration', 1) for cls in items["classes"])
            labs = 0
            shared_labs = 0
            for lab in items["labs"]:
                slot_count = _lab_slot_count(lab)
                if getattr(lab, "second_instructor", None):
                    shared_labs += slot_count
                else:
                    labs += slot_count
            total = lectures + labs + shared_labs
            return {
                "lectures": _format_load(lectures),
                "labs": _format_load(labs),
                "shared_labs": _format_load(shared_labs),
                "total": _format_load(total),
            }

        teacher_map = OrderedDict()
        for cls in all_classes:
            for t in [cls.instructor] + list(getattr(cls, "co_instructors", []) or []):
                if t and t not in teacher_map:
                    teacher_map[t] = {"classes": [], "labs": []}
                if t:
                    teacher_map[t]["classes"].append(cls)
        for lab in all_labs:
            teachers = [lab.instructor, getattr(lab, "second_instructor", None)] + list(getattr(lab, "co_instructors", []) or [])
            for t in teachers:
                if t and t not in teacher_map:
                    teacher_map[t] = {"classes": [], "labs": []}
                if t:
                    teacher_map[t]["labs"].append(lab)

        tables = []
        for teacher, data in teacher_map.items():
            grid = {}
            for day in DAYS:
                grid[day] = {}
            for cls in data["classes"]:
                if cls.meeting_time:
                    day = cls.meeting_time.day
                    slot = int(cls.meeting_time.time)
                    if day in grid:
                        grid[day].setdefault(slot, {"classes": [], "labs": []})
                        grid[day][slot]["classes"].append(cls)
            for lab in data["labs"]:
                if lab.meeting_times:
                    first_mt = lab.meeting_times[0]
                    day = first_mt.day
                    slot = int(first_mt.time)
                    if day in grid:
                        grid[day].setdefault(slot, {"classes": [], "labs": []})
                        grid[day][slot]["labs"].append(lab)

            rows = []
            for day in DAYS:
                cells = []
                skip_until = 0
                for s in range(1, 10):
                    if s <= skip_until:
                        continue
                    if s == 5:
                        cells.append({"type": "lunch", "colspan": 1, "slot_number": s})
                        continue
                    cell_data = grid[day].get(s, {"classes": [], "labs": []})
                    if cell_data["labs"]:
                        lab_span = max((len(lb.meeting_times) for lb in cell_data["labs"] if lb.meeting_times), default=getattr(cell_data["labs"][0], 'duration', LAB_DURATION))
                        cells.append({
                            "type": "lab", "colspan": lab_span, "slot_number": s,
                            "labs": cell_data["labs"],
                        })
                        skip_until = s + lab_span - 1
                    elif cell_data["classes"]:
                        cls_dur = getattr(cell_data["classes"][0], 'duration', 1)
                        cells.append({
                            "type": "class", "colspan": cls_dur, "slot_number": s,
                            "classes": cell_data["classes"],
                        })
                        if cls_dur > 1:
                            skip_until = s + cls_dur - 1
                    else:
                        cells.append({"type": "empty", "colspan": 1, "slot_number": s})
                rows.append({"day": day, "cells": cells})

            tables.append({"teacher": teacher, "rows": rows, "workload": _teacher_workload(data)})
        return tables


if "build_room_tables" not in globals():
    def build_room_tables(all_classes, all_labs, user=None):
        from collections import OrderedDict
        room_map = OrderedDict()
        room_qs = Room.objects.filter(user=user) if user else Room.objects.all()
        for room in room_qs.select_related("department").order_by("r_number"):
            room_map[room] = {"classes": [], "labs": []}
        for cls in all_classes:
            r = cls.room
            if r and r not in room_map:
                room_map[r] = {"classes": [], "labs": []}
            if r:
                room_map[r]["classes"].append(cls)
        for lab in all_labs:
            r = lab.room
            if r and r not in room_map:
                room_map[r] = {"classes": [], "labs": []}
            if r:
                room_map[r]["labs"].append(lab)

        tables = []
        for room, data in room_map.items():
            grid = {}
            for day in DAYS:
                grid[day] = {}
            for cls in data["classes"]:
                if cls.meeting_time:
                    day = cls.meeting_time.day
                    slot = int(cls.meeting_time.time)
                    if day in grid:
                        grid[day].setdefault(slot, {"classes": [], "labs": []})
                        grid[day][slot]["classes"].append(cls)
            for lab in data["labs"]:
                if lab.meeting_times:
                    first_mt = lab.meeting_times[0]
                    day = first_mt.day
                    slot = int(first_mt.time)
                    if day in grid:
                        grid[day].setdefault(slot, {"classes": [], "labs": []})
                        grid[day][slot]["labs"].append(lab)

            total_slots = sum(1 for d in grid.values() for s, v in d.items() if v["classes"] or v["labs"])
            max_slots = len(DAYS) * 8
            optimization = round(total_slots / max_slots * 100) if max_slots else 0

            rows = []
            for day in DAYS:
                cells = []
                skip_until = 0
                for s in range(1, 10):
                    if s <= skip_until:
                        continue
                    if s == 5:
                        cells.append({"type": "lunch", "colspan": 1, "slot_number": s})
                        continue
                    cell_data = grid[day].get(s, {"classes": [], "labs": []})
                    if cell_data["labs"]:
                        lab_span = max((len(lb.meeting_times) for lb in cell_data["labs"] if lb.meeting_times), default=getattr(cell_data["labs"][0], 'duration', LAB_DURATION))
                        cells.append({
                            "type": "lab", "colspan": lab_span, "slot_number": s,
                            "labs": cell_data["labs"],
                        })
                        skip_until = s + lab_span - 1
                    elif cell_data["classes"]:
                        cls_dur = getattr(cell_data["classes"][0], 'duration', 1)
                        cells.append({
                            "type": "class", "colspan": cls_dur, "slot_number": s,
                            "classes": cell_data["classes"],
                        })
                        if cls_dur > 1:
                            skip_until = s + cls_dur - 1
                    else:
                        cells.append({"type": "empty", "colspan": 1, "slot_number": s})
                rows.append({"day": day, "cells": cells})

            room_dept = getattr(room, "department", None)
            room_dept_code = getattr(room_dept, "code", "") or ""
            room_dept_name = getattr(room_dept, "name", None) or getattr(room_dept, "dept_name", "") or room_dept_code
            tables.append({
                "room": room,
                "rows": rows,
                "optimization": optimization,
                "optimization_percentage": optimization,
                "dept_codes": [room_dept_code] if room_dept_code else [],
                "dept_names": [room_dept_name] if room_dept_name else [],
            })
        return tables


if "timetable" not in globals():
    def timetable(request):
        return _runtime_unavailable_response(request)


if "get_meeting_time" not in globals():
    def get_meeting_time(day, slot, user=None):
        if slot is None:
            return None
        try:
            slot_str = str(int(slot))
        except (TypeError, ValueError):
            return None
        qs = MeetingTime.objects.filter(day=day, time=slot_str)
        if user:
            qs = qs.filter(user=user)
        return qs.first()


if "normalize_lab_category" not in globals():
    def normalize_lab_category(value):
        """Fallback: normalize a lab category string to match LAB_CATEGORY_CHOICES."""
        if not value:
            return ""
        value = value.strip()
        from ttgen.models import LAB_CATEGORY_CHOICES
        valid = {v.lower(): v for v, _ in LAB_CATEGORY_CHOICES}
        lower = value.lower()
        if lower in valid:
            return valid[lower]
        for key, canonical in valid.items():
            if lower in key or key in lower:
                return canonical
        return value


if "normalize_lab_categories" not in globals():
    def normalize_lab_categories(value):
        if not value:
            return []
        tokens = [token.strip() for token in re.split(r"[;\n]+", str(value)) if token.strip()]
        categories = []
        seen = set()
        for token in tokens:
            normalized = normalize_lab_category(token)
            if not normalized:
                continue
            key = normalized.casefold()
            if key in seen:
                continue
            seen.add(key)
            categories.append(normalized)
        return categories


if "normalize_lab_categories_value" not in globals():
    def normalize_lab_categories_value(value):
        return ";".join(normalize_lab_categories(value))


if "lab_category_matches" not in globals():
    def lab_category_matches(room_category, required_categories):
        normalized_room = normalize_lab_category(room_category)
        required = normalize_lab_categories(required_categories)
        if not required:
            return True
        room_key = normalized_room.casefold()
        return any(room_key == option.casefold() for option in required)


if "normalize_specific_rooms" not in globals():
    def normalize_specific_rooms(value):
        if not value:
            return ""
        tokens = [token.strip() for token in re.split(r"[;,\n]+", str(value)) if token.strip()]
        return ";".join(tokens)


if "teacher_payload" not in globals():
    def teacher_payload(name, designation="", max_workload=""):
        """Fallback: return sensible defaults for teacher designation & workload."""
        resolved_designation = designation.strip() if designation else "Assistant Professor"
        try:
            resolved_workload = int(max_workload)
        except (TypeError, ValueError):
            resolved_workload = 12
        return resolved_designation, resolved_workload


if "normalize_section_id_list" not in globals():
    def normalize_section_id_list(raw_value):
        if not raw_value:
            return ""
        sections = []
        seen = set()
        for token in re.split(r"[;,\n]+", str(raw_value)):
            value = token.strip()
            normalized = re.sub(r"\s+", " ", value)
            key = normalized.casefold()
            if not normalized or key in seen:
                continue
            sections.append(normalized)
            seen.add(key)
        return ";".join(sections)


for _runtime_view_name in (
    "delete_slot",
    "add_slot",
    "update_slot",
    "move_slot_dragdrop",
    "substitute_teacher",
    "substitute_lab_teacher",
    "save_timetable",
    "saved_timetable",
    "saved_timetable_list",
    "teachertimetable_list",
    "saved_teacher_timetable",
    "saved_delete_slot",
    "saved_add_slot",
    "saved_update_slot",
):
    if _runtime_view_name not in globals():
        def _fallback_runtime_view(request, *args, **kwargs):
            return _runtime_unavailable_response(request)
        globals()[_runtime_view_name] = _fallback_runtime_view


# ================================================================
# SAVED TIMETABLE VIEWS (overrides private runtime stubs)
# ================================================================

def _rebuild_classes_and_labs_from_saved(saved_t):
    """Rebuild in-memory Class and Lab objects from ScheduledSlot records."""
    slots = saved_t.slots.select_related(
        "section", "section__department", "subject", "instructor", "second_instructor", "room", "meeting_time",
    ).prefetch_related("lab_slots").all()

    classes = []
    labs = []
    lab_slots_list = []

    # Pre-fetch all meeting times grouped by day for multi-slot class reconstruction
    from ttgen.models import MeetingTime as MT
    all_mts = list(MT.objects.filter(user=saved_t.user))
    mts_by_day = {}
    slot_order = ["1","2","3","4","5","6","7","8","9"]
    for m in all_mts:
        mts_by_day.setdefault(m.day, []).append(m)
    for d in mts_by_day:
        mts_by_day[d].sort(key=lambda x: slot_order.index(x.time) if x.time in slot_order else 99)

    for slot in slots:
        if slot.is_lab:
            lab_slots_list.append(slot)
        else:
            cls = Class(
                id=0,
                dept=slot.section.department,
                section=slot.section.section_id,
                subject=slot.subject,
            )
            cls.instructor = slot.instructor
            cls.room = slot.room
            cls.meeting_time = slot.meeting_time
            # Rebuild meeting_times for multi-slot classes
            cls_dur = getattr(slot.subject, 'duration', 1)
            if cls_dur > 1:
                day_mts = mts_by_day.get(slot.meeting_time.day, [])
                start_idx = next((i for i, m in enumerate(day_mts) if m.time == slot.meeting_time.time), None)
                if start_idx is not None and start_idx + cls_dur <= len(day_mts):
                    cls.meeting_times = day_mts[start_idx:start_idx + cls_dur]
                else:
                    cls.meeting_times = [slot.meeting_time]
            else:
                cls.meeting_times = [slot.meeting_time]
            classes.append(cls)

    # Group labs by (section, starting day/time) to compute batch numbers
    from collections import defaultdict
    lab_groups = defaultdict(list)
    for slot in lab_slots_list:
        key = (slot.section.section_id, slot.meeting_time.day, slot.meeting_time.time)
        lab_groups[key].append(slot)

    for key, group in lab_groups.items():
        total_batches = len(group)
        for batch_num, slot in enumerate(group, start=1):
            lab_obj = Lab(
                id=0,
                dept=slot.section.department,
                section=slot.section.section_id,
                subject=slot.subject,
                batch=batch_num,
                total_batches=total_batches,
            )
            lab_obj.instructor = slot.instructor
            lab_obj.second_instructor = slot.second_instructor
            lab_obj.room = slot.room
            lab_obj.meeting_times = list(slot.lab_slots.all())
            labs.append(lab_obj)

    if (not classes and not labs) and getattr(saved_t, "snapshot", None):
        snapshot = saved_t.snapshot or {}
        for entry in list(snapshot.get("classes") or []):
            restored = _prefill_restore_class(entry, saved_t.user)
            if restored is not None:
                classes.append(restored)
        for entry in list(snapshot.get("labs") or []):
            restored = _prefill_restore_lab(entry, saved_t.user)
            if restored is not None:
                labs.append(restored)

    return classes, labs


def _room_usage_counts_from_entities(classes, labs):
    usage_counts = {}
    for cls in classes or []:
        room = getattr(cls, "room", None)
        if room is None:
            continue
        usage_counts[room.pk] = usage_counts.get(room.pk, 0) + max(int(getattr(cls, "duration", 1) or 1), 1)
    for lab in labs or []:
        room = getattr(lab, "room", None)
        if room is None:
            continue
        usage_counts[room.pk] = usage_counts.get(room.pk, 0) + max(len(getattr(lab, "meeting_times", None) or []), 1)
    return usage_counts


def _section_subject_mapping(section_obj, subject):
    if section_obj is None or getattr(subject, "pk", None) is None:
        return None
    return SectionSubjectInstructor.objects.filter(
        user=getattr(section_obj, "user", None),
        section=section_obj,
        subject=subject,
    ).select_related("instructor", "second_instructor").first()


def _subject_teacher_candidates(section_obj, subject):
    mapping = _section_subject_mapping(section_obj, subject)
    ordered = []
    seen = set()

    def _add(teacher):
        if not teacher:
            return
        key = getattr(teacher, "pk", None) or id(teacher)
        if key in seen:
            return
        seen.add(key)
        ordered.append(teacher)

    if mapping is not None:
        _add(mapping.instructor)
        teacher_by_id = {
            teacher.id: teacher
            for teacher in Instructor.objects.filter(
                id__in=list(filter(None, getattr(mapping, "group_instructor_ids", []) or [])),
                user=getattr(section_obj, "user", None),
            )
        }
        for teacher_id in getattr(mapping, "group_instructor_ids", []) or []:
            _add(teacher_by_id.get(teacher_id))

    for teacher in subject.instructors.all().order_by("name"):
        _add(teacher)

    return ordered, mapping


def _subject_room_candidates(user, section_obj, subject, classes, labs):
    is_lab = getattr(subject, "room_required", "") == "Lab"
    usage_counts = _room_usage_counts_from_entities(classes, labs)
    specific_tokens = [
        token for token in normalize_specific_rooms(getattr(subject, "specific_rooms", "")).split(";")
        if token
    ]

    rooms = []
    if specific_tokens:
        for token in specific_tokens:
            try:
                rooms.append(
                    _resolve_room_for_user(
                        token,
                        user,
                        department=getattr(section_obj, "department", None),
                    )
                )
            except Room.DoesNotExist:
                continue
    else:
        required_room_type = (getattr(subject, "room_required", "") or "").strip()
        if is_lab:
            room_qs = Room.objects.filter(user=user, room_type="Lab")
        elif required_room_type:
            room_qs = Room.objects.filter(user=user, room_type=required_room_type)
        else:
            room_qs = Room.objects.filter(user=user, room_type="Lecture Hall")

        if section_obj is not None and getattr(section_obj, "department", None) is not None:
            local_rooms = list(room_qs.filter(department=section_obj.department).select_related("department"))
            rooms = local_rooms or list(room_qs.select_related("department"))
        else:
            rooms = list(room_qs.select_related("department"))

    if is_lab:
        required_categories = normalize_lab_categories(getattr(subject, "required_lab_category", ""))
        if required_categories:
            rooms = [room for room in rooms if lab_category_matches(getattr(room, "lab_category", ""), required_categories)]

    rooms.sort(key=lambda room: (-usage_counts.get(room.pk, 0), str(getattr(room, "r_number", "") or "").lower(), room.pk))
    return rooms


def _section_subject_daily_count(classes, section_id, subject, day):
    count = 0
    for cls in classes or []:
        meeting_time = getattr(cls, "meeting_time", None)
        if not meeting_time:
            continue
        if str(getattr(cls, "section", "")) != str(section_id):
            continue
        if getattr(cls, "subject", None) != subject:
            continue
        if getattr(meeting_time, "day", "") != day:
            continue
        count += 1
    return count


def _required_subject_occurrences(section_obj, subject):
    group_counter = globals().get("get_section_subject_group_count")
    if callable(group_counter):
        try:
            group_count = max(1, int(group_counter(section_obj, subject) or 1))
        except Exception:
            group_count = 1
    else:
        try:
            mapping = SectionSubjectMapping.objects.filter(section=section_obj, subject=subject).only("group_count").first()
            group_count = max(1, getattr(mapping, "group_count", 1) or 1)
        except Exception:
            group_count = 1
    return max(1, int(getattr(subject, "classes_per_week", 1) or 1) * group_count)


def _assigned_subject_occurrences(classes, labs, section_id, subject):
    if getattr(subject, "room_required", "") == "Lab":
        return sum(1 for lab in labs or [] if str(getattr(lab, "section", "")) == str(section_id) and getattr(lab, "subject", None) == subject)
    return sum(1 for cls in classes or [] if str(getattr(cls, "section", "")) == str(section_id) and getattr(cls, "subject", None) == subject)


def _slot_block(day, start_slot, duration, user):
    block = []
    for offset in range(max(int(duration or 1), 1)):
        slot_number = int(start_slot) + offset
        if slot_number > 9 or str(slot_number) == str(LUNCH_SLOT):
            return None
        meeting_time = get_meeting_time(day, slot_number, user=user)
        if meeting_time is None:
            return None
        block.append(meeting_time)
    return block


def _entity_occupied_slots(entity):
    meeting_times = list(getattr(entity, "meeting_times", None) or [])
    if meeting_times:
        return {(mt.day, str(mt.time)) for mt in meeting_times if mt is not None}

    meeting_time = getattr(entity, "meeting_time", None)
    if meeting_time is None:
        return set()

    duration = max(1, int(getattr(entity, "duration", 1) or 1))
    start_slot = int(meeting_time.time)
    return {(meeting_time.day, str(start_slot + offset)) for offset in range(duration)}


def _section_block_is_free(section_id, block, classes, labs):
    target_slots = {(mt.day, str(mt.time)) for mt in block or []}
    if not target_slots:
        return False

    for cls in classes or []:
        if str(getattr(cls, "section", "")) != str(section_id):
            continue
        if _entity_occupied_slots(cls) & target_slots:
            return False

    for lab in labs or []:
        if str(getattr(lab, "section", "")) != str(section_id):
            continue
        if _entity_occupied_slots(lab) & target_slots:
            return False

    return True


def _explain_slot_conflict_for_entities(day, slot, teacher, room, section_id, subject, classes, labs, co_instructors=None):
    slot_str = str(slot)

    moving_teachers = set()
    if teacher is not None:
        moving_teachers.add(teacher)
    for co_teacher in (co_instructors or []):
        if co_teacher:
            moving_teachers.add(co_teacher)

    def _occupant_teachers(obj):
        staff = set()
        instructor = getattr(obj, "instructor", None)
        if instructor:
            staff.add(instructor)
        second_instructor = getattr(obj, "second_instructor", None)
        if second_instructor:
            staff.add(second_instructor)
        for co_teacher in (getattr(obj, "co_instructors", None) or []):
            if co_teacher:
                staff.add(co_teacher)
        return staff

    same_subject_slots = []
    for cls in classes or []:
        meeting_time = getattr(cls, "meeting_time", None)
        if meeting_time is None:
            continue
        if str(getattr(cls, "section", "")) != str(section_id):
            continue
        if getattr(cls, "subject", None) != subject:
            continue
        if getattr(meeting_time, "day", "") != day:
            continue
        for _, occupied_slot in _entity_occupied_slots(cls):
            same_subject_slots.append(int(occupied_slot))

    if len(same_subject_slots) == 1:
        existing_slot = same_subject_slots[0]
        if abs(existing_slot - int(slot)) == 1:
            return f"{subject.subject_name} already has a nearby slot on {day} at slot {existing_slot}, so adjacent scheduling is blocked."

    for cls in classes or []:
        if (day, slot_str) not in _entity_occupied_slots(cls):
            continue
        if room is not None and getattr(cls, "room", None) == room:
            return f"Room {room.r_number} is already occupied by {cls.subject.subject_name} (Section {cls.section}) in this slot."
        clash_teacher = moving_teachers & _occupant_teachers(cls)
        if clash_teacher:
            teacher_name = next(iter(clash_teacher)).name
            return f"{teacher_name} is already teaching {cls.subject.subject_name} (Section {cls.section}) in this slot."
        if str(getattr(cls, "section", "")) == str(section_id):
            return f"Section {section_id} already has {cls.subject.subject_name} in this slot."

    for lab in labs or []:
        if (day, slot_str) not in _entity_occupied_slots(lab):
            continue
        if room is not None and getattr(lab, "room", None) == room:
            return f"Room {room.r_number} is already occupied by {lab.subject.subject_name} Lab (Section {lab.section}) in this slot."
        clash_teacher = moving_teachers & _occupant_teachers(lab)
        if clash_teacher:
            teacher_name = next(iter(clash_teacher)).name
            return f"{teacher_name} is already assigned to {lab.subject.subject_name} Lab (Section {lab.section}) in this slot."
        if str(getattr(lab, "section", "")) == str(section_id):
            return f"Section {section_id} already has a lab running in this slot."

    return None


def _find_missing_subject_placement(user, section_obj, subject, classes, labs):
    if section_obj is None or subject is None:
        return None, "Section or subject was not found."

    teacher_candidates, mapping = _subject_teacher_candidates(section_obj, subject)
    room_candidates = _subject_room_candidates(user, section_obj, subject, classes, labs)
    if not teacher_candidates:
        return None, "No teacher is mapped to this subject."
    if not room_candidates:
        room_type = "lab room" if getattr(subject, "room_required", "") == "Lab" else "lecture hall"
        return None, f"No matching {room_type} is available for this subject in the allowed department scope."

    is_lab = getattr(subject, "room_required", "") == "Lab"
    default_duration = LAB_DURATION if is_lab else 1
    duration = max(1, int(getattr(subject, "duration", default_duration) or default_duration))
    start_slots = [int(slot) for slot in VALID_LAB_START_SLOTS] if is_lab else [slot for slot in range(1, 10) if str(slot) != str(LUNCH_SLOT)]
    sorted_teachers = sorted(
        teacher_candidates,
        key=lambda teacher: (_compute_teacher_workload(teacher, classes, labs), str(getattr(teacher, "name", "") or "").lower(), teacher.pk),
    )
    last_reason = "No conflict-free slot could be created for this subject."
    candidate_blocks = []

    for day in DAYS:
        if not is_lab and _section_subject_daily_count(classes, section_obj.section_id, subject, day) >= 2:
            last_reason = f"{subject.subject_name} already has the maximum allowed classes on {day}."
            continue
        for start_slot in start_slots:
            block = _slot_block(day, start_slot, duration, user)
            if not block:
                if not is_lab:
                    last_reason = f"This subject does not fit from slot {start_slot} on {day}."
                continue

            candidate_blocks.append({
                "day": day,
                "start_slot": start_slot,
                "block": block,
                "section_free": _section_block_is_free(section_obj.section_id, block, classes, labs),
            })

    candidate_blocks.sort(key=lambda item: (0 if item["section_free"] else 1, DAYS.index(item["day"]), int(item["start_slot"])))
    section_free_windows = sum(1 for item in candidate_blocks if item["section_free"])

    for candidate in candidate_blocks:
        day = candidate["day"]
        start_slot = candidate["start_slot"]
        block = candidate["block"]
        if not candidate["section_free"]:
            continue

        if is_lab:
            free_teachers = [teacher for teacher in sorted_teachers if _teacher_is_free_for_lab(teacher, block, classes, labs)]
        else:
            free_teachers = [
                teacher for teacher in sorted_teachers
                if all(_teacher_is_free_at(teacher, day, int(mt.time), classes, labs) for mt in block)
            ]
        if not free_teachers:
            last_reason = f"Section {section_obj.section_id} is free here, but all mapped teachers are busy on {day} at slot {start_slot}."
            continue

        for room in room_candidates:
            for teacher in free_teachers:
                blocked = False
                for meeting_time in block:
                    conflict_reason = _explain_slot_conflict_for_entities(
                        day,
                        int(meeting_time.time),
                        teacher,
                        room,
                        section_obj.section_id,
                        subject,
                        classes,
                        labs,
                        co_instructors=[],
                    )
                    if conflict_reason:
                        last_reason = conflict_reason
                        blocked = True
                        break
                if blocked:
                    continue
                return {
                    "section": section_obj,
                    "subject": subject,
                    "teacher": teacher,
                    "second_instructor": getattr(mapping, "second_instructor", None) if is_lab else None,
                    "room": room,
                    "meeting_times": block,
                    "is_lab": is_lab,
                }, ""

    if section_free_windows <= 0:
        return None, f"Section {section_obj.section_id} does not have any fully empty {'lab block' if is_lab else 'slot'} left for {subject.subject_name}."

    if last_reason == "No conflict-free slot could be created for this subject.":
        last_reason = f"Scanned {section_free_windows} section-free window(s), but no teacher-room combination was conflict-free for {subject.subject_name}."

    return None, last_reason


def _apply_generated_missing_subject_placement(request, section_obj, subject):
    state = _get_user_state(request.user.id)
    classes = list(state.get("classes") or GLOBAL_CLASSES or [])
    labs = list(state.get("labs") or GLOBAL_LABS or [])
    placement, failure_reason = _find_missing_subject_placement(request.user, section_obj, subject, classes, labs)
    if placement is None:
        return None, failure_reason

    if placement["is_lab"]:
        new_lab = Lab(
            _next_in_memory_class_id(),
            section_obj.department,
            section_obj.section_id,
            subject,
        )
        new_lab.set_instructor(placement["teacher"])
        if placement.get("second_instructor") is not None:
            new_lab.set_second_instructor(placement["second_instructor"])
        new_lab.set_room(placement["room"])
        new_lab.set_meetingTimes(placement["meeting_times"])
        labs.append(new_lab)
        state["labs"] = labs
        return {"kind": "lab", "slot": placement["meeting_times"][0]}, ""

    new_class = Class(
        _next_in_memory_class_id(),
        section_obj.department,
        section_obj.section_id,
        subject,
    )
    new_class.set_instructor(placement["teacher"])
    new_class.set_room(placement["room"])
    new_class.set_meetingTime(placement["meeting_times"][0])
    new_class.meeting_times = placement["meeting_times"]
    new_class.duration = len(placement["meeting_times"])
    classes.append(new_class)
    state["classes"] = classes
    return {"kind": "class", "slot": placement["meeting_times"][0]}, ""


def _apply_generated_missing_subject_placements(request, section_obj, subject):
    state = _get_user_state(request.user.id)
    classes = list(state.get("classes") or GLOBAL_CLASSES or [])
    labs = list(state.get("labs") or GLOBAL_LABS or [])
    required = _required_subject_occurrences(section_obj, subject)
    assigned = _assigned_subject_occurrences(classes, labs, section_obj.section_id, subject)
    missing = max(0, required - assigned)
    if missing <= 0:
        return {"created": 0, "required": required, "remaining": 0, "kind": "lab" if getattr(subject, "room_required", "") == "Lab" else "class"}, "This subject is already fully scheduled."

    created = 0
    first_slot = None
    last_slot = None
    failure_reason = ""
    kind = "lab" if getattr(subject, "room_required", "") == "Lab" else "class"

    while created < missing:
        result, failure_reason = _apply_generated_missing_subject_placement(request, section_obj, subject)
        if result is None:
            break
        created += 1
        first_slot = first_slot or result.get("slot")
        last_slot = result.get("slot") or last_slot
        kind = result.get("kind") or kind

    return {
        "created": created,
        "required": required,
        "remaining": max(0, missing - created),
        "kind": kind,
        "first_slot": first_slot,
        "last_slot": last_slot,
    }, failure_reason


def _apply_saved_missing_subject_placement(saved_t, section_obj, subject):
    classes, labs = _rebuild_classes_and_labs_from_saved(saved_t)
    placement, failure_reason = _find_missing_subject_placement(saved_t.user, section_obj, subject, classes, labs)
    if placement is None:
        return None, failure_reason

    with transaction.atomic():
        saved_slot = ScheduledSlot.objects.create(
            timetable=saved_t,
            section=section_obj,
            subject=subject,
            instructor=placement["teacher"],
            second_instructor=placement.get("second_instructor") if placement["is_lab"] else None,
            room=placement["room"],
            meeting_time=placement["meeting_times"][0],
            is_lab=placement["is_lab"],
        )
        if placement["is_lab"]:
            saved_slot.lab_slots.set(placement["meeting_times"])
    return {"kind": "lab" if placement["is_lab"] else "class", "slot": placement["meeting_times"][0]}, ""


def _apply_saved_missing_subject_placements(saved_t, section_obj, subject):
    classes, labs = _rebuild_classes_and_labs_from_saved(saved_t)
    required = _required_subject_occurrences(section_obj, subject)
    assigned = _assigned_subject_occurrences(classes, labs, section_obj.section_id, subject)
    missing = max(0, required - assigned)
    if missing <= 0:
        return {"created": 0, "required": required, "remaining": 0, "kind": "lab" if getattr(subject, "room_required", "") == "Lab" else "class"}, "This subject is already fully scheduled."

    created = 0
    first_slot = None
    last_slot = None
    failure_reason = ""
    kind = "lab" if getattr(subject, "room_required", "") == "Lab" else "class"

    while created < missing:
        result, failure_reason = _apply_saved_missing_subject_placement(saved_t, section_obj, subject)
        if result is None:
            break
        created += 1
        first_slot = first_slot or result.get("slot")
        last_slot = result.get("slot") or last_slot
        kind = result.get("kind") or kind

    return {
        "created": created,
        "required": required,
        "remaining": max(0, missing - created),
        "kind": kind,
        "first_slot": first_slot,
        "last_slot": last_slot,
    }, failure_reason


@login_required
def saved_reshuffle_missing_subject(request, tid, section_id, subject_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "Invalid request method."}, status=405)

    saved_t = _get_saved_timetable_or_404(tid, request.user)
    section_obj = Section.objects.filter(user=request.user, section_id=section_id).select_related("department").first()
    if section_obj is None:
        return JsonResponse({"ok": False, "message": "Section not found."}, status=404)

    subject = Subject.objects.filter(user=request.user, pk=subject_id).first()
    if subject is None:
        return JsonResponse({"ok": False, "message": "Subject not found."}, status=404)

    result, failure_reason = _apply_saved_missing_subject_placements(saved_t, section_obj, subject)
    if not result or result.get("created", 0) <= 0:
        return JsonResponse({"ok": False, "message": failure_reason or "No conflict-free slot could be created for this subject."}, status=409)

    first_slot = result.get("first_slot")
    slot_label = SLOT_LABELS.get(str(first_slot.time), f"Slot {first_slot.time}") if first_slot else "a valid slot"
    created = result.get("created", 0)
    remaining = result.get("remaining", 0)
    if remaining > 0:
        message = f"Created {created} {result['kind']} slot(s) for {subject.subject_name}. {remaining} still could not be placed. First slot: {first_slot.day} at {slot_label}."
    else:
        message = f"Created {created} {result['kind']} slot(s) for {subject.subject_name}. First slot: {first_slot.day} at {slot_label}."
    return JsonResponse({
        "ok": True,
        "message": message,
    })


def _compute_teacher_workloads(classes, labs):
    """Compute teacher workload summary from classes and labs."""
    workloads = {}

    def _ensure_teacher(teacher):
        if teacher not in workloads:
            workloads[teacher] = {
                "lectures": 0,
                "labs": 0,
                "shared_labs": 0,
                "total": 0,
                "departments": set(),
            }
        return workloads[teacher]

    def _add_department(teacher, item_obj):
        data = _ensure_teacher(teacher)
        dept = getattr(item_obj, "department", None)
        dept_name = getattr(dept, "name", "") or getattr(dept, "code", "")
        if dept_name:
            data["departments"].add(str(dept_name))

    # Elective sessions are taught once physically but exist as one object per
    # shared section. Track seen (teacher, physical-session) pairs so an elective
    # adds to a teacher's load only once instead of once per section.
    seen_elective = set()

    def _is_dup_elective(item_obj, kind, first_mt, assigned_teacher):
        if not getattr(item_obj, "is_elective", False):
            return False
        key = (
            getattr(assigned_teacher, "pk", None) or id(assigned_teacher),
            kind,
            getattr(getattr(item_obj, "subject", None), "pk", None),
            getattr(getattr(item_obj, "room", None), "pk", None),
            getattr(first_mt, "day", None),
            getattr(first_mt, "time", None),
            getattr(item_obj, "group", None),
        )
        if key in seen_elective:
            return True
        seen_elective.add(key)
        return False

    for cls in classes:
        teacher = cls.instructor
        cls_dur = getattr(cls, 'duration', 1)
        teachers = [teacher] + list(getattr(cls, "co_instructors", []) or [])
        for assigned_teacher in teachers:
            if not assigned_teacher:
                continue
            if _is_dup_elective(cls, "class", getattr(cls, "meeting_time", None), assigned_teacher):
                continue
            data = _ensure_teacher(assigned_teacher)
            data["lectures"] += cls_dur
            data["total"] += cls_dur
            _add_department(assigned_teacher, cls)
    for lab in labs:
        teacher = lab.instructor
        lab_slot_count = len(lab.meeting_times) if lab.meeting_times else getattr(lab, 'duration', LAB_DURATION)
        second_teacher = getattr(lab, "second_instructor", None)
        co_teachers = list(getattr(lab, "co_instructors", []) or [])
        _lab_first_mt = (lab.meeting_times or [None])[0]
        if second_teacher:
            shared_load = lab_slot_count
            for assigned_teacher in [teacher, second_teacher] + co_teachers:
                if not assigned_teacher:
                    continue
                if _is_dup_elective(lab, "lab", _lab_first_mt, assigned_teacher):
                    continue
                data = _ensure_teacher(assigned_teacher)
                data["shared_labs"] += shared_load
                data["total"] += shared_load
                _add_department(assigned_teacher, lab)
        else:
            for assigned_teacher in [teacher] + co_teachers:
                if not assigned_teacher:
                    continue
                if _is_dup_elective(lab, "lab", _lab_first_mt, assigned_teacher):
                    continue
                data = _ensure_teacher(assigned_teacher)
                data["labs"] += lab_slot_count
                data["total"] += lab_slot_count
                _add_department(assigned_teacher, lab)

    for data in workloads.values():
        data["departments"] = ", ".join(sorted(data["departments"])) or "-"
        for key in ("lectures", "labs", "shared_labs", "total"):
            value = data.get(key, 0)
            if isinstance(value, float) and value.is_integer():
                data[key] = int(value)
    return workloads


def _get_program_filter_options(user):
    """Return distinct, non-empty program names for section-level filtering."""
    return sorted(
        {
            (name or "").strip()
            for name in Section.objects.filter(user=user).values_list("program_name", flat=True)
            if (name or "").strip()
        },
        key=lambda value: value.lower(),
    )


def _filter_entities_by_program(classes, labs, user, selected_program):
    """Filter in-memory timetable entities by section program name."""
    selected_program = (selected_program or "all").strip()
    if not selected_program or selected_program.lower() == "all":
        return classes, labs, "all"

    allowed_sections = set(
        Section.objects.filter(user=user, program_name__iexact=selected_program).values_list("section_id", flat=True)
    )
    filtered_classes = [cls for cls in classes if str(getattr(cls, "section", "")) in allowed_sections]
    filtered_labs = [lab for lab in labs if str(getattr(lab, "section", "")) in allowed_sections]
    return filtered_classes, filtered_labs, selected_program


def _get_department_filter_options(user):
    """Return the user's departments as {code, name} dicts for filtering."""
    options = []
    for dept in Department.objects.filter(user=user).order_by("code", "name"):
        code = (dept.code or "").strip()
        if not code:
            continue
        options.append({"code": code, "name": (dept.name or "").strip() or code})
    return options


def _get_saved_department_filter_options(user, classes, labs):
    all_options = _get_department_filter_options(user)
    names_by_code = {opt["code"]: opt["name"] for opt in all_options}
    relevant_codes = set()

    def _add_code(code):
        code = (code or "").strip()
        if code:
            relevant_codes.add(code)

    for cls in classes:
        dept = getattr(getattr(cls, "section", None), "department", None) or getattr(cls, "department", None)
        _add_code(getattr(dept, "code", ""))
        teacher_dept = getattr(getattr(cls, "instructor", None), "department", None)
        _add_code(getattr(teacher_dept, "code", ""))
        for co_teacher in getattr(cls, "co_instructors", []) or []:
            teacher_dept = getattr(co_teacher, "department", None)
            _add_code(getattr(teacher_dept, "code", ""))

    for lab in labs:
        dept = getattr(getattr(lab, "section", None), "department", None) or getattr(lab, "department", None)
        _add_code(getattr(dept, "code", ""))
        for teacher in [getattr(lab, "instructor", None), getattr(lab, "second_instructor", None)] + list(getattr(lab, "co_instructors", []) or []):
            teacher_dept = getattr(teacher, "department", None)
            _add_code(getattr(teacher_dept, "code", ""))

    options = []
    for code in sorted(relevant_codes):
        options.append({"code": code, "name": names_by_code.get(code, code)})
    return options


def _teacher_home_dept_fallback(classes, labs):
    """Derive a fallback home-department code per teacher from where they teach
    most. Used only when an Instructor has no stored department."""
    from collections import defaultdict
    counts = defaultdict(lambda: defaultdict(int))

    def _add(teacher, dept):
        if teacher is None or dept is None:
            return
        code = (getattr(dept, "code", "") or "").strip()
        if code:
            counts[getattr(teacher, "id", None)][code] += 1

    for cls in classes:
        _add(getattr(cls, "instructor", None), getattr(cls, "department", None))
    for lab in labs:
        _add(getattr(lab, "instructor", None), getattr(lab, "department", None))
        _add(getattr(lab, "second_instructor", None), getattr(lab, "department", None))

    fallback = {}
    for tid, dept_counts in counts.items():
        if dept_counts:
            fallback[tid] = max(dept_counts.items(), key=lambda kv: kv[1])[0]
    return fallback


def _teacher_home_dept_code(teacher, fallback_map):
    """Home department code of a teacher: stored Instructor.department first,
    else the workload-derived fallback."""
    dept = getattr(teacher, "department", None)
    code = (getattr(dept, "code", "") or "").strip() if dept else ""
    if not code:
        code = fallback_map.get(getattr(teacher, "id", None), "") or ""
    return code


def _filter_section_tables_by_department(tables, selected_department):
    if not selected_department or selected_department.lower() == "all":
        return tables
    sel = selected_department.strip().lower()
    out = []
    for table in tables:
        section = table.get("section")
        dept = getattr(section, "department", None) if section else None
        code = (getattr(dept, "code", "") or "").strip().lower() if dept else ""
        if code == sel:
            out.append(table)
    return out


def _collect_department_codes_from_rows(rows):
    codes = set()
    for row in rows or []:
        for cell in row.get("cells", []):
            for cls in cell.get("classes", []):
                dept = getattr(getattr(cls, "section", None), "department", None) or getattr(cls, "department", None)
                code = (getattr(dept, "code", "") or "").strip().lower()
                if code:
                    codes.add(code)
            for lab in cell.get("labs", []):
                dept = getattr(getattr(lab, "section", None), "department", None) or getattr(lab, "department", None)
                code = (getattr(dept, "code", "") or "").strip().lower()
                if code:
                    codes.add(code)
    return codes


def _table_has_scheduled_entries(table):
    for row in table.get("rows", []):
        for cell in row.get("cells", []):
            if cell.get("classes") or cell.get("labs"):
                return True
    return False


def _filter_room_tables_by_department(room_tables, selected_department):
    if not selected_department or selected_department.lower() == "all":
        return room_tables
    sel = selected_department.strip().lower()
    out = []
    for table in room_tables:
        codes = {
            str(c).strip().lower()
            for c in table.get("dept_codes", [])
            if str(c).strip()
        }
        codes.update(_collect_department_codes_from_rows(table.get("rows", [])))
        if sel in codes and _table_has_scheduled_entries(table):
            out.append(table)
    return out


def _filter_teacher_tables_by_department(teacher_tables, selected_department, fallback_map):
    """Keep only teachers whose home department matches the selection.

    Saved timetable teacher/workload views are grouped by teacher home
    department, not by every department they teach into.
    """
    if not selected_department or selected_department.lower() == "all":
        return teacher_tables
    sel = selected_department.strip().lower()
    out = []
    for table in teacher_tables:
        teacher = table.get("teacher")
        home = (table.get("home_dept_code") or _teacher_home_dept_code(teacher, fallback_map) or "").strip().lower()
        if home == sel and _table_has_scheduled_entries(table):
            out.append(table)
    return out


def _filter_workloads_for_teacher_tables(workloads, teacher_tables, selected_department):
    if not selected_department or selected_department.lower() == "all":
        return workloads
    visible_teacher_ids = {
        getattr(table.get("teacher"), "id", None)
        for table in teacher_tables
        if table.get("teacher") is not None
    }
    out = {}
    for teacher, data in workloads.items():
        if getattr(teacher, "id", None) in visible_teacher_ids:
            out[teacher] = data
    return out


def _normalize_selected_department(user, selected_department):
    selected_department = (selected_department or "all").strip()
    if not selected_department or selected_department.lower() == "all":
        return "all"
    valid_codes = {
        (code or "").strip().lower()
        for code in Department.objects.filter(user=user).values_list("code", flat=True)
        if (code or "").strip()
    }
    if selected_department.lower() in valid_codes:
        return selected_department
    return "all"


def _decorate_section_tables_with_department_meta(tables):
    for table in tables:
        section = table.get("section")
        dept = getattr(section, "department", None) if section else None
        table["section_department_code"] = (getattr(dept, "code", "") or "").strip()
        table["section_department_name"] = (getattr(dept, "name", "") or "").strip()
    return tables


def _saved_filtered_entities_for_request(request, saved_t):
    classes, labs = _rebuild_classes_and_labs_from_saved(saved_t)
    selected_program = request.GET.get("program", "all")
    classes, labs, selected_program = _filter_entities_by_program(
        classes, labs, request.user, selected_program
    )
    selected_department = _normalize_selected_department(
        request.user, request.GET.get("department", "all")
    )
    return classes, labs, selected_program, selected_department



def _get_saved_timetable_or_404(tid, user):
    """Fetch a saved timetable ensuring it belongs to the user."""
    try:
        saved_t = SavedTimetable.objects.get(id=tid)
    except SavedTimetable.DoesNotExist:
        raise Http404("Timetable does not exist")
    if saved_t.user != user:
        raise Http404("Timetable does not exist")
    return saved_t


def _get_plan_permissions(user):
    """Return plan permission flags for a user.

    Feature locks have been removed — edit/delete, substitute and drag-and-drop
    are available to everyone regardless of subscription plan.
    """
    return {
        "can_edit_delete": True,
        "can_substitute": True,
        "can_drag_drop": True,
    }


def _get_slot_sequence(day, start_slot, span, user):
    slots = []
    for offset in range(max(int(span or 1), 1)):
        meeting_time = get_meeting_time(day, int(start_slot) + offset, user=user)
        if not meeting_time:
            return None
        slots.append(meeting_time)
    return slots


def _build_saved_parking_context(saved_t, tables):
    reservation_map = defaultdict(dict)
    reservations = saved_t.slot_room_reservations.select_related("section", "meeting_time", "room")
    for reservation in reservations:
        reservation_map[reservation.section.section_id][
            (reservation.meeting_time.day, reservation.meeting_time.time)
        ] = reservation

    parking_map = defaultdict(list)
    parked_slots = saved_t.parked_slots.select_related(
        "section",
        "subject",
        "instructor",
        "second_instructor",
        "original_room",
        "original_meeting_time",
    )
    for parked in parked_slots:
        parking_map[parked.section.section_id].append(parked)

    for table in tables:
        section_id = table["section"].section_id
        table["parking_items"] = parking_map.get(section_id, [])
        for row in table.get("rows", []):
            for cell in row.get("cells", []):
                reservation = reservation_map.get(section_id, {}).get((row["day"], str(cell.get("slot_number"))))
                cell["reserved_room"] = reservation.room if reservation else None
                cell["has_room_context"] = bool(reservation)

    return tables


def _reserve_saved_slot_room(saved_t, section, meeting_times, room):
    for meeting_time in meeting_times:
        SavedSlotRoomReservation.objects.update_or_create(
            timetable=saved_t,
            section=section,
            meeting_time=meeting_time,
            defaults={"room": room},
        )


def _clear_saved_slot_room_reservations(saved_t, section, meeting_times):
    SavedSlotRoomReservation.objects.filter(
        timetable=saved_t,
        section=section,
        meeting_time__in=meeting_times,
    ).delete()


@login_required
def saved_timetable_list(request):
    timetables = SavedTimetable.objects.filter(user=request.user)
    return render(request, "saved_timetable_list.html", {
        "timetables": timetables,
        "timetable_count": timetables.count(),
        "save_limit": 15,
        "save_usage_percent": min(100, int((timetables.count() / 15) * 100) if 15 else 0),
    })


@login_required
def saved_timetable(request, tid):
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    classes, labs, selected_program, selected_department = _saved_filtered_entities_for_request(request, saved_t)
    department_options = _get_saved_department_filter_options(request.user, classes, labs)
    valid_department_codes = {
        (option.get("code") or "").strip().lower()
        for option in department_options
        if (option.get("code") or "").strip()
    }
    if selected_department != "all" and selected_department.strip().lower() not in valid_department_codes:
        selected_department = "all"
    # Fallback home-department map (used only for teachers with no stored department).
    home_dept_fallback = _teacher_home_dept_fallback(classes, labs)

    # Section / room tables are filtered to the selected department directly.
    tables = build_section_tables(classes, labs, user=request.user)
    tables = _decorate_section_tables_with_department_meta(tables)
    tables = _build_saved_parking_context(saved_t, tables)
    tables = _filter_section_tables_by_department(tables, selected_department)
    room_tables = build_room_tables(classes, labs, user=request.user)
    room_tables = _filter_room_tables_by_department(room_tables, selected_department)

    # Teacher tables / workloads are built from the FULL set so each teacher's
    # grid stays complete, then the LIST is restricted to teachers whose HOME
    # department matches the selection.
    teacher_tables = build_teacher_tables(classes, labs, user=request.user)
    teacher_tables = _filter_teacher_tables_by_department(teacher_tables, selected_department, home_dept_fallback)
    teacher_workloads = _compute_teacher_workloads(classes, labs)
    teacher_workloads = _filter_workloads_for_teacher_tables(teacher_workloads, teacher_tables, selected_department)

    permissions = _get_plan_permissions(request.user)

    context = {
        "saved": saved_t,
        "tables": tables,
        "room_tables": room_tables,
        "teacher_tables": teacher_tables,
        "teacher_workloads": teacher_workloads,
        "program_options": _get_program_filter_options(request.user),
        "active_program": selected_program,
        "department_options": department_options,
        "active_department": selected_department,
        "SLOT_LABELS": SLOT_LABELS,
        "can_edit_delete": permissions["can_edit_delete"],
        "can_substitute": permissions["can_substitute"],
        "can_drag_drop": permissions["can_drag_drop"],
    }
    if request.session.get("is_superadmin") and request.session.get("sa_impersonate_uid"):
        context["superadmin_exit_url"] = reverse("superadmin_stop_impersonate")
        context["superadmin_exit_title"] = "Back to Super Admin"
    return render(request, "saved_timetable.html", context)


def _drop_room_type_for_subject(subject, original_room=None):
    room_type = (getattr(original_room, "room_type", "") or "").strip()
    if room_type:
        return room_type
    required = (getattr(subject, "room_required", "") or "").strip()
    if required == "Lab":
        return "Lab"
    return required or "Lecture Hall"


def _saved_drag_room_usage_counts(saved_timetable, ignore_slot=None):
    usage_counts = {}
    slots = saved_timetable.slots.select_related("room").prefetch_related("lab_slots")
    for scheduled in slots:
        if ignore_slot is not None and scheduled.pk == getattr(ignore_slot, "pk", None):
            continue
        room = getattr(scheduled, "room", None)
        if room is None:
            continue
        span = len(list(scheduled.lab_slots.all())) if getattr(scheduled, "is_lab", False) else 1
        usage_counts[room.pk] = usage_counts.get(room.pk, 0) + max(span, 1)
    return usage_counts


def _saved_drag_room_candidates(user, department, subject, original_room=None, usage_counts=None):
    usage_counts = usage_counts or {}
    room_type = _drop_room_type_for_subject(subject, original_room)
    candidate_qs = Room.objects.filter(user=user, room_type=room_type)
    if department is not None:
        candidate_qs = candidate_qs.filter(department=department)
    candidates = list(candidate_qs.select_related("department"))

    preferred = []
    preferred_pk = getattr(original_room, "pk", None)
    if preferred_pk is not None:
        preferred = [room for room in candidates if room.pk == preferred_pk]
        candidates = [room for room in candidates if room.pk != preferred_pk]

    candidates.sort(key=lambda room: (-usage_counts.get(room.pk, 0), str(room.r_number or "").lower(), room.pk))
    return preferred + candidates


@login_required
def saved_substitute_teacher(request, tid, section, day, slot):
    """Substitute a theory class instructor in a saved timetable."""
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    mt = get_meeting_time(day, slot, user=request.user)
    if mt is None:
        messages.error(request, "Invalid time slot.")
        return redirect("saved_timetable", tid=tid)

    try:
        sec = Section.objects.get(section_id=section, user=request.user)
    except Section.DoesNotExist:
        messages.error(request, "Section not found.")
        return redirect("saved_timetable", tid=tid)

    scheduled = saved_t.slots.filter(section=sec, meeting_time=mt, is_lab=False).first()
    if not scheduled:
        messages.error(request, "No theory class found at this slot.")
        return redirect("saved_timetable", tid=tid)

    available_teachers = Instructor.objects.filter(user=request.user).exclude(id=scheduled.instructor.id)

    if request.method == "POST":
        teacher_id = request.POST.get("teacher")
        if teacher_id:
            try:
                new_teacher = Instructor.objects.get(id=teacher_id, user=request.user)
                # Check for conflict: new teacher already has a slot at this time
                conflict = saved_t.slots.filter(
                    instructor=new_teacher, meeting_time=mt,
                ).exclude(id=scheduled.id).exists()
                if conflict:
                    messages.error(request, f"{new_teacher.name} already has a class at this time.")
                else:
                    scheduled.instructor = new_teacher
                    scheduled.save(update_fields=["instructor"])
                    messages.success(request, f"Teacher substituted to {new_teacher.name}.")
                    return redirect("saved_timetable", tid=tid)
            except Instructor.DoesNotExist:
                messages.error(request, "Selected teacher not found.")

    return render(request, "saved_substitute_teacher.html", {
        "saved": saved_t,
        "scheduled": scheduled,
        "section": sec,
        "day": day,
        "slot": slot,
        "available_teachers": available_teachers,
        "SLOT_LABELS": SLOT_LABELS,
    })


@login_required
def saved_substitute_lab_teacher(request, tid, section, day, slot):
    """Substitute a lab instructor in a saved timetable."""
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    mt = get_meeting_time(day, slot, user=request.user)
    if mt is None:
        messages.error(request, "Invalid time slot.")
        return redirect("saved_timetable", tid=tid)

    try:
        sec = Section.objects.get(section_id=section, user=request.user)
    except Section.DoesNotExist:
        messages.error(request, "Section not found.")
        return redirect("saved_timetable", tid=tid)

    scheduled = saved_t.slots.filter(section=sec, meeting_time=mt, is_lab=True).first()
    if not scheduled:
        messages.error(request, "No lab found at this slot.")
        return redirect("saved_timetable", tid=tid)

    available_teachers = Instructor.objects.filter(user=request.user).exclude(id=scheduled.instructor.id)

    if request.method == "POST":
        teacher_id = request.POST.get("teacher")
        if teacher_id:
            try:
                new_teacher = Instructor.objects.get(id=teacher_id, user=request.user)
                lab_times = list(scheduled.lab_slots.all())
                conflict = saved_t.slots.filter(
                    instructor=new_teacher, meeting_time__in=lab_times,
                ).exclude(id=scheduled.id).exists()
                if conflict:
                    messages.error(request, f"{new_teacher.name} already has a class during this lab.")
                else:
                    scheduled.instructor = new_teacher
                    scheduled.save(update_fields=["instructor"])
                    messages.success(request, f"Lab teacher substituted to {new_teacher.name}.")
                    return redirect("saved_timetable", tid=tid)
            except Instructor.DoesNotExist:
                messages.error(request, "Selected teacher not found.")

    return render(request, "saved_substitute_teacher.html", {
        "saved": saved_t,
        "scheduled": scheduled,
        "section": sec,
        "day": day,
        "slot": slot,
        "available_teachers": available_teachers,
        "SLOT_LABELS": SLOT_LABELS,
        "is_lab": True,
    })


@login_required
def saved_move_slot_dragdrop(request, tid, section, day, slot):
    """Drag-and-drop move a slot in a saved timetable."""
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST required."}, status=405)

    saved_t = _get_saved_timetable_or_404(tid, request.user)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "message": "Invalid payload."}, status=400)

    target_day = payload.get("target_day")
    target_slot = payload.get("target_slot")
    move_type = payload.get("move_type", "class")

    if not target_day or not target_slot:
        return JsonResponse({"ok": False, "message": "Missing target day/slot."}, status=400)

    source_mt = get_meeting_time(day, slot, user=request.user)
    target_mt = get_meeting_time(target_day, target_slot, user=request.user)

    if not source_mt or not target_mt:
        return JsonResponse({"ok": False, "message": "Invalid time slot."}, status=400)

    try:
        sec = Section.objects.get(section_id=section, user=request.user)
    except Section.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Section not found."}, status=404)

    if move_type == "lab":
        scheduled = saved_t.slots.filter(section=sec, meeting_time=source_mt, is_lab=True).first()
    else:
        scheduled = saved_t.slots.filter(section=sec, meeting_time=source_mt, is_lab=False).first()

    if not scheduled:
        return JsonResponse({"ok": False, "message": "No slot found at source."}, status=404)

    if move_type == "lab":
        source_lab_times = list(scheduled.lab_slots.all().order_by("time"))
        if not source_lab_times:
            return JsonResponse({"ok": False, "message": "Lab has no time slots."}, status=400)

        # Calculate offset
        source_start_slot = int(source_lab_times[0].time)
        target_start_slot = int(target_mt.time)
        offset = target_start_slot - source_start_slot

        new_lab_times = []
        for lt in source_lab_times:
            new_slot_num = int(lt.time) + offset
            new_lt = get_meeting_time(target_day, new_slot_num, user=request.user)
            if not new_lt:
                return JsonResponse({"ok": False, "message": f"Target slot {new_slot_num} on {target_day} does not exist."}, status=400)
            new_lab_times.append(new_lt)

        # Check non-room conflicts for all new lab times
        for nlt in new_lab_times:
            sec_conflict = saved_t.slots.filter(section=sec, meeting_time=nlt).exclude(id=scheduled.id).exists()
            if sec_conflict:
                return JsonResponse({"ok": False, "message": f"Section conflict at {target_day} slot {nlt.time}."}, status=409)
            teacher_conflict = saved_t.slots.filter(instructor=scheduled.instructor, meeting_time=nlt).exclude(id=scheduled.id).exists()
            if teacher_conflict:
                return JsonResponse({"ok": False, "message": f"Teacher conflict at {target_day} slot {nlt.time}."}, status=409)

        usage_counts = _saved_drag_room_usage_counts(saved_t, ignore_slot=scheduled)
        selected_room = None
        for room in _saved_drag_room_candidates(
            request.user,
            sec.department,
            scheduled.subject,
            original_room=scheduled.room,
            usage_counts=usage_counts,
        ):
            room_conflict = saved_t.slots.filter(room=room, meeting_time__in=new_lab_times).exclude(id=scheduled.id).exists()
            if not room_conflict:
                selected_room = room
                break
        if selected_room is None:
            room_label = getattr(getattr(scheduled, "room", None), "r_number", "this room")
            return JsonResponse({"ok": False, "message": f"No available { _drop_room_type_for_subject(scheduled.subject, scheduled.room) } room was found in {sec.department.name} for the selected time slot."}, status=409)

        # Apply move
        scheduled.meeting_time = new_lab_times[0]
        scheduled.room = selected_room
        scheduled.save(update_fields=["meeting_time", "room"])
        scheduled.lab_slots.set(new_lab_times)

    else:
        sec_conflict = saved_t.slots.filter(section=sec, meeting_time=target_mt).exclude(id=scheduled.id).exists()
        if sec_conflict:
            return JsonResponse({"ok": False, "message": "Section already has a class at target slot."}, status=409)
        teacher_conflict = saved_t.slots.filter(instructor=scheduled.instructor, meeting_time=target_mt).exclude(id=scheduled.id).exists()
        if teacher_conflict:
            return JsonResponse({"ok": False, "message": "Teacher already has a class at target slot."}, status=409)

        usage_counts = _saved_drag_room_usage_counts(saved_t, ignore_slot=scheduled)
        selected_room = None
        for room in _saved_drag_room_candidates(
            request.user,
            sec.department,
            scheduled.subject,
            original_room=scheduled.room,
            usage_counts=usage_counts,
        ):
            room_conflict = saved_t.slots.filter(room=room, meeting_time=target_mt).exclude(id=scheduled.id).exists()
            if not room_conflict:
                selected_room = room
                break
        if selected_room is None:
            return JsonResponse({"ok": False, "message": f"No available {_drop_room_type_for_subject(scheduled.subject, scheduled.room)} room was found in {sec.department.name} for the selected time slot."}, status=409)

        scheduled.meeting_time = target_mt
        scheduled.room = selected_room
        scheduled.save(update_fields=["meeting_time", "room"])

    return JsonResponse({"ok": True, "message": "Slot moved successfully."})


@login_required
def saved_park_slot(request, tid, section, day, slot):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST required."}, status=405)

    saved_t = _get_saved_timetable_or_404(tid, request.user)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "message": "Invalid payload."}, status=400)

    move_type = payload.get("move_type", "class")
    source_mt = get_meeting_time(day, slot, user=request.user)
    if not source_mt:
        return JsonResponse({"ok": False, "message": "Invalid time slot."}, status=400)

    try:
        sec = Section.objects.get(section_id=section, user=request.user)
    except Section.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Section not found."}, status=404)

    scheduled = saved_t.slots.filter(
        section=sec,
        meeting_time=source_mt,
        is_lab=(move_type == "lab"),
    ).select_related("subject", "instructor", "second_instructor", "room", "meeting_time").first()
    if not scheduled:
        return JsonResponse({"ok": False, "message": "No slot found at source."}, status=404)

    slot_times = list(scheduled.lab_slots.all().order_by("time")) if scheduled.is_lab else [scheduled.meeting_time]
    if not slot_times:
        slot_times = [scheduled.meeting_time]

    with transaction.atomic():
        _reserve_saved_slot_room(saved_t, sec, slot_times, scheduled.room)
        SavedParkingSlot.objects.create(
            timetable=saved_t,
            section=sec,
            subject=scheduled.subject,
            instructor=scheduled.instructor,
            second_instructor=scheduled.second_instructor,
            original_room=scheduled.room,
            original_meeting_time=scheduled.meeting_time,
            is_lab=scheduled.is_lab,
            slot_span=len(slot_times),
        )
        scheduled.delete()

    return JsonResponse({
        "ok": True,
        "message": "Slot moved to parking.",
    })


@login_required
def saved_restore_parked_slot(request, tid, parking_id):
    if request.method != "POST":
        return JsonResponse({"ok": False, "message": "POST required."}, status=405)

    saved_t = _get_saved_timetable_or_404(tid, request.user)

    try:
        payload = json.loads(request.body.decode("utf-8"))
    except (json.JSONDecodeError, UnicodeDecodeError):
        return JsonResponse({"ok": False, "message": "Invalid payload."}, status=400)

    target_day = payload.get("target_day")
    target_slot = payload.get("target_slot")
    target_section = payload.get("target_section")
    if not target_day or not target_slot or not target_section:
        return JsonResponse({"ok": False, "message": "Missing target section/day/slot."}, status=400)

    parked = saved_t.parked_slots.select_related(
        "section",
        "subject",
        "instructor",
        "second_instructor",
    ).filter(id=parking_id).first()
    if not parked:
        return JsonResponse({"ok": False, "message": "Parking item not found."}, status=404)
    if parked.section.section_id != target_section:
        return JsonResponse({"ok": False, "message": "Parking is locked to its section."}, status=409)

    target_mt = get_meeting_time(target_day, target_slot, user=request.user)
    if not target_mt:
        return JsonResponse({"ok": False, "message": "Invalid target time slot."}, status=400)

    try:
        sec = Section.objects.get(section_id=target_section, user=request.user)
    except Section.DoesNotExist:
        return JsonResponse({"ok": False, "message": "Section not found."}, status=404)

    target_times = _get_slot_sequence(target_day, target_slot, parked.slot_span, request.user)
    if not target_times:
        return JsonResponse({"ok": False, "message": "Target span does not fit in the timetable."}, status=400)

    reservations = list(
        SavedSlotRoomReservation.objects.filter(
            timetable=saved_t,
            section=sec,
            meeting_time__in=target_times,
        ).select_related("room", "meeting_time")
    )
    reservation_map = {reservation.meeting_time_id: reservation for reservation in reservations}
    if len(reservation_map) != len(target_times):
        return JsonResponse({"ok": False, "message": "Target slot has no parked-room context yet."}, status=409)

    room = reservation_map[target_times[0].id].room
    if any(reservation.room_id != room.id for reservation in reservations):
        return JsonResponse({"ok": False, "message": "Target parking room context is inconsistent."}, status=409)

    for meeting_time in target_times:
        section_conflict = saved_t.slots.filter(section=sec, meeting_time=meeting_time).exists()
        if section_conflict:
            return JsonResponse({"ok": False, "message": f"Section already has a slot at {target_day} slot {meeting_time.time}."}, status=409)

        teacher_conflict = saved_t.slots.filter(
            Q(instructor=parked.instructor) | Q(second_instructor=parked.instructor),
            meeting_time=meeting_time,
        ).exists()
        if teacher_conflict:
            return JsonResponse({"ok": False, "message": f"Teacher already has a slot at {target_day} slot {meeting_time.time}."}, status=409)

        if parked.second_instructor:
            second_conflict = saved_t.slots.filter(
                Q(instructor=parked.second_instructor) | Q(second_instructor=parked.second_instructor),
                meeting_time=meeting_time,
            ).exists()
            if second_conflict:
                return JsonResponse({"ok": False, "message": f"Second teacher already has a slot at {target_day} slot {meeting_time.time}."}, status=409)

    with transaction.atomic():
        scheduled = ScheduledSlot.objects.create(
            timetable=saved_t,
            section=sec,
            subject=parked.subject,
            instructor=parked.instructor,
            second_instructor=parked.second_instructor,
            room=room,
            meeting_time=target_times[0],
            is_lab=parked.is_lab,
        )
        if parked.is_lab:
            scheduled.lab_slots.set(target_times)

        _clear_saved_slot_room_reservations(saved_t, sec, target_times)
        parked.delete()

    return JsonResponse({"ok": True, "message": "Parking item restored successfully."})


# ================================================================
# PUBLISH / TEACHER READ-ONLY VIEWS
# ================================================================

@login_required
def publish_timetable(request, tid):
    """HOD publishes a saved timetable with a custom code."""
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    if request.method == "POST":
        code = re.sub(r"\s+", "", request.POST.get("publish_code", "")).upper().strip()
        if not code:
            messages.error(request, "Please enter a publish code.")
            return redirect("saved_timetable", tid=tid)
        conflict = SavedTimetable.objects.filter(
            publish_code__iexact=code, is_published=True
        ).exclude(id=tid).exists()
        if conflict:
            messages.error(request, "This code is already used for another timetable.")
            return redirect("saved_timetable", tid=tid)
        saved_t.is_published = True
        saved_t.publish_code = code
        saved_t.save(update_fields=["is_published", "publish_code"])
        messages.success(request, f"Timetable published with code: {code}")
        return redirect("saved_timetable_publish_notifications", tid=tid)
    return redirect("saved_timetable", tid=tid)


@login_required
def saved_timetable_publish_notifications(request, tid):
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    if not saved_t.is_published or not saved_t.publish_code:
        messages.error(request, "Publish this timetable first to send notifications.")
        return redirect("saved_timetable", tid=tid)
    teacher_recipients = _collect_saved_teacher_recipients(saved_t)
    search_teacher_recipients = _collect_admin_teacher_recipients()
    coordinator_recipients = _coordinator_recipients()
    context = {
        "saved": saved_t,
        "teacher_recipients": teacher_recipients,
        "teacher_count": len(teacher_recipients),
        "search_teacher_recipients": search_teacher_recipients,
        "search_teacher_count": len(search_teacher_recipients),
        "coordinator_recipients": coordinator_recipients,
        "coordinator_count": len(coordinator_recipients),
        "timetable_title": _saved_timetable_title(saved_t),
    }
    return render(request, "publish_notify_center.html", context)


def _send_publish_notifications(request, saved_t, recipients, role_label, action_label):
    sent_count = 0
    failed_count = 0
    timetable_title = _saved_timetable_title(saved_t)
    for recipient in recipients:
        try:
            _send_publish_notification_email(
                recipient_name=recipient.get("name") or role_label,
                recipient_email=recipient.get("email") or "",
                publish_code=saved_t.publish_code,
                request=request,
                role_label=role_label,
                timetable_title=timetable_title,
            )
            sent_count += 1
        except Exception:
            failed_count += 1
            logger.exception("Publish notification email failed for %s", recipient.get("email"))
    _send_publish_summary_email(request, saved_t, action_label, sent_count, failed_count)
    return sent_count, failed_count


@login_required
def publish_notify_all_teachers(request, tid):
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    if request.method != "POST":
        return redirect("saved_timetable_publish_notifications", tid=tid)
    recipients = _collect_saved_teacher_recipients(saved_t)
    sent_count, failed_count = _send_publish_notifications(request, saved_t, recipients, "Teacher", "Teacher notification broadcast")
    messages.success(request, f"Teacher notifications sent: {sent_count}. Failed: {failed_count}.")
    return redirect("saved_timetable_publish_notifications", tid=tid)


@login_required
def publish_notify_single_teacher(request, tid):
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    if request.method != "POST":
        return redirect("saved_timetable_publish_notifications", tid=tid)
    teacher_id = request.POST.get("teacher_id", "").strip()
    recipients = _collect_admin_teacher_recipients()
    selected = next((item for item in recipients if str(item.get("id")) == teacher_id), None)
    if not selected:
        messages.error(request, "Select a valid teacher to notify.")
        return redirect("saved_timetable_publish_notifications", tid=tid)
    sent_count, failed_count = _send_publish_notifications(request, saved_t, [selected], "Teacher", f"Single teacher notification for {selected.get('name')}")
    if sent_count:
        messages.success(request, f"Publish code mailed to {selected.get('name')}.")
    else:
        messages.error(request, f"Could not send email to {selected.get('name')}. Failed: {failed_count}.")
    return redirect("saved_timetable_publish_notifications", tid=tid)


@login_required
def publish_notify_all_coordinators(request, tid):
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    if request.method != "POST":
        return redirect("saved_timetable_publish_notifications", tid=tid)
    recipients = _coordinator_recipients()
    sent_count, failed_count = _send_publish_notifications(request, saved_t, recipients, "Timetable Coordinator", "Coordinator notification broadcast")
    messages.success(request, f"Coordinator notifications sent: {sent_count}. Failed: {failed_count}.")
    return redirect("saved_timetable_publish_notifications", tid=tid)


@login_required
def unpublish_timetable(request, tid):
    """HOD unpublishes a timetable."""
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    if request.method == "POST":
        saved_t.is_published = False
        saved_t.publish_code = ""
        saved_t.save(update_fields=["is_published", "publish_code"])
        messages.success(request, "Timetable unpublished.")
    return redirect("saved_timetable", tid=tid)


@login_required
def teacher_enter_code(request):
    """Legacy route: connect the HOD publish code and return to the timetable page."""
    profile, locked_response = _get_teacher_profile_or_locked_response(request.user)
    if locked_response:
        return render(request, 'role_locked.html', {'current_role': locked_response})

    _ensure_teacher_role(profile)

    if request.method == "POST":
        code = request.POST.get("access_code", "").strip()
        if not code:
            messages.error(request, "Please enter the publish code shared by your HOD.")
            return redirect("teacher_published_timetable")
        timetable, error_message = _connect_teacher_timetable(profile, code)
        if error_message:
            messages.error(request, error_message)
            return redirect("teacher_published_timetable")
        messages.success(request, f"HOD published timetable connected with code {timetable.publish_code}.")
        return redirect("teacher_published_timetable")
    return redirect("teacher_published_timetable")


@login_required
def teacher_view_timetable(request, tid):
    """Read-only timetable view for teachers accessing via publish code."""
    try:
        saved_t = SavedTimetable.objects.get(id=tid, is_published=True)
    except SavedTimetable.DoesNotExist:
        messages.error(request, "This timetable is not available.")
        return redirect("teacher_published_timetable")

    classes, labs = _rebuild_classes_and_labs_from_saved(saved_t)
    owner = saved_t.user
    selected_program = request.GET.get("program", "all")
    classes, labs, selected_program = _filter_entities_by_program(classes, labs, owner, selected_program)
    tables = build_section_tables(classes, labs, user=owner)
    room_tables = build_room_tables(classes, labs, user=owner)
    teacher_tables = build_teacher_tables(classes, labs, user=owner)
    teacher_workloads = _compute_teacher_workloads(classes, labs)

    context = {
        "saved": saved_t,
        "tables": tables,
        "room_tables": room_tables,
        "teacher_tables": teacher_tables,
        "teacher_workloads": teacher_workloads,
        "program_options": _get_program_filter_options(owner),
        "active_program": selected_program,
        "SLOT_LABELS": SLOT_LABELS,
        "can_edit_delete": False,
        "can_substitute": False,
        "can_drag_drop": False,
        "is_readonly": True,
    }
    return render(request, "saved_timetable.html", context)


# ---------------- CRUD VIEWS ----------------
import csv
from django.contrib import messages
from django.shortcuts import redirect, render
from django.db import transaction

@login_required
def addSubjects(request):
    edit_subject = _get_required_step_edit_object(request, Subject, user=request.user) if request.method == "POST" and request.POST.get("edit_id") else _get_step_edit_object(request, Subject, user=request.user)
    form = SubjectForm(request.POST or None, instance=edit_subject, user=request.user)

    # ============================
    # MANUAL ADD SUBJECT
    # ============================
    if request.method == "POST" and "add_subject" in request.POST:
        if form.is_valid():
            subject = form.save(commit=False)
            subject.user = request.user
            raw_room = (subject.room_required or "").strip().lower()
            subject.room_required = {"lab": "Lab", "lecture hall": "Lecture Hall"}.get(raw_room, subject.room_required)
            subject.required_lab_category = normalize_lab_categories_value(subject.required_lab_category)
            subject.specific_rooms = normalize_specific_rooms(subject.specific_rooms)

            if subject.room_required == "Lab" and not subject.required_lab_category:
                messages.error(request, "Lab subjects must have a Required Lab Category.")
                return redirect("addSubjects")
            if subject.room_required not in {"Lab", "Lecture Hall"}:
                subject.required_lab_category = ""

            subject.save()
            form.save_m2m()

            reset_global_schedule_cache(request.user.id)
            messages.success(request, "Subject updated successfully!" if edit_subject else "Subject added successfully!")
            return redirect("addSubjects")

    # ============================
    # CSV UPLOAD SUBJECTS
    # ============================
    if request.method == "POST" and "csv_upload" in request.POST:
        csv_file = request.FILES.get("csv_file")

        if not csv_file or not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a valid CSV file.")
            return redirect("addSubjects")

        try:
            decoded_file = csv_file.read().decode("utf-8").splitlines()
            reader = csv.DictReader(decoded_file, skipinitialspace=True)
            reader.fieldnames = [
                ((name or "").lstrip("\ufeff")).strip()
                for name in (reader.fieldnames or [])
            ]

            fieldnames = reader.fieldnames or []
            required_columns = [
                "department_code",
                "subject_number",
                "subject_name",
                "room_required",
                "classes_per_week",
            ]

            # Header validation
            for col in required_columns:
                if col not in fieldnames:
                    messages.error(request, f"Missing required column: {col}")
                    return redirect("addSubjects")

            if not ({"required_lab_category", "lab_category_required"} & set(fieldnames)):
                messages.error(request, "Missing required column: required_lab_category or lab_category_required")
                return redirect("addSubjects")

            created_count = 0
            skipped_count = 0
            issues = []

            with transaction.atomic():
                for row_number, row in enumerate(reader, start=2):
                    dept_code = (row.get("department_code") or "").strip().upper()
                    subject_number = (row.get("subject_number") or "").strip()
                    subject_name = (row.get("subject_name") or "").strip()
                    raw_room = (row.get("room_required") or "").strip()
                    raw_classes_per_week = (row.get("classes_per_week") or "").strip()

                    if not any((value or "").strip() for value in row.values()):
                        continue
                    if not dept_code:
                        skipped_count += 1
                        _csv_issue(issues, row_number, "department_code is blank")
                        continue
                    if not subject_number:
                        skipped_count += 1
                        _csv_issue(issues, row_number, "subject_number is blank")
                        continue
                    if not subject_name:
                        skipped_count += 1
                        _csv_issue(issues, row_number, "subject_name is blank")
                        continue
                    if not raw_room:
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"room_required is blank for subject '{subject_number}'")
                        continue
                    if not raw_classes_per_week:
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"classes_per_week is blank for subject '{subject_number}'")
                        continue

                    room_required = {"lab": "Lab", "lecture hall": "Lecture Hall"}.get(raw_room.lower(), raw_room)
                    required_lab_category = normalize_lab_categories_value(
                        row.get("required_lab_category") or row.get("lab_category_required")
                    )
                    specific_rooms = normalize_specific_rooms(
                        row.get("specific_equipment/software_lab") or row.get("specific_rooms")
                    )

                    try:
                        subject_department = Department.objects.get(code=dept_code, user=request.user)
                    except Department.DoesNotExist:
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"department_code '{dept_code}' does not exist")
                        continue

                    # Skip duplicates
                    if Subject.objects.filter(
                        subject_number=subject_number,
                        user=request.user,
                        department=subject_department,
                    ).exists():
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"subject_number '{subject_number}' already exists in department '{dept_code}'")
                        continue

                    if room_required == "Lab" and not required_lab_category:
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"subject '{subject_number}' is Lab but required_lab_category is blank")
                        continue
                    if room_required not in {"Lab", "Lecture Hall"}:
                        required_lab_category = ""

                    try:
                        classes_per_week = int(raw_classes_per_week)
                    except (TypeError, ValueError):
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"classes_per_week '{raw_classes_per_week}' is not a whole number")
                        continue

                    raw_max_students = (row.get("max_numb_students", 70) or 70)
                    raw_duration = row.get("duration") or row.get("duration (in hr)") or 1
                    try:
                        max_numb_students = int(raw_max_students)
                    except (TypeError, ValueError):
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"max_numb_students '{raw_max_students}' is not a whole number")
                        continue

                    try:
                        duration = int(raw_duration)
                    except (TypeError, ValueError):
                        skipped_count += 1
                        _csv_issue(issues, row_number, f"duration '{raw_duration}' is not a whole number")
                        continue

                    subject = Subject.objects.create(
                        user=request.user,
                        subject_number=subject_number,
                        subject_name=subject_name,
                        department=subject_department,
                        room_required=room_required,
                        required_lab_category=required_lab_category,
                        specific_rooms=specific_rooms,
                        classes_per_week=classes_per_week,
                        max_numb_students=max_numb_students,
                        duration=duration,
                    )

                    created_count += 1

            reset_global_schedule_cache(request.user.id)
            messages.success(
                request, f"{created_count} subjects uploaded successfully! {skipped_count} skipped."
            )
            _emit_csv_issues(request, issues)

        except Exception as e:
            messages.error(request, f"CSV upload failed: {str(e)}")

        return redirect("addSubjects")

    return render(request, "addSubjects.html", {"form": form, "edit_subject": edit_subject})




@login_required
def subject_list_view(request):
    subjects = Subject.objects.filter(user=request.user).select_related("department").prefetch_related("instructors").order_by("department__code", "subject_number")
    return render(request, 'subjectslist.html', {'subjects': subjects})


@login_required
def delete_subject(request, pk):
    if request.method == 'POST':
        Subject.objects.filter(pk=pk, user=request.user).delete()
        reset_global_schedule_cache(request.user.id)
        return redirect('editsubject')


@login_required
def delete_all_subjects(request):
    return _delete_all_step_entries(
        request,
        Subject.objects.filter(user=request.user),
        "addSubjects",
        "subjects",
    )


@login_required
def addInstructor(request):
    edit_instructor = _get_required_step_edit_object(request, Instructor, user=request.user) if request.method == "POST" and request.POST.get("edit_id") else _get_step_edit_object(request, Instructor, user=request.user)
    form = InstructorForm(request.POST or None, instance=edit_instructor, user=request.user)

    # ================================
    # FETCH FROM ERP API
    # ================================
    if request.method == "POST" and "fetch_api" in request.POST:
        url = "http://localhost:1000/api/teachers/"
        try:
            response = requests.get(url)
            response.raise_for_status()
            request.session["api_teachers"] = response.json()
            messages.success(request, "Teachers fetched successfully!")
        except Exception as e:
            print("API ERROR:", e)
            messages.error(request, "Failed to fetch teachers from API.")
        return redirect("addInstructors")

    # ================================
    # CONFIRM ADD FROM API
    # ================================
    if request.method == "POST" and "confirm_add_api" in request.POST:
        data = request.session.get("api_teachers", [])
        added = 0

        for t in data:
            uid = t.get("teacherId")
            name = t.get("teacherName")

            if uid and name and not Instructor.objects.filter(uid=uid, user=request.user).exists():
                designation, workload = teacher_payload(name)
                Instructor.objects.create(
                    user=request.user,
                    uid=uid,
                    name=name,
                    designation=designation,
                    max_workload=workload,
                )
                added += 1

        request.session["api_teachers"] = []
        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"{added} teachers added successfully!")
        return redirect("addInstructors")

    # ================================
    # MANUAL ADD TEACHER
    # ================================
    if request.method == "POST" and "add_teacher" in request.POST:
        if form.is_valid():
            instructor = form.save(commit=False)
            instructor.user = request.user
            instructor.save()
            reset_global_schedule_cache(request.user.id)
            messages.success(request, "Teacher updated successfully!" if edit_instructor else "Teacher added successfully!")
        else:
            messages.error(request, "Invalid input.")
        return redirect("addInstructors")

    # ================================
    # CSV UPLOAD
    # ================================
    if request.method == "POST" and "csv_upload" in request.POST:
        csv_file = request.FILES.get("csv_file")

        if not csv_file or not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a valid CSV file.")
            return redirect("addInstructors")

        import csv
        reader = csv.reader(csv_file.read().decode("utf-8").splitlines())

        first = True

        added = 0
        skipped = 0
        issues = []
        for row_number, row in enumerate(reader, start=1):
            if not row:
                continue

            if first:
                first = False
                continue

            if len(row) < 2:
                skipped += 1
                _csv_issue(issues, row_number, "expected at least uid and name columns")
                continue

            uid, name = row[0].strip(), row[1].strip()
            designation = row[2].strip() if len(row) > 2 else ""
            max_workload = row[3].strip() if len(row) > 3 else ""
            email = row[4].strip() if len(row) > 4 else ""
            contact_number = row[5].strip() if len(row) > 5 else ""
            department_code = row[6].strip() if len(row) > 6 else ""
            if not uid:
                skipped += 1
                _csv_issue(issues, row_number, "uid is blank")
                continue
            if not name:
                skipped += 1
                _csv_issue(issues, row_number, f"name is blank for uid '{uid}'")
                continue
            if not email:
                skipped += 1
                _csv_issue(issues, row_number, f"email is blank for uid '{uid}'")
                continue
            if not contact_number:
                skipped += 1
                _csv_issue(issues, row_number, f"contact_number is blank for uid '{uid}'")
                continue
            if Instructor.objects.filter(uid=uid, user=request.user).exists():
                skipped += 1
                _csv_issue(issues, row_number, f"uid '{uid}' already exists")
                continue
            resolved_designation, resolved_workload = teacher_payload(
                name,
                designation,
                max_workload,
            )
            department_obj = None
            if department_code:
                department_obj = Department.objects.filter(
                    user=request.user, code__iexact=department_code
                ).first()
                if department_obj is None:
                    department_obj = Department.objects.filter(
                        user=request.user, name__iexact=department_code
                    ).first()
            Instructor.objects.create(
                user=request.user,
                uid=uid,
                name=name,
                email=email,
                contact_number=contact_number,
                designation=resolved_designation,
                max_workload=resolved_workload,
                department=department_obj,
            )
            added += 1

        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"{added} teachers imported successfully! {skipped} skipped.")
        _emit_csv_issues(request, issues)
        return redirect("addInstructors")

    popup_data = request.session.get("api_teachers", [])
    return render(request, "addInstructors.html", {
        "form": form,
        "edit_instructor": edit_instructor,
        "popup_data": popup_data,
    })

@login_required
def map_section_subjects(request):
    sections = Section.objects.filter(user=request.user).order_by("section_id")
    subjects = Subject.objects.filter(user=request.user).order_by("subject_number")
    edit_mapping = None
    raw_edit_mapping_id = request.POST.get("edit_mapping_id") if request.method == "POST" else request.GET.get("edit")
    if raw_edit_mapping_id:
        try:
            edit_mapping = SectionSubjectMapping.objects.filter(pk=int(raw_edit_mapping_id), section__user=request.user).select_related("section", "subject").first()
        except (TypeError, ValueError):
            edit_mapping = None

    def parse_group_count(raw_value):
        value = str(raw_value or "").strip()
        if not value:
            return 1
        try:
            parsed = int(value)
        except (TypeError, ValueError):
            raise ValueError("Group count must be a whole number.")
        if parsed < 1:
            raise ValueError("Group count must be at least 1.")
        return parsed

    def resolve_section(section_identifier):
        section_identifier = (section_identifier or "").strip()
        if not section_identifier:
            raise Section.DoesNotExist

        try:
            return Section.objects.get(section_id=section_identifier, user=request.user)
        except Section.DoesNotExist:
            if section_identifier.isdigit():
                return Section.objects.get(pk=int(section_identifier), user=request.user)
            raise

    if request.method == "POST" and "manual_add" in request.POST:
        section_identifier = request.POST.get("section_id", "").strip()
        selected_subject_ids = request.POST.getlist("subjects")
        raw_group_count = request.POST.get("group_count", "1")
        raw_elective_section_ids = request.POST.get("elective_section_id", "")

        if not section_identifier or not selected_subject_ids:
            messages.error(request, "Please select a section and at least one subject.")
            return redirect("map_section_subjects")
        if edit_mapping is not None and len(selected_subject_ids) != 1:
            messages.error(request, "Select exactly one subject while updating a mapping row.")
            return redirect(f"{reverse('map_section_subjects')}?edit={edit_mapping.id}")

        try:
            group_count = parse_group_count(raw_group_count)
        except ValueError as exc:
            messages.error(request, str(exc))
            return redirect("map_section_subjects")

        elective_section_ids = normalize_section_id_list(raw_elective_section_ids)

        try:
            section = resolve_section(section_identifier)
        except Section.DoesNotExist:
            messages.error(request, f"Section not found: {section_identifier}")
            return redirect("map_section_subjects")

        valid_subjects = list(Subject.objects.filter(pk__in=selected_subject_ids, user=request.user))
        if not valid_subjects:
            messages.error(request, "No valid subjects were selected.")
            return redirect("map_section_subjects")

        created = 0
        updated = 0
        if edit_mapping is not None:
            subj = valid_subjects[0]
            edit_mapping.section = section
            edit_mapping.subject = subj
            edit_mapping.group_count = group_count
            edit_mapping.elective_section_ids = elective_section_ids
            edit_mapping.save()
            updated = 1
        else:
            for subj in valid_subjects:
                mapping, was_created = SectionSubjectMapping.objects.update_or_create(
                    section=section,
                    subject=subj,
                    defaults={
                        "group_count": group_count,
                        "elective_section_ids": elective_section_ids,
                    },
                )
                if was_created:
                    created += 1
                elif mapping.group_count == group_count:
                    updated += 1

        reset_global_schedule_cache(request.user.id)
        messages.success(
            request,
            f"{created} subject mappings saved for {section.section_id}. {updated} group counts updated."
        )
        return redirect("map_section_subjects")

    if request.method == "POST" and "csv_upload" in request.POST:
        csv_file = request.FILES.get("csv_file")

        if not csv_file:
            messages.error(request, "No CSV file selected.")
            return redirect("map_section_subjects")

        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Invalid file format. Upload CSV only.")
            return redirect("map_section_subjects")

        decoded = csv_file.read().decode("utf-8").splitlines()
        reader = csv.reader(decoded)

        added = 0
        updated = 0
        skipped = 0
        first = True
        issues = []

        for row_number, row in enumerate(reader, start=1):
            if not row:
                continue

            if first:
                first = False
                continue

            if len(row) < 2:
                skipped += 1
                _csv_issue(issues, row_number, "expected at least section_id and subject_number columns")
                continue

            section_identifier = row[0].strip()
            subject_number = row[1].strip()
            elective_section_ids = normalize_section_id_list(row[3] if len(row) > 3 else "")

            if not section_identifier:
                skipped += 1
                _csv_issue(issues, row_number, "section_id is blank")
                continue
            if not subject_number:
                skipped += 1
                _csv_issue(issues, row_number, f"subject_number is blank for section '{section_identifier}'")
                continue

            try:
                group_count = parse_group_count(row[2] if len(row) > 2 else "1")
            except ValueError as exc:
                skipped += 1
                _csv_issue(issues, row_number, str(exc))
                continue

            try:
                section = resolve_section(section_identifier)
            except Section.DoesNotExist:
                skipped += 1
                _csv_issue(issues, row_number, f"section not found '{section_identifier}'")
                continue

            try:
                subj = _resolve_subject_for_user(subject_number, request.user, department=section.department)
            except Subject.DoesNotExist:
                skipped += 1
                _csv_issue(issues, row_number, f"subject not found '{subject_number}'")
                continue

            _, was_created = SectionSubjectMapping.objects.update_or_create(
                section=section,
                subject=subj,
                defaults={
                    "group_count": group_count,
                    "elective_section_ids": elective_section_ids,
                },
            )
            if was_created:
                added += 1
            else:
                updated += 1

        reset_global_schedule_cache(request.user.id)
        messages.success(
            request,
            f"{added} section-subject mappings added. {updated} group counts updated. {skipped} skipped."
        )
        _emit_csv_issues(request, issues)
        return redirect("map_section_subjects")

    saved_prefills = []
    for prefill in SavedPrefill.objects.filter(user=request.user):
        snapshot = prefill.snapshot or {}
        slot_count = len(snapshot.get("classes") or []) + len(snapshot.get("labs") or []) + len(snapshot.get("parking_items") or [])
        section_ids = list(snapshot.get("section_ids") or [])
        saved_prefills.append({
            "item": prefill,
            "slot_count": slot_count,
            "section_count": len(section_ids),
        })

    return render(request, "map_section_subjects.html", {
        "sections": sections,
        "subjects": subjects,
        "saved_prefills": saved_prefills,
        "edit_mapping": edit_mapping,
    })


@login_required
def generate_selected_prefills(request):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    raw_prefill_ids = request.POST.getlist("prefill_ids")
    prefill_ids = []
    for raw_id in raw_prefill_ids:
        try:
            prefill_ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue

    prefills = list(SavedPrefill.objects.filter(user=request.user, pk__in=prefill_ids))
    if not prefills:
        messages.error(request, "Select at least one saved prefill before generating with prefilled slots.")
        return redirect("map_section_subjects")

    snapshot = _combine_prefill_snapshots(prefills)
    if not _activate_prefill_snapshot(request, snapshot):
        messages.error(request, "Selected saved prefills do not contain section data.")
        return redirect("map_section_subjects")

    _clear_active_saved_prefill(request)
    messages.success(request, f"Generation will use {len(prefills)} saved prefill(s).")
    return redirect(f"{reverse('generate_timetable_loading')}?use_pso=1&use_prefilled_slots=1")


@login_required
def generate_without_prefills(request):
    state = _get_user_state(request)
    state["prefill_mode"] = False
    state["prefill_locked_classes"] = []
    state["prefill_locked_labs"] = []
    for key in ("prefill_mode", "prefill_section_ids", "prefill_saved_slots", "current_index"):
        request.session.pop(key, None)
    _clear_active_saved_prefill(request)
    return redirect(f"{reverse('generate_timetable_loading')}?use_pso=1")


@login_required
def view_section_subjects(request):
    sections = Section.objects.filter(user=request.user).select_related("department").order_by("department__code", "section_id")
    section_mappings = [
        {
            "section": section,
            "subjects": list(
                SectionSubjectMapping.objects.filter(section=section)
                .select_related("subject", "subject__department")
                .order_by("subject__subject_number")
            ),
        }
        for section in sections
    ]
    return render(
        request,
        "view_section_subjects.html",
        {"section_mappings": section_mappings},
    )


@login_required
def view_teacher_subject_mappings(request):
    mappings = list(
        SectionSubjectInstructor.objects.filter(user=request.user)
        .select_related("section", "section__department", "subject", "subject__department", "instructor", "second_instructor")
        .order_by("section__department__code", "section__section_id", "subject__subject_number")
    )
    for mapping in mappings:
        mapping.primary_teacher_uids = _teacher_uid_string_from_ids(
            getattr(mapping, "group_instructor_ids", []),
            request.user,
            getattr(mapping, "instructor", None),
        )
        mapping.primary_teacher_names = _teacher_name_string_from_ids(
            getattr(mapping, "group_instructor_ids", []),
            request.user,
            getattr(mapping, "instructor", None),
        )
        mapping.second_teacher_uids = _teacher_uid_string_from_ids(
            getattr(mapping, "group_second_instructor_ids", []),
            request.user,
            getattr(mapping, "second_instructor", None),
        )
        mapping.second_teacher_names = _teacher_name_string_from_ids(
            getattr(mapping, "group_second_instructor_ids", []),
            request.user,
            getattr(mapping, "second_instructor", None),
        )
    return render(request, "view_teacher_subject_mappings.html", {"mappings": mappings})

@login_required
def map_teacher_subjects(request):
    edit_mapping = None
    raw_edit_mapping_id = request.POST.get("edit_mapping_id") if request.method == "POST" else request.GET.get("edit")
    if raw_edit_mapping_id:
        try:
            edit_mapping = SectionSubjectInstructor.objects.filter(
                pk=int(raw_edit_mapping_id),
                user=request.user,
            ).select_related("section", "subject", "instructor", "second_instructor").first()
        except (TypeError, ValueError):
            edit_mapping = None

    def _resolve_mapping_instructor(value):
        value = (value or "").strip()
        if not value:
            return None
        try:
            return Instructor.objects.get(uid=value, user=request.user)
        except Instructor.DoesNotExist:
            try:
                return Instructor.objects.get(name=value, user=request.user)
            except (Instructor.DoesNotExist, Instructor.MultipleObjectsReturned):
                return None

    def _split_mapping_teacher_tokens(raw_value):
        text = str(raw_value or "").strip()
        if not text:
            return []
        return [token.strip() for token in text.split(";")]

    if request.method == "POST" and "manual_save_mapping" in request.POST:
        section_pk = request.POST.get("section_id")
        subject_pk = request.POST.get("subject_id")
        primary_raw = request.POST.get("primary_teacher_uids", "")
        secondary_raw = request.POST.get("second_teacher_uids", "")
        shared_lab = bool(request.POST.get("shared_lab"))

        section = Section.objects.filter(pk=section_pk, user=request.user).first()
        subject = Subject.objects.filter(pk=subject_pk, user=request.user).first()
        if section is None or subject is None:
            messages.error(request, "Select a valid section and subject.")
            return redirect("map_teacher_subjects")

        primary_tokens = _split_mapping_teacher_tokens(primary_raw)
        secondary_tokens = _split_mapping_teacher_tokens(secondary_raw)
        if not primary_tokens:
            messages.error(request, "Enter at least one primary teacher UID.")
            return redirect("map_teacher_subjects")

        section_subject_mapping = SectionSubjectMapping.objects.filter(section=section, subject=subject).only("group_count").first()
        group_count = max(
            1,
            section_subject_mapping.group_count if section_subject_mapping else 1,
            len(primary_tokens),
            len(secondary_tokens),
        )
        if len(primary_tokens) == 1 and group_count > 1:
            primary_tokens = primary_tokens * group_count
        elif len(primary_tokens) < group_count:
            primary_tokens = primary_tokens + ([""] * (group_count - len(primary_tokens)))

        if secondary_tokens and len(secondary_tokens) == 1 and group_count > 1:
            secondary_tokens = secondary_tokens * group_count
        elif len(secondary_tokens) < group_count:
            secondary_tokens = secondary_tokens + ([""] * (group_count - len(secondary_tokens)))

        primary_teachers = []
        for index, token in enumerate(primary_tokens[:group_count], start=1):
            teacher = _resolve_mapping_instructor(token)
            if teacher is None:
                messages.error(request, f"Primary teacher not found for group {index}: {token}")
                return redirect("map_teacher_subjects")
            primary_teachers.append(teacher)

        secondary_teachers = []
        for index, token in enumerate(secondary_tokens[:group_count], start=1):
            if not token:
                secondary_teachers.append(None)
                continue
            teacher = _resolve_mapping_instructor(token)
            if teacher is None:
                messages.error(request, f"Second teacher not found for group {index}: {token}")
                return redirect("map_teacher_subjects")
            if teacher.id == primary_teachers[index - 1].id:
                messages.error(request, f"Primary and second teacher cannot be same for group {index}.")
                return redirect("map_teacher_subjects")
            secondary_teachers.append(teacher)

        if secondary_tokens and any(secondary_teachers):
            shared_lab = True

        saved_mapping, _ = SectionSubjectInstructor.objects.update_or_create(
            user=request.user,
            section=section,
            subject=subject,
            defaults={
                "instructor": primary_teachers[0],
                "second_instructor": next((teacher for teacher in secondary_teachers if teacher is not None), None) if shared_lab else None,
                "group_instructor_ids": [teacher.id for teacher in primary_teachers],
                "group_second_instructor_ids": [teacher.id if teacher is not None else None for teacher in secondary_teachers] if shared_lab else [],
            },
        )
        if edit_mapping is not None and edit_mapping.id != saved_mapping.id:
            edit_mapping.delete()
        for teacher in primary_teachers + [teacher for teacher in secondary_teachers if teacher is not None]:
            subject.instructors.add(teacher)
        reset_global_schedule_cache(request.user.id)
        messages.success(request, "Teacher-subject mapping updated successfully!" if edit_mapping else "Teacher-subject mapping saved successfully!")
        return redirect("map_teacher_subjects")
    # Step 7 CSV supports legacy, shared-lab, and group-wise teacher mappings.

    # =========================
    # CSV UPLOAD HANDLER
    # =========================
    if request.method == "POST" and "csv_upload" in request.POST:

        csv_file = request.FILES.get("csv_file")

        if not csv_file:
            messages.error(request, "No CSV file selected.")
            return redirect("map_teacher_subjects")

        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Invalid file format. Upload CSV only.")
            return redirect("map_teacher_subjects")

        decoded = csv_file.read().decode("utf-8").splitlines()
        reader = csv.reader(decoded)

        added = 0
        skipped = 0
        first = True
        validation_errors = []

        def _resolve_instructor(value):
            try:
                return Instructor.objects.get(uid=value, user=request.user)
            except Instructor.DoesNotExist:
                try:
                    return Instructor.objects.get(name=value, user=request.user)
                except (Instructor.DoesNotExist, Instructor.MultipleObjectsReturned):
                    return None

        def _split_teacher_tokens(raw_value):
            if raw_value is None:
                return []
            text = str(raw_value).strip()
            if not text:
                return []
            return [token.strip() for token in text.split(";")]

        def _resolve_group_assignments(raw_value, slot_count, field_label, allow_empty=False):
            tokens = _split_teacher_tokens(raw_value)
            if not tokens:
                return [None] * slot_count

            if len(tokens) == 1 and slot_count > 1:
                tokens = tokens * slot_count
            elif len(tokens) < slot_count:
                tokens = tokens + ([""] * (slot_count - len(tokens)))

            resolved = []
            for index, token in enumerate(tokens[:slot_count], start=1):
                if not token:
                    if allow_empty:
                        resolved.append(None)
                        continue
                    validation_errors.append(
                        f"Row {row_number}: {field_label} missing for group {index}"
                    )
                    return None
                instructor_obj = _resolve_instructor(token)
                if not instructor_obj:
                    validation_errors.append(
                        f"Row {row_number}: {field_label} not found '{token}'"
                    )
                    return None
                resolved.append(instructor_obj)
            return resolved

        valid_bool_values = {"", "true", "false", "1", "0", "yes", "no"}
        row_number = 0

        for row in reader:
            row_number += 1
            if not row:
                continue

            if first:
                first = False
                continue

            if len(row) < 3:
                skipped += 1
                validation_errors.append(
                    f"Row {row_number}: expected at least section_id, subject_number, instructor_uid columns"
                )
                continue

            section_id     = row[0].strip()
            subject_number = row[1].strip()
            instructor_uid = row[2].strip()

            if not section_id:
                skipped += 1
                validation_errors.append(f"Row {row_number}: section_id is blank")
                continue
            if not subject_number:
                skipped += 1
                validation_errors.append(f"Row {row_number}: subject_number is blank")
                continue
            if not instructor_uid:
                skipped += 1
                validation_errors.append(
                    f"Row {row_number}: primary instructor is blank for subject '{subject_number}'"
                )
                continue

            # Optional shared-lab columns (col 3 = shared_lab bool, col 4 = second instructor uid)
            shared_lab_flag = False
            second_instructor_uid = ""
            raw_shared_value = ""
            if len(row) >= 4:
                raw_shared_value = row[3].strip().lower()
                if raw_shared_value not in valid_bool_values:
                    skipped += 1
                    validation_errors.append(
                        f"Row {row_number}: invalid shared_lab value '{row[3].strip()}'"
                    )
                    continue
                shared_lab_flag = raw_shared_value in ("true", "1", "yes")
            if len(row) >= 5:
                second_instructor_uid = row[4].strip()

            # -------------------------
            # VALIDATION
            # -------------------------
            try:
                section = Section.objects.get(section_id=section_id, user=request.user)
            except Section.DoesNotExist:
                skipped += 1
                validation_errors.append(
                    f"Row {row_number}: section not found '{section_id}'"
                )
                continue

            try:
                subj = _resolve_subject_for_user(subject_number, request.user, department=section.department)
            except Subject.DoesNotExist:
                skipped += 1
                validation_errors.append(
                    f"Row {row_number}: subject not found '{subject_number}'"
                )
                continue

            section_group_count = 1
            section_subject_mapping = SectionSubjectMapping.objects.filter(section=section, subject=subj).only("group_count").first()
            if section_subject_mapping is not None:
                section_group_count = max(1, section_subject_mapping.group_count or 1)

            primary_tokens = _split_teacher_tokens(instructor_uid)
            secondary_tokens = _split_teacher_tokens(second_instructor_uid)
            if section_group_count > 1:
                inferred_group_count = max(
                    section_group_count,
                    len(primary_tokens) if primary_tokens else 0,
                    len(secondary_tokens) if secondary_tokens else 0,
                    1,
                )
            else:
                inferred_group_count = max(
                    len(primary_tokens) if primary_tokens else 0,
                    len(secondary_tokens) if secondary_tokens else 0,
                    1,
                )

            primary_group_instructors = _resolve_group_assignments(
                instructor_uid,
                inferred_group_count,
                "primary instructor",
                allow_empty=False,
            )
            if primary_group_instructors is None:
                skipped += 1
                validation_errors.append(
                    f"Row {row_number}: primary instructor not found '{instructor_uid}'"
                )
                continue

            instructor = primary_group_instructors[0]

            # Resolve second instructor when shared_lab flag is enabled
            second_instructor = None
            second_group_instructors = [None] * inferred_group_count
            if shared_lab_flag:
                if not second_instructor_uid:
                    skipped += 1
                    validation_errors.append(
                        f"Row {row_number}: shared_lab=true but second_instructor_uid missing"
                    )
                    continue
                second_group_instructors = _resolve_group_assignments(
                    second_instructor_uid,
                    inferred_group_count,
                    "second instructor",
                    allow_empty=True,
                )
                if second_group_instructors is None:
                    skipped += 1
                    continue
                second_instructor = next((teacher for teacher in second_group_instructors if teacher is not None), None)
                for index, primary_teacher in enumerate(primary_group_instructors):
                    paired_second = second_group_instructors[index] if index < len(second_group_instructors) else None
                    if paired_second is not None and paired_second.id == primary_teacher.id:
                        skipped += 1
                        validation_errors.append(
                            f"Row {row_number}: primary and second instructor cannot be same for group {index + 1}"
                        )
                        second_group_instructors = None
                        break
                if second_group_instructors is None:
                    continue

            # -------------------------
            # CREATE / UPDATE MAPPING
            # -------------------------
            SectionSubjectInstructor.objects.update_or_create(
                user=request.user,
                section=section,
                subject=subj,
                defaults={
                    "instructor": instructor,
                    "second_instructor": second_instructor,
                    "group_instructor_ids": [teacher.id for teacher in primary_group_instructors],
                    "group_second_instructor_ids": [teacher.id if teacher is not None else None for teacher in second_group_instructors],
                },
            )
            # Also ensure instructor is in Subject.instructors (for validation)
            for teacher in primary_group_instructors:
                subj.instructors.add(teacher)
            for teacher in second_group_instructors:
                if teacher is not None:
                    subj.instructors.add(teacher)
            added += 1

        messages.success(
            request,
            f"{added} section–subject–teacher mappings saved. {skipped} skipped."
        )
        if validation_errors:
            _emit_csv_issues(request, validation_errors)
        reset_global_schedule_cache(request.user.id)
        return redirect("map_teacher_subjects")

    # =========================
    # DISPLAY EXISTING MAPPINGS
    # =========================
    mappings = SectionSubjectInstructor.objects.filter(
        user=request.user
    ).select_related(
        "section", "subject", "instructor", "second_instructor"
    ).order_by("section__section_id", "subject__subject_number")

    return render(
        request,
        "map_teacher_subjects.html",
        {
            "mappings": mappings,
            "sections": Section.objects.filter(user=request.user).order_by("section_id"),
            "subjects": Subject.objects.filter(user=request.user).order_by("subject_number"),
            "instructors": Instructor.objects.filter(user=request.user).order_by("uid"),
            "edit_mapping": edit_mapping,
            "edit_primary_teacher_uids": _teacher_uid_string_from_ids(getattr(edit_mapping, "group_instructor_ids", []), request.user, getattr(edit_mapping, "instructor", None)) if edit_mapping else "",
            "edit_second_teacher_uids": _teacher_uid_string_from_ids(getattr(edit_mapping, "group_second_instructor_ids", []), request.user, getattr(edit_mapping, "second_instructor", None)) if edit_mapping else "",
        },
    )


@login_required
def delete_all_section_subject_mappings(request):
    return _delete_all_step_entries(
        request,
        SectionSubjectMapping.objects.filter(section__user=request.user),
        "map_section_subjects",
        "section-subject mappings",
    )


def _resolve_subject_for_user(subject_number, user, department=None):
    subject_number = (subject_number or "").strip()
    if not subject_number:
        raise Subject.DoesNotExist

    queryset = Subject.objects.filter(subject_number=subject_number, user=user)
    if department is not None:
        department_match = queryset.filter(department=department).first()
        if department_match is not None:
            return department_match

    match = queryset.first()
    if match is None:
        raise Subject.DoesNotExist
    return match


def _resolve_room_for_user(room_number, user, department=None, case_insensitive=False):
    room_number = (room_number or "").strip()
    if not room_number:
        raise Room.DoesNotExist

    queryset = Room.objects.filter(user=user)
    queryset = queryset.filter(r_number__iexact=room_number) if case_insensitive else queryset.filter(r_number=room_number)
    if department is not None:
        department_match = queryset.filter(department=department).first()
        if department_match is not None:
            return department_match

    match = queryset.first()
    if match is None:
        raise Room.DoesNotExist
    return match


def _teacher_uid_string_from_ids(id_values, user, fallback_teacher=None):
    ids = [teacher_id for teacher_id in (id_values or []) if teacher_id]
    teachers = Instructor.objects.filter(id__in=ids, user=user)
    teacher_by_id = {teacher.id: teacher for teacher in teachers}
    values = [teacher_by_id[teacher_id].uid for teacher_id in ids if teacher_id in teacher_by_id]
    if values:
        return ";".join(values)
    return getattr(fallback_teacher, "uid", "") or ""


def _teacher_name_string_from_ids(id_values, user, fallback_teacher=None):
    ids = [teacher_id for teacher_id in (id_values or []) if teacher_id]
    teachers = Instructor.objects.filter(id__in=ids, user=user)
    teacher_by_id = {teacher.id: teacher for teacher in teachers}
    values = [teacher_by_id[teacher_id].name for teacher_id in ids if teacher_id in teacher_by_id]
    if values:
        return "; ".join(values)
    return getattr(fallback_teacher, "name", "") or ""


@login_required
def delete_all_teacher_subject_mappings(request):
    return _delete_all_step_entries(
        request,
        SectionSubjectInstructor.objects.filter(user=request.user),
        "map_teacher_subjects",
        "section-subject-teacher mappings",
    )


@login_required
def delete_teacher_subject_mapping(request, subject_number, instructor_id):
    if request.method == "POST":
        try:
            subj = _resolve_subject_for_user(subject_number, request.user)
            instructor = Instructor.objects.get(id=instructor_id, user=request.user)
            subj.instructors.remove(instructor)
            messages.success(request, "Mapping removed successfully.")
        except Exception as e:
            messages.error(request, f"Error removing mapping: {e}")
    return redirect("map_teacher_subjects")


@login_required
def delete_sci_mapping(request, mapping_id):
    if request.method == "POST":
        SectionSubjectInstructor.objects.filter(
            id=mapping_id, user=request.user
        ).delete()
        reset_global_schedule_cache(request.user.id)
        messages.success(request, "Mapping removed.")
    return redirect("map_teacher_subjects")




@login_required
def inst_list_view(request):
    return render(
        request,
        'inslist.html',
        {'instructors': Instructor.objects.filter(user=request.user).select_related("department").order_by("uid")}
    )


@login_required
def dashboard_inst_list_view(request):
    return render(
        request,
        'dashboard_inslist.html',
        {'instructors': Instructor.objects.filter(user=request.user)}
    )


@login_required
def delete_instructor(request, pk):
    if request.method == 'POST':
        Instructor.objects.filter(pk=pk, user=request.user).delete()
        reset_global_schedule_cache(request.user.id)
        return redirect('editinstructor')


@login_required
def delete_all_instructors(request):
    return _delete_all_step_entries(
        request,
        Instructor.objects.filter(user=request.user),
        "addInstructors",
        "teachers",
    )




@login_required
def addRooms(request):
    edit_room = _get_required_step_edit_object(request, Room, user=request.user) if request.method == "POST" and request.POST.get("edit_id") else _get_step_edit_object(request, Room, user=request.user)
    form = RoomForm(request.POST or None, instance=edit_room, user=request.user)

    def resolve_department(dept_identifier):
        dept_identifier = (dept_identifier or "").strip()
        if not dept_identifier:
            raise Department.DoesNotExist
        if dept_identifier.isdigit():
            return Department.objects.get(pk=int(dept_identifier), user=request.user)
        try:
            return Department.objects.get(code=dept_identifier.upper(), user=request.user)
        except Department.DoesNotExist:
            return Department.objects.get(name=dept_identifier, user=request.user)

    # ---------------------------
    # 1) MANUAL ADD ROOM
    # ---------------------------
    if request.method == "POST" and "add_room" in request.POST:
        if form.is_valid():
            room = form.save(commit=False)
            room.user = request.user
            raw_room_type = (room.room_type or "").strip().lower()
            room.room_type = {
                "lecture hall": "Lecture Hall",
                "lab": "Lab",
                "seminar room": "Seminar Room",
            }.get(raw_room_type, room.room_type)
            room.lab_category = normalize_lab_categories_value(room.lab_category) if room.room_type == "Lab" else ""
            if room.room_type == "Lab" and not room.lab_category:
                messages.error(request, "Lab rooms must have a Lab Category.")
                return redirect("addRooms")
            room.save()

            reset_global_schedule_cache(request.user.id)
            messages.success(request, "Room updated successfully!" if edit_room else "Room added successfully!")
            return redirect("addRooms")
        else:
            messages.error(request, "Please fill out all required fields.")
            return redirect("addRooms")

    # ---------------------------
    # 2) CSV UPLOAD ROOMS
    # ---------------------------
    if request.method == "POST" and "csv_upload" in request.POST:
        csv_file = request.FILES.get("csv_file")

        if not csv_file or not csv_file.name.endswith(".csv"):
            messages.error(request, "Please upload a valid CSV file.")
            return redirect("addRooms")

        import csv
        decoded = csv_file.read().decode("utf-8").splitlines()
        reader = csv.reader(decoded)

        added = 0
        updated = 0
        first = True
        skipped = 0
        issues = []
        skipped = 0
        issues = []

        for row_number, row in enumerate(reader, start=1):
            if not row:
                continue

            if first:
                first = False
                header = [h.strip().lower() for h in row]

                rid_index = header.index("r_number") if "r_number" in header else 0
                dept_index = header.index("department") if "department" in header else (
                    header.index("department_code") if "department_code" in header else (
                        header.index("department_id") if "department_id" in header else 1
                    )
                )
                cap_index = header.index("seating_capacity") if "seating_capacity" in header else 2
                type_index = header.index("room_type") if "room_type" in header else 3
                if "lab_category" not in header:
                    messages.error(request, "Missing required column: lab_category")
                    return redirect("addRooms")
                category_index = header.index("lab_category")
                continue

            if len(row) < 5:
                skipped += 1
                _csv_issue(issues, row_number, "expected at least 5 columns: r_number, department, seating_capacity, room_type, lab_category")
                continue

            r_number = row[rid_index].strip()
            department_value = row[dept_index].strip()
            seating_capacity = row[cap_index].strip()
            room_type = row[type_index].strip()
            lab_category = normalize_lab_category(row[category_index].strip()) if category_index < len(row) else ""

            if not r_number:
                skipped += 1
                _csv_issue(issues, row_number, "r_number is blank")
                continue
            if not department_value:
                skipped += 1
                _csv_issue(issues, row_number, f"department is blank for room '{r_number}'")
                continue

            room_map = {
                "lecture hall": "Lecture Hall",
                "lab": "Lab",
                "seminar room": "Seminar Room",
            }
            room_type = room_map.get(room_type.lower(), room_type)
            if room_type != "Lab":
                lab_category = ""
            elif not lab_category:
                # Lab rooms must have a category in CSV.
                skipped += 1
                _csv_issue(issues, row_number, f"lab_category is blank for lab room '{r_number}'")
                continue

            if not seating_capacity.isdigit():
                skipped += 1
                _csv_issue(issues, row_number, f"seating_capacity '{seating_capacity}' is not a whole number")
                continue

            seating_capacity = int(seating_capacity)
            try:
                department = resolve_department(department_value)
            except Department.DoesNotExist:
                skipped += 1
                _csv_issue(issues, row_number, f"department '{department_value}' not found")
                continue

            if not Room.objects.filter(
                r_number=r_number,
                user=request.user,
                department=department,
            ).exists():
                Room.objects.create(
                    user=request.user,
                    r_number=r_number,
                    seating_capacity=seating_capacity,
                    room_type=room_type,
                    lab_category=lab_category,
                    department=department
                )
                added += 1
            else:
                skipped += 1
                _csv_issue(issues, row_number, f"room '{r_number}' already exists in department '{department.code}'")

        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"{added} room(s) added from CSV! {skipped} skipped.")
        _emit_csv_issues(request, issues)
        return redirect("addRooms")

    return render(request, "addRooms.html", {"form": form, "edit_room": edit_room})


@login_required
def room_list(request):
    rooms = Room.objects.filter(user=request.user).select_related("department").order_by("department__code", "r_number")
    return render(request, 'roomslist.html', {'rooms': rooms})


@login_required
def delete_room(request, pk):
    if request.method == 'POST':
        Room.objects.filter(pk=pk, user=request.user).delete()
        reset_global_schedule_cache(request.user.id)
        return redirect('editrooms')


@login_required
def delete_all_rooms(request):
    return _delete_all_step_entries(
        request,
        Room.objects.filter(user=request.user),
        "addRooms",
        "rooms",
    )


@login_required
def addTimings(request):
    edit_meeting_time = _get_required_step_edit_object(request, MeetingTime, user=request.user) if request.method == "POST" and request.POST.get("edit_id") else _get_step_edit_object(request, MeetingTime, user=request.user)
    form = MeetingTimeForm(request.POST or None, instance=edit_meeting_time)

    # -------------------------
    # 1. MANUAL ADD TIMING
    # -------------------------
    if request.method == "POST" and "add_timing" in request.POST:
        if form.is_valid():
            mt = form.save(commit=False)
            mt.user = request.user
            mt.save()
            reset_global_schedule_cache(request.user.id)
            messages.success(request, "Timing updated successfully!" if edit_meeting_time else "Timing added successfully!")
            return redirect("addTimings")
        else:
            messages.error(request, "Please fill all fields.")
            return redirect("addTimings")

    # -------------------------
    # 2. CSV UPLOAD TIMINGS
    # -------------------------
    if request.method == "POST" and "csv_upload" in request.POST:

        csv_file = request.FILES.get("csv_file")

        if not csv_file:
            messages.error(request, "Please select a CSV file.")
            return redirect("addTimings")

        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Invalid file type. Upload CSV only.")
            return redirect("addTimings")

        import csv
        decoded = csv_file.read().decode("utf-8").splitlines()
        reader = csv.reader(decoded)

        added = 0
        updated = 0
        first = True
        skipped = 0
        issues = []

        for row_number, row in enumerate(reader, start=1):

            # Skip empty or incomplete rows
            if not row:
                continue

            # Skip header row
            if first:
                first = False
                if row[0].strip().lower() == "pid":
                    continue

            if len(row) < 3:
                skipped += 1
                _csv_issue(issues, row_number, "expected pid, time, and day columns")
                continue

            pid = row[0].strip()
            time = row[1].strip()
            day_raw = row[2].strip()

            if not pid:
                skipped += 1
                _csv_issue(issues, row_number, "pid is blank")
                continue
            if not time:
                skipped += 1
                _csv_issue(issues, row_number, f"time is blank for pid '{pid}'")
                continue
            if not day_raw:
                skipped += 1
                _csv_issue(issues, row_number, f"day is blank for pid '{pid}'")
                continue

            # --------------------------
            # FIXED DAY PARSING
            # --------------------------
            # Normalize capitalization
            day_value = day_raw.capitalize()

            # Map numeric day → weekday
            day_map = {
                '1': 'Monday',
                '2': 'Tuesday',
                '3': 'Wednesday',
                '4': 'Thursday',
                '5': 'Friday',
                '6': 'Saturday',
                '7': 'Sunday',
            }

            # If the value is numeric, map it
            day = day_map.get(day_value.lower(), day_value.capitalize())

            # Validate day
            valid_days = [d[0] for d in DAYS_OF_WEEK]  # ['Monday', 'Tuesday',...]
            if day not in valid_days:
                skipped += 1
                _csv_issue(issues, row_number, f"day '{day_raw}' is invalid after normalization")
                continue

            # Validate time slot
            valid_times = [t[0] for t in TIME_SLOTS]  # ['1','2','3','4',...]
            if time not in valid_times:
                skipped += 1
                _csv_issue(issues, row_number, f"time '{time}' is not a valid slot")
                continue

            # Avoid duplicates
            if not MeetingTime.objects.filter(pid=pid, user=request.user).exists():
                MeetingTime.objects.create(user=request.user, pid=pid, day=day, time=time)
                added += 1
            else:
                skipped += 1
                _csv_issue(issues, row_number, f"pid '{pid}' already exists")

        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"{added} timing(s) added from CSV! {skipped} skipped.")
        _emit_csv_issues(request, issues)
        return redirect("addTimings")



    return render(request, "addTimings.html", {"form": form, "edit_meeting_time": edit_meeting_time})


@login_required
def meeting_list_view(request):
    meeting_times = MeetingTime.objects.filter(user=request.user).order_by("day", "time", "pid")
    return render(request, 'mtlist.html', {'meeting_times': meeting_times})


@login_required
def delete_meeting_time(request, pk):
    if request.method == 'POST':
        MeetingTime.objects.filter(pk=pk, user=request.user).delete()
        reset_global_schedule_cache(request.user.id)
        return redirect('editmeetingtime')


@login_required
def delete_all_meeting_times(request):
    return _delete_all_step_entries(
        request,
        MeetingTime.objects.filter(user=request.user),
        "addTimings",
        "timings",
    )


@login_required
def addDepts(request):
    edit_department = _get_required_step_edit_object(request, Department, user=request.user) if request.method == "POST" and request.POST.get("edit_id") else _get_step_edit_object(request, Department, user=request.user)
    form = DepartmentForm(request.POST or None, instance=edit_department)

    # ------------------------------------
    # 1) MANUAL ADD DEPARTMENT
    # ------------------------------------
    if request.method == "POST" and "add_department" in request.POST:
        if form.is_valid():
            dept = form.save(commit=False)
            dept.user = request.user
            dept.save()
            reset_global_schedule_cache(request.user.id)
            messages.success(request, "Department updated successfully!" if edit_department else "Department added successfully!")
            return redirect("addDepts")
        else:
            messages.error(request, "Please fill all required fields.")
            return redirect("addDepts")

    # ------------------------------------
    # 2) CSV UPLOAD DEPARTMENTS
    # ------------------------------------
    if request.method == "POST" and "csv_upload" in request.POST:

        csv_file = request.FILES.get("csv_file")

        if not csv_file:
            messages.error(request, "Please select a CSV file.")
            return redirect("addDepts")

        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Invalid file type. Upload CSV only.")
            return redirect("addDepts")

        import csv
        decoded = csv_file.read().decode("utf-8").splitlines()
        reader = csv.reader(decoded)

        added = 0
        updated = 0
        first = True
        skipped = 0
        issues = []

        for row_number, row in enumerate(reader, start=1):

            # Skip blank rows
            if not row:
                continue

            # Handle header row
            if first:
                first = False
                header = [h.strip().lower() for h in row]
                
                # Detect column positions
                try:
                    # Try both 'name' and 'dept_name'
                    if 'name' in header:
                        name_i = header.index('name')
                    elif 'dept_name' in header:
                        name_i = header.index('dept_name')
                    else:
                        name_i = 0  # Default first column
                    
                    # Try both 'code' and 'dept_code'
                    if 'code' in header:
                        code_i = header.index('code')
                    elif 'dept_code' in header:
                        code_i = header.index('dept_code')
                    else:
                        code_i = 1  # Default second column
                        
                except ValueError as e:
                    messages.error(request, f"CSV format error: {e}")
                    return redirect("addDepts")
                
                continue

            if len(row) < 2:
                skipped += 1
                _csv_issue(issues, row_number, "expected name and code columns")
                continue

            # Extract values
            dept_name = row[name_i].strip()
            dept_code = row[code_i].strip().upper()

            if not dept_name:
                skipped += 1
                _csv_issue(issues, row_number, "department name is blank")
                continue
            if not dept_code:
                skipped += 1
                _csv_issue(issues, row_number, f"department code is blank for '{dept_name}'")
                continue

            # Skip if department already exists (check both name and code)
            if Department.objects.filter(code=dept_code, user=request.user).exists():
                skipped += 1
                _csv_issue(issues, row_number, f"department code '{dept_code}' already exists")
                continue

            if Department.objects.filter(name=dept_name, user=request.user).exists():
                skipped += 1
                _csv_issue(issues, row_number, f"department name '{dept_name}' already exists")
                continue

            # Create Department
            try:
                Department.objects.create(
                    user=request.user,
                    name=dept_name,
                    code=dept_code
                )
                added += 1
            except Exception as e:
                skipped += 1
                _csv_issue(issues, row_number, f"could not create department '{dept_name}': {e}")
                continue

        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"{added} department(s) added from CSV! {skipped} skipped.")
        _emit_csv_issues(request, issues)
        return redirect("addDepts")

    return render(request, 'addDepts.html', {'form': form, 'edit_department': edit_department})


@login_required
def department_list(request):
    departments = Department.objects.filter(user=request.user).order_by("code", "name")
    return render(request, 'deptlist.html', {'departments': departments})


@login_required
def dashboard_department_list(request):
    return render(request, 'dashboard_deptlist.html', {'departments': Department.objects.filter(user=request.user)})


@login_required
def delete_department(request, pk):
    if request.method == 'POST':
        Department.objects.filter(pk=pk, user=request.user).delete()
        reset_global_schedule_cache(request.user.id)
        return redirect('editdepartment')


@login_required
def delete_all_departments(request):
    return _delete_all_step_entries(
        request,
        Department.objects.filter(user=request.user),
        "addDepts",
        "departments",
    )


@login_required
def addSections(request):
    edit_section = _get_required_step_edit_object(request, Section, user=request.user) if request.method == "POST" and request.POST.get("edit_id") else _get_step_edit_object(request, Section, user=request.user)
    form = SectionForm(request.POST or None, instance=edit_section, user=request.user)

    # -------------------------------------------
    # 1) MANUAL ADDING OF SECTION
    # -------------------------------------------
    if request.method == "POST" and "add_section" in request.POST:
        if form.is_valid():
            section = form.save(commit=False)
            section.user = request.user
            section.save()
            template_section = None if edit_section else clone_section_subjects_from_similar(section)
            reset_global_schedule_cache(request.user.id)
            if edit_section:
                messages.success(request, "Section updated successfully!")
            elif template_section:
                messages.success(
                    request,
                    f"Section added successfully! Subjects copied from {template_section.section_id}.",
                )
            else:
                messages.success(request, "Section added successfully!")
            return redirect("addSections")
        else:
            messages.error(request, "Please fill all required fields.")
            return redirect("addSections")

    # -------------------------------------------
    # 2) CSV UPLOAD SECTION DATA
    # -------------------------------------------
    if request.method == "POST" and "csv_upload" in request.POST:

        csv_file = request.FILES.get("csv_file")

        if not csv_file:
            messages.error(request, "Please select a CSV file.")
            return redirect("addSections")

        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Invalid file type. Upload CSV only.")
            return redirect("addSections")

        import csv
        decoded = csv_file.read().decode("utf-8").splitlines()
        reader = csv.reader(decoded)

        added = 0
        updated = 0
        first = True
        skipped = 0
        issues = []

        def resolve_department(dept_identifier):
            dept_identifier = (dept_identifier or "").strip()
            if not dept_identifier:
                raise Department.DoesNotExist
            if dept_identifier.isdigit():
                return Department.objects.get(pk=int(dept_identifier), user=request.user)
            try:
                return Department.objects.get(code=dept_identifier.upper(), user=request.user)
            except Department.DoesNotExist:
                return Department.objects.get(name=dept_identifier, user=request.user)

        def row_value(row, index):
            if index is None or index >= len(row):
                return ""
            return row[index].strip()

        for row_number, row in enumerate(reader, start=1):

            # Skip empty or insufficient rows
            if not row:
                continue

            # Auto-detect column positions using header row
            if first:
                first = False
                header = [h.strip().lower() for h in row]

                def index(name, default=None):
                    try:
                        return header.index(name)
                    except:
                        return default

                section_i = index("section_id", 0)
                program_i = index("program_name", index("program", 1))
                dept_i = index("department", index("department_code", index("department_id", 1)))
                if dept_i == 1 and program_i == 1:
                    dept_i = index("department", index("department_code", index("department_id", 2)))
                strength_i = index("student_strength", index("section_strength", 2 if program_i is None else 3))

                # Skip header
                continue

            if len(row) < 2:
                skipped += 1
                _csv_issue(issues, row_number, "expected at least section_id and department columns")
                continue

            # Extract values
            section_id = row_value(row, section_i)
            program_name = row_value(row, program_i)
            dept_name = row_value(row, dept_i)
            student_strength = row_value(row, strength_i) or "70"

            if not section_id:
                skipped += 1
                _csv_issue(issues, row_number, "section_id is blank")
                continue
            if not dept_name:
                skipped += 1
                _csv_issue(issues, row_number, f"department is blank for section '{section_id}'")
                continue

            # Validate Department
            try:
                dept = resolve_department(dept_name)
            except Department.DoesNotExist:
                skipped += 1
                _csv_issue(issues, row_number, f"department '{dept_name}' not found")
                continue

            if not student_strength.isdigit():
                skipped += 1
                _csv_issue(issues, row_number, f"student_strength '{student_strength}' is not a whole number")
                continue

            existing_section = Section.objects.filter(section_id=section_id, user=request.user).first()
            if existing_section:
                existing_section.program_name = program_name
                existing_section.department = dept
                existing_section.student_strength = int(student_strength)
                existing_section.save(update_fields=["program_name", "department", "student_strength"])
                updated += 1
                continue

            # Create the Section object
            section = Section.objects.create(
                user=request.user,
                section_id=section_id,
                program_name=program_name,
                student_strength=int(student_strength),
                department=dept,
            )
            clone_section_subjects_from_similar(section)

            added += 1

        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"{added} section(s) added and {updated} section(s) updated from CSV! {skipped} skipped.")
        _emit_csv_issues(request, issues)
        return redirect("addSections")

    return render(request, "addSections.html", {"form": form, "edit_section": edit_section})


@login_required
def prefilled_timetable_setup(request):
    if request.method == "POST":
        csv_file = request.FILES.get("csv_file")
        if not csv_file:
            messages.error(request, "Please select a sections CSV file.")
            return redirect("prefilled_timetable_setup")
        if not csv_file.name.endswith(".csv"):
            messages.error(request, "Invalid file type. Upload CSV only.")
            return redirect("prefilled_timetable_setup")

        decoded = csv_file.read().decode("utf-8-sig").splitlines()
        reader = csv.reader(decoded)
        section_ids = []
        issues = []
        first = True

        def resolve_department(dept_identifier):
            dept_identifier = (dept_identifier or "").strip()
            if not dept_identifier:
                raise Department.DoesNotExist
            if dept_identifier.isdigit():
                return Department.objects.get(pk=int(dept_identifier), user=request.user)
            try:
                return Department.objects.get(code=dept_identifier.upper(), user=request.user)
            except Department.DoesNotExist:
                return Department.objects.get(name=dept_identifier, user=request.user)

        def row_value(row, index):
            if index is None or index >= len(row):
                return ""
            return row[index].strip()

        section_i = 0
        program_i = 1
        dept_i = 2
        strength_i = 3

        for row_number, row in enumerate(reader, start=1):
            if not row:
                continue
            if first:
                first = False
                header = [h.strip().lower() for h in row]

                def index(name, default=None):
                    try:
                        return header.index(name)
                    except ValueError:
                        return default

                section_i = index("section_id", 0)
                program_i = index("program_name", index("program", 1))
                dept_i = index("department", index("department_code", index("department_id", 1)))
                if dept_i == 1 and program_i == 1:
                    dept_i = index("department", index("department_code", index("department_id", 2)))
                strength_i = index("student_strength", index("section_strength", 2 if program_i is None else 3))
                continue

            section_id = row_value(row, section_i)
            program_name = row_value(row, program_i)
            dept_name = row_value(row, dept_i)
            student_strength = row_value(row, strength_i) or "70"

            if not section_id:
                _csv_issue(issues, row_number, "section_id is blank")
                continue
            if not dept_name:
                _csv_issue(issues, row_number, f"department is blank for section '{section_id}'")
                continue
            if not student_strength.isdigit():
                _csv_issue(issues, row_number, f"student_strength '{student_strength}' is not a whole number")
                continue

            try:
                dept = resolve_department(dept_name)
            except Department.DoesNotExist:
                _csv_issue(issues, row_number, f"department '{dept_name}' not found")
                continue

            section, created = Section.objects.update_or_create(
                user=request.user,
                section_id=section_id,
                defaults={
                    "program_name": program_name,
                    "department": dept,
                    "student_strength": int(student_strength),
                },
            )
            if created:
                clone_section_subjects_from_similar(section)
            section_ids.append(section.section_id)

        if not section_ids:
            messages.error(request, "No valid sections found in the CSV.")
            _emit_csv_issues(request, issues)
            return redirect("prefilled_timetable_setup")

        state = _get_user_state(request.user.id)
        state["classes"] = []
        state["labs"] = []
        state["schedules"] = [{"classes": [], "labs": [], "stats": {}, "reco_block": {}}]
        state["prefill_mode"] = True
        state["prefill_section_ids"] = section_ids
        state["prefill_locked_classes"] = []
        state["prefill_locked_labs"] = []
        _reset_generated_drag_state(state, current_index=1)
        request.session["current_index"] = 1
        request.session["prefill_mode"] = True
        request.session["prefill_section_ids"] = section_ids
        request.session.pop("prefill_saved_slots", None)
        request.session.pop("active_saved_prefill_id", None)
        request.session.modified = True
        reset_global_schedule_cache(request.user.id)
        messages.success(request, f"Empty prefilled timetable created for {len(section_ids)} section(s).")
        _emit_csv_issues(request, issues)
        return redirect("prefilled_timetable_view")

    return render(request, "prefilled_timetable_setup.html")


@login_required
def prefilled_timetable_view(request):
    state = _get_user_state(request.user.id)
    section_ids = list(state.get("prefill_section_ids") or request.session.get("prefill_section_ids") or [])
    if not (state.get("prefill_mode") or request.session.get("prefill_mode")) or not section_ids:
        messages.info(request, "Upload a sections CSV to create an empty prefilled timetable.")
        return redirect("prefilled_timetable_setup")

    try:
        state["prefill_mode"] = True
        state["prefill_section_ids"] = section_ids
        if state.get("classes") is None:
            state["classes"] = []
        if state.get("labs") is None:
            state["labs"] = []
        if not state.get("schedules"):
            state["schedules"] = [{"classes": [], "labs": [], "stats": {}, "reco_block": {}}]
        restore_prefill_session_snapshot(request, state, section_ids)
        ensure_manual_prefill_slot_uids(state)
        state["view_mode"] = "editing"
        request.session["current_index"] = 1

        global GLOBAL_CLASSES, GLOBAL_LABS, GLOBAL_GENERATED_SCHEDULES, CURRENT_VIEW_MODE
        GLOBAL_CLASSES = state["classes"]
        GLOBAL_LABS = state["labs"]
        GLOBAL_GENERATED_SCHEDULES = state["schedules"]
        CURRENT_VIEW_MODE = "editing"

        tables = build_section_tables(GLOBAL_CLASSES, GLOBAL_LABS, user=request.user)
        allowed_section_ids = set(section_ids)
        tables = [table for table in tables if table["section"].section_id in allowed_section_ids]
        tables = _decorate_generated_tables_with_parking(tables, state)

        departments = []
        seen_depts = set()
        for table in tables:
            section = table["section"]
            dept = section.department
            table["section_department_code"] = dept.code
            table["section_department_name"] = dept.name
            table["subject_counts"] = []
            table["missed_labs"] = []
            table["total_missing_classes"] = 0
            if dept.code not in seen_depts:
                seen_depts.add(dept.code)
                departments.append({"code": dept.code, "name": dept.name})
            section_id = section.section_id
            table["create_slot_subjects"] = []
            table["create_slot_teachers"] = []
            table["create_slot_rooms"] = []
            table["create_slot_subjects_id"] = f"create-slot-subjects-{section_id}"
            table["create_slot_teachers_id"] = f"create-slot-teachers-{section_id}"
            table["create_slot_rooms_id"] = f"create-slot-rooms-{section_id}"

        return render(request, "gentimetable.html", {
            "tables": tables,
            "teacher_tables": [],
            "room_tables": [],
            "SLOT_LABELS": SLOT_LABELS,
            "index": 1,
            "teacher_workloads": {},
            "demo_mode": False,
            "active_dept": "",
            "active_program": "all",
            "program_options": [],
            "departments": departments,
            "prefill_mode": True,
            "active_saved_prefill": _get_active_saved_prefill(request),
            "can_edit_delete": True,
            "can_substitute": True,
            "can_drag_drop": True,
        })
    except Exception:
        logger.exception("Failed to open prefilled timetable for user %s", request.user.id)
        state["classes"] = []
        state["labs"] = []
        state["generated_parking_items"] = []
        request.session.pop("prefill_saved_slots", None)
        request.session.pop("current_index", None)
        request.session.modified = True
        messages.error(request, "Could not open the prefilled timetable. Please upload the sections CSV again.")
        return redirect("prefilled_timetable_setup")


@login_required
def save_prefilled_timetable(request):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    state = _get_user_state(request.user.id)
    snapshot = save_prefill_session_snapshot(request, state) or request.session.get("prefill_saved_slots")
    if not snapshot or not snapshot.get("section_ids"):
        messages.error(request, "Create a prefilled table before saving.")
        return redirect("prefilled_timetable_setup")

    total_slots = len(snapshot.get("classes") or []) + len(snapshot.get("labs") or []) + len(snapshot.get("parking_items") or [])
    name = (request.POST.get("name") or "").strip()
    active_prefill = _get_active_saved_prefill(request)
    explicit_name = bool(name)
    if not name and active_prefill is None:
        section_preview = ", ".join(list(snapshot.get("section_ids") or [])[:2])
        if len(snapshot.get("section_ids") or []) > 2:
            section_preview += " + more"
        name = f"Prefill {section_preview}" if section_preview else "Saved Prefill"

    if active_prefill is not None:
        update_fields = ["snapshot", "updated_at"]
        if explicit_name:
            active_prefill.name = name[:120]
            update_fields.append("name")
        active_prefill.snapshot = snapshot
        active_prefill.save(update_fields=update_fields)
        messages.success(request, f"Saved prefill updated with {total_slots} slot(s).")
    else:
        active_prefill = SavedPrefill.objects.create(user=request.user, name=name[:120], snapshot=snapshot)
        _set_active_saved_prefill(request, active_prefill)
        messages.success(request, f"Prefilled table saved with {total_slots} slot(s).")
    return redirect("saved_prefill_list")


@login_required
def export_prefill_csv(request):
    state = _get_user_state(request.user.id)
    snapshot = save_prefill_session_snapshot(request, state) or request.session.get("prefill_saved_slots") or {}
    section_ids = list(snapshot.get("section_ids") or [])
    if not section_ids:
        messages.error(request, "Create a prefilled timetable before downloading CSV.")
        return redirect("prefilled_timetable_view")

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="prefill_slots.csv"'
    writer = csv.DictWriter(response, fieldnames=_PREFILL_CSV_COLUMNS)
    writer.writeheader()

    for section_id in section_ids:
        writer.writerow({"record_type": "section", "section_id": section_id})
    for entry in list(snapshot.get("classes") or []):
        writer.writerow(_prefill_csv_row_from_entry(entry, parking=False))
    for entry in list(snapshot.get("labs") or []):
        writer.writerow(_prefill_csv_row_from_entry(entry, parking=False))
    for entry in list(snapshot.get("parking_items") or []):
        writer.writerow(_prefill_csv_row_from_entry(entry, parking=True))
    return response


@login_required
def import_prefill_csv(request):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    csv_file = request.FILES.get("csv_file")
    if not csv_file or not str(getattr(csv_file, "name", "")).lower().endswith(".csv"):
        messages.error(request, "Please upload a valid prefill CSV file.")
        return redirect("prefilled_timetable_view")

    try:
        decoded_lines = csv_file.read().decode("utf-8-sig").splitlines()
        snapshot = _build_prefill_csv_snapshot(decoded_lines)
    except UnicodeDecodeError:
        messages.error(request, "CSV file must be UTF-8 encoded.")
        return redirect("prefilled_timetable_view")
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("prefilled_timetable_view")

    if not snapshot.get("section_ids"):
        messages.error(request, "Prefill CSV has no selected sections.")
        return redirect("prefilled_timetable_view")

    request.session["prefill_saved_slots"] = snapshot
    request.session["prefill_mode"] = True
    request.session["prefill_section_ids"] = list(snapshot.get("section_ids") or [])
    request.session.modified = True
    _clear_active_saved_prefill(request)
    if not _activate_prefill_snapshot(request, snapshot):
        messages.error(request, "Could not restore any slots from the uploaded CSV.")
    else:
        total_slots = len(snapshot.get("classes") or []) + len(snapshot.get("labs") or []) + len(snapshot.get("parking_items") or [])
        messages.success(request, f"Prefill CSV loaded with {total_slots} slot(s).")
    return redirect("prefilled_timetable_view")


@login_required
def saved_prefill_list(request):
    prefills = []
    for prefill in SavedPrefill.objects.filter(user=request.user):
        snapshot = prefill.snapshot or {}
        section_ids = list(snapshot.get("section_ids") or [])
        slot_count = len(snapshot.get("classes") or []) + len(snapshot.get("labs") or []) + len(snapshot.get("parking_items") or [])
        prefills.append({
            "item": prefill,
            "section_count": len(section_ids),
            "sections": ", ".join(section_ids[:4]) + (" + more" if len(section_ids) > 4 else ""),
            "slot_count": slot_count,
        })
    return render(request, "saved_prefill_list.html", {"prefills": prefills})


@login_required
def rename_saved_prefill(request, prefill_id):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")
    prefill = SavedPrefill.objects.filter(pk=prefill_id, user=request.user).first()
    if prefill is None:
        raise Http404("Saved prefill not found")
    name = (request.POST.get("name") or "").strip()
    if not name:
        messages.error(request, "Enter a name before renaming the saved prefill.")
        return redirect("saved_prefill_list")
    prefill.name = name[:120]
    prefill.save(update_fields=["name", "updated_at"])
    messages.success(request, "Saved prefill renamed.")
    return redirect("saved_prefill_list")


@login_required
def open_saved_prefill(request, prefill_id):
    try:
        prefill = SavedPrefill.objects.get(pk=prefill_id, user=request.user)
    except SavedPrefill.DoesNotExist:
        raise Http404("Saved prefill not found")
    if not _activate_prefill_snapshot(request, prefill.snapshot or {}):
        messages.error(request, "This saved prefill has no section data.")
        return redirect("saved_prefill_list")
    _set_active_saved_prefill(request, prefill)
    messages.success(request, "Saved prefill loaded for editing.")
    return redirect("prefilled_timetable_view")


@login_required
def generate_saved_prefill(request, prefill_id):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")
    try:
        prefill = SavedPrefill.objects.get(pk=prefill_id, user=request.user)
    except SavedPrefill.DoesNotExist:
        raise Http404("Saved prefill not found")
    if not _activate_prefill_snapshot(request, prefill.snapshot or {}):
        messages.error(request, "This saved prefill has no section data.")
        return redirect("saved_prefill_list")
    _set_active_saved_prefill(request, prefill)
    return redirect(f"{reverse('generate_timetable_loading')}?use_pso=1&use_prefilled_slots=1")


@login_required
def delete_saved_prefill(request, prefill_id):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")
    SavedPrefill.objects.filter(pk=prefill_id, user=request.user).delete()
    if str(request.session.get("active_saved_prefill_id") or "") == str(prefill_id):
        _clear_active_saved_prefill(request)
    messages.success(request, "Saved prefill deleted.")
    return redirect("saved_prefill_list")


@login_required
def section_list(request):
    sections = Section.objects.filter(user=request.user).select_related("department").order_by("department__code", "section_id")
    return render(request, 'seclist.html', {'sections': sections})


@login_required
def dashboard_section_list(request):
    return render(request, 'dashboard_seclist.html', {'sections': Section.objects.filter(user=request.user)})


@login_required
def delete_section(request, pk):
    if request.method == 'POST':
        Section.objects.filter(pk=pk, user=request.user).delete()
        reset_global_schedule_cache(request.user.id)
        return redirect('editsection')


@login_required
def delete_all_sections(request):
    return _delete_all_step_entries(
        request,
        Section.objects.filter(user=request.user),
        "addSections",
        "sections",
    )


@login_required
def generate(request):
    return render(request, 'generate.html')


def generate(request):
    return render(request, 'generate.html')




def delete_saved_timetable(request, tid):
    SavedTimetable.objects.filter(id=tid, user=request.user).delete()
    messages.success(request, "Saved timetable deleted.")
    return redirect('saved_timetable_list')


@login_required
def rename_saved_timetable(request, tid):
    if request.method != "POST":
        return HttpResponseForbidden("Invalid request method.")

    saved_t = SavedTimetable.objects.filter(id=tid, user=request.user).first()
    if saved_t is None:
        raise Http404("Saved timetable not found")

    name = (request.POST.get("name") or "").strip()
    if not name:
        messages.error(request, "Enter a name before renaming the saved timetable.")
        return redirect("saved_timetable_list")

    saved_t.name = name[:120]
    saved_t.save(update_fields=["name"])
    messages.success(request, "Saved timetable renamed.")
    return redirect("saved_timetable_list")

def expand_labs_for_pdf(rows):
    new_rows = []

    for row in rows:
        expanded = []
        for cell in row["cells"]:
            if cell.get("type") == "lab":
                span = cell.get("colspan", 1)
                for _ in range(span):
                    expanded.append(cell)
            else:
                expanded.append(cell)

        row_copy = row.copy()
        row_copy["cells"] = expanded
        new_rows.append(row_copy)

    return new_rows




from django.template.loader import render_to_string
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.http import HttpResponse, Http404, HttpResponseForbidden

@login_required
def saved_timetable_download_center(request, tid):
    saved_t = _get_saved_timetable_or_404(tid, request.user)
    return render(request, "download_center.html", {
        "saved_mode": True,
        "saved_id": saved_t.id,
        "departments": _get_department_filter_options(request.user),
        "college_name": COLLEGE_NAME,
    })


@login_required
def download_saved_timetable_pdf(request, tid, view_type='section'):
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse(
            "PDF generation dependencies are not installed on this machine yet.",
            status=503,
        )

    saved_t = _get_saved_timetable_or_404(tid, request.user)
    view_type = (view_type or 'section').lower()
    if view_type not in {"section", "room", "teacher", "workload"}:
        view_type = "section"

    classes, labs, _selected_program, selected_department = _saved_filtered_entities_for_request(request, saved_t)
    dept_filter = _parse_dept_filter(request)
    if not dept_filter and selected_department and selected_department != "all":
        dept_filter = {selected_department}
    blocks = []
    is_workload = False
    workload_rows = []
    view_label = "Section"

    if view_type == "section":
        view_label = "Section"
        tables = build_section_tables(classes, labs, user=request.user)
        tables = _decorate_section_tables_with_department_meta(tables)
        if dept_filter:
            tables = [
                table for table in tables
                if (getattr(getattr(table.get("section"), "department", None), "code", "") or "") in dept_filter
            ]
        blocks = [_section_pdf_block(t) for t in tables]
    elif view_type == "room":
        view_label = "Room"
        room_tables = build_room_tables(classes, labs, user=request.user)
        if dept_filter:
            room_tables = [
                table for table in room_tables
                if any(code in dept_filter for code in table.get("dept_codes", []))
            ]
        blocks = [_room_pdf_block(t) for t in room_tables]
    elif view_type == "teacher":
        view_label = "Teacher"
        fallback_map = _teacher_home_dept_fallback(classes, labs)
        teacher_tables = build_teacher_tables(classes, labs, user=request.user)
        if dept_filter:
            teacher_tables = [
                table for table in teacher_tables
                if (_teacher_home_dept_code(table.get("teacher"), fallback_map) or "") in dept_filter
            ]
        blocks = [_teacher_pdf_block(t) for t in teacher_tables]
    else:
        view_label = "Workload"
        is_workload = True
        fallback_map = _teacher_home_dept_fallback(classes, labs)
        workloads = _compute_teacher_workloads(classes, labs)
        if dept_filter:
            workloads = {
                teacher: data for teacher, data in workloads.items()
                if (_teacher_home_dept_code(teacher, fallback_map) or "") in dept_filter
            }
        for teacher, data in workloads.items():
            workload_rows.append({
                "name": teacher.name,
                "uid": getattr(teacher, "uid", ""),
                "departments": data.get("departments", "-"),
                "lectures": data.get("lectures", 0),
                "labs": data.get("labs", 0),
                "shared_labs": data.get("shared_labs", 0),
                "total": data.get("total", 0),
            })
        workload_rows.sort(key=lambda row: str(row["name"]).lower())

    html = render_to_string("generated_timetable_pdf.html", {
        "blocks": blocks,
        "is_workload": is_workload,
        "workload_rows": workload_rows,
        "view_label": view_label,
        "slot_labels": SLOT_LABELS,
        "college_name": COLLEGE_NAME,
        "college_logo": settings.STATIC_URL + "img/college_logo.png",
        "brand_logo": settings.STATIC_URL + "img/logo_email.png",
    })

    response = HttpResponse(content_type='application/pdf')
    inline = request.GET.get("inline") == "1"
    disposition = "inline" if inline else "attachment"
    response['Content-Disposition'] = f'{disposition}; filename="saved_{view_type}_timetable_{tid}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=_pdf_link_callback)
    if pisa_status.err:
        return HttpResponse("PDF creation failed.", status=500)
    return response


def _safe_get(obj, *attrs, default=""):
    for attr in attrs:
        value = getattr(obj, attr, None)
        if value not in (None, ""):
            return value
    return default


def _pick_slot_cell(table_rows, day, slot_number):
    for table_row in table_rows:
        if table_row.get("day") != day:
            continue
        for cell_data in table_row.get("cells", []):
            if str(cell_data.get("slot_number")) == str(slot_number):
                return cell_data
    return None


def _build_timetable_excel_response(classes, labs, user, filename, view_type='section', dept_filter=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return HttpResponse("Excel dependency missing", status=503)

    allowed_views = {"section", "room", "teacher", "workload", "all"}
    if view_type not in allowed_views:
        view_type = "section"

    classes = list(classes or [])
    labs = list(labs or [])

    section_tables = build_section_tables(classes, labs, user=user)
    room_tables = build_room_tables(classes, labs, user=user)
    teacher_tables = build_teacher_tables(classes, labs, user=user)
    teacher_workloads = _compute_teacher_workloads(classes, labs)

    if dept_filter:
        fb = _teacher_home_dept_fallback(classes, labs)
        section_tables = [
            t for t in section_tables
            if (getattr(getattr(t.get("section"), "department", None), "code", "") or "") in dept_filter
        ]
        room_tables = [
            rt for rt in room_tables
            if any(code in dept_filter for code in rt.get("dept_codes", []))
        ]
        teacher_tables = [
            t for t in teacher_tables
            if (_teacher_home_dept_code(t.get("teacher"), fb) or "") in dept_filter
        ]
        teacher_workloads = {
            teacher: data for teacher, data in teacher_workloads.items()
            if (_teacher_home_dept_code(teacher, fb) or "") in dept_filter
        }

    wb = Workbook()
    wb.remove(wb.active)

    # ---- STYLES ----
    day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    day_font = Font(bold=True, size=10)

    lunch_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    lunch_font = Font(bold=True, size=9)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # =====================================================
    # COMMON WRITER (DAYS → ROWS, SLOTS → COLUMNS)
    # =====================================================
    def write_timetable(ws, tables, title_func):
        row = 1

        def _format_class_entry(cls):
            subj = getattr(cls, "subject", None)
            instructor = getattr(cls, "instructor", None)
            room = getattr(cls, "room", None)
            lines = [
                str(_safe_get(subj, "subject_number", "subject_name", default="Class")),
                str(_safe_get(instructor, "name", "uid", default="TBD")),
                str(_safe_get(room, "r_number", default="Room TBD")),
            ]
            return "\n".join(lines)

        def _format_lab_entry(lab):
            subj = getattr(lab, "subject", None)
            instructor = getattr(lab, "instructor", None)
            room = getattr(lab, "room", None)
            lines = [
                f"{_safe_get(subj, 'subject_number', 'subject_name', default='Lab')} (Lab)",
                str(_safe_get(instructor, "name", "uid", default="TBD")),
                str(_safe_get(room, "r_number", default="Room TBD")),
            ]
            return "\n".join(lines)

        for table in tables:
            # ---- TITLE ----
            ws[f"A{row}"] = title_func(table)
            ws[f"A{row}"].font = Font(bold=True, size=12, color="1F4E78")
            row += 1

            # ---- HEADER ----
            ws[f"A{row}"] = "Day"

            for slot in range(1, 10):
                cell = ws.cell(row=row, column=slot + 1,
                               value=SLOT_LABELS.get(str(slot), f"Slot {slot}"))
                cell.fill = day_fill
                cell.font = day_font
                cell.alignment = center_align

            row += 1

            # ---- DATA ----
            for day in DAYS:
                ws[f"A{row}"] = day
                ws[f"A{row}"].fill = day_fill
                ws[f"A{row}"].font = day_font

                for slot in range(1, 10):
                    xl_cell = ws.cell(row=row, column=slot + 1)
                    xl_cell.border = border
                    xl_cell.alignment = center_align

                    cell_data = _pick_slot_cell(table.get("rows", []), day, slot)

                    if not cell_data:
                        continue

                    cell_type = cell_data.get("type")

                    if cell_type == "lunch":
                        xl_cell.value = "LUNCH"
                        xl_cell.fill = lunch_fill
                        xl_cell.font = lunch_font

                    elif cell_type == "class":
                        class_items = cell_data.get("classes", [])
                        if class_items:
                            xl_cell.value = "\n\n".join(
                                _format_class_entry(cls) for cls in class_items
                            )

                    elif cell_type == "lab":
                        lab_items = cell_data.get("labs", [])
                        if lab_items:
                            xl_cell.value = "\n\n".join(
                                _format_lab_entry(lab) for lab in lab_items
                            )

                row += 1

            row += 2

        # ---- WIDTH ----
        ws.column_dimensions["A"].width = 18
        for col in range(2, 11):
            ws.column_dimensions[chr(64 + col)].width = 25

    # =====================================================
    # SECTION
    # =====================================================
    if view_type in {"section", "all"}:
        ws = wb.create_sheet("Section Timetable")

        def section_title(table):
            section = table.get("section")
            dept = getattr(section, "department", None)
            return f"{_safe_get(section, 'section_id')} ({_safe_get(dept, 'name', default='Dept')})"

        write_timetable(ws, section_tables, section_title)

    # =====================================================
    # ROOM
    # =====================================================
    if view_type in {"room", "all"}:
        ws = wb.create_sheet("Room Timetable")

        def room_title(table):
            room = table.get("room")
            return f"Room: {_safe_get(room, 'r_number', default='Room')}"

        write_timetable(ws, room_tables, room_title)

    # =====================================================
    # TEACHER
    # =====================================================
    if view_type in {"teacher", "all"}:
        ws = wb.create_sheet("Teacher Timetable")

        def teacher_title(table):
            teacher = table.get("teacher")
            return f"Teacher: {_safe_get(teacher, 'name', 'uid', default='Teacher')}"

        write_timetable(ws, teacher_tables, teacher_title)

    # =====================================================
    # WORKLOAD
    # =====================================================
    if view_type in {"workload", "all"}:
        ws = wb.create_sheet("Teacher Workload")

        ws["A1"] = "Teacher Workload Summary"
        ws["A1"].font = Font(bold=True, size=12)

        headers = ["Teacher Name", "Lectures", "Lab Workload", "Shared Lab Workload", "Total Hours"]
        for col, h in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=h)

        row = 3
        for teacher, workload in teacher_workloads.items():
            lectures = workload.get("lectures", workload.get("classes", 0))
            labs_count = workload.get("labs", 0)
            shared_labs = workload.get("shared_labs", 0)
            total_hours = workload.get("total", lectures + labs_count + shared_labs)

            ws.cell(row=row, column=1, value=_safe_get(teacher, "name", "uid"))
            ws.cell(row=row, column=2, value=lectures)
            ws.cell(row=row, column=3, value=labs_count)
            ws.cell(row=row, column=4, value=shared_labs)
            ws.cell(row=row, column=5, value=total_hours)

            row += 1

    # ---- RESPONSE ----
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


# =====================================================
# DOWNLOAD FUNCTIONS (UNCHANGED)
# =====================================================
@login_required
def download_timetable_excel(request, tid, view_type='section'):
    try:
        saved_t = SavedTimetable.objects.get(id=tid)
    except SavedTimetable.DoesNotExist:
        raise Http404("Timetable does not exist")

    if saved_t.user != request.user:
        return HttpResponseForbidden("You do not have permission")

    classes, labs = _rebuild_classes_and_labs_from_saved(saved_t)

    # Apply the same program + department filters used by the saved view so the
    # Excel export matches whatever the user is currently looking at.
    selected_program = request.GET.get("program", "all")
    classes, labs, _selected_program = _filter_entities_by_program(
        classes, labs, request.user, selected_program
    )

    dept_filter = None
    multi_dept_filter = _parse_dept_filter(request)
    if multi_dept_filter:
        dept_filter = multi_dept_filter
    else:
        selected_department = _normalize_selected_department(
            request.user, request.GET.get("department", "all")
        )
        if selected_department and selected_department != "all":
            dept_filter = {selected_department}

    return _build_timetable_excel_response(
        classes=classes,
        labs=labs,
        user=request.user,
        filename=f"timetable_{tid}.xlsx",
        view_type=view_type,
        dept_filter=dept_filter,
    )


@login_required
def download_generated_timetable_excel(request, index, view_type='section'):
    try:
        idx = int(index)
    except:
        raise Http404("Invalid index")

    state = _get_user_state(request.user.id)
    schedules = state.get("schedules") or GLOBAL_GENERATED_SCHEDULES or []

    if not schedules or idx < 1 or idx > len(schedules):
        raise Http404("Invalid timetable")

    selected = schedules[idx - 1]
    classes = list(selected.get("classes", []))
    labs = list(selected.get("labs", []))

    return _build_timetable_excel_response(
        classes=classes,
        labs=labs,
        user=request.user,
        filename=f"generated_timetable_{idx}.xlsx",
        view_type=view_type,
        dept_filter=_parse_dept_filter(request),
    )


# ============================================================
# DOWNLOAD CENTER — dept selection + clean PDF export
# ============================================================
COLLEGE_NAME = "J.C. Bose University of Science and Technology, YMCA (Formerly YMCA UST)"


def _pdf_link_callback(uri, rel):
    """Resolve static/media URIs to absolute filesystem paths for xhtml2pdf."""
    import os as _os
    from django.conf import settings as _settings

    if uri.startswith(("http://", "https://", "data:")):
        return uri

    static_url = _settings.STATIC_URL or "/static/"
    media_url = getattr(_settings, "MEDIA_URL", "") or ""

    path = None
    if media_url and uri.startswith(media_url):
        path = _os.path.join(getattr(_settings, "MEDIA_ROOT", ""), uri[len(media_url):])
    elif uri.startswith(static_url):
        rel_path = uri[len(static_url):]
        for base in [getattr(_settings, "STATIC_ROOT", None)] + list(getattr(_settings, "STATICFILES_DIRS", [])):
            if not base:
                continue
            candidate = _os.path.join(str(base), rel_path)
            if _os.path.isfile(candidate):
                path = candidate
                break
    if path and _os.path.isfile(path):
        return _os.path.abspath(path)
    return uri


def _generated_tables_for_download(request, index, dept_filter=None):
    """Return (selected_schedule, decorated tables, departments) for a generated index.

    dept_filter: optional set of dept codes to keep. None / empty = all.
    """
    idx = int(index)
    state = _get_user_state(request.user.id)
    schedules = state.get("schedules") or GLOBAL_GENERATED_SCHEDULES or []
    if not schedules or idx < 1 or idx > len(schedules):
        return None, [], []

    selected = schedules[idx - 1]
    classes = list(selected.get("classes", []))
    labs = list(selected.get("labs", []))
    tables = build_section_tables(classes, labs, user=request.user)

    departments = []
    seen = set()
    decorated = []
    for table in tables:
        section = table.get("section")
        try:
            dept = section.department
            dept_code = getattr(dept, "code", "") or ""
            dept_name = getattr(dept, "name", "") or dept_code
        except Exception:
            continue
        table["section_department_code"] = dept_code
        table["section_department_name"] = dept_name
        if dept_code and dept_code not in seen:
            seen.add(dept_code)
            departments.append({"code": dept_code, "name": dept_name})
        if dept_filter and dept_code not in dept_filter:
            continue
        decorated.append(table)

    departments.sort(key=lambda d: d["name"])
    return selected, decorated, departments


def _parse_dept_filter(request):
    raw = (request.GET.get("depts") or "").strip()
    if not raw or raw.lower() == "all":
        return None
    return {code.strip() for code in raw.split(",") if code.strip()}


def _pdf_cell_width(cell):
    return 85 * cell.get("colspan", 1)


def _pdf_block_is_compact(rows):
    max_entries_in_cell = 0
    max_lines_in_entry = 0
    total_lines = 0
    for row in rows:
        for cell in row.get("cells", []):
            entries = cell.get("entries", []) or []
            if entries:
                max_entries_in_cell = max(max_entries_in_cell, len(entries))
            for entry in entries:
                lines = len(entry.get("lines", []) or [])
                max_lines_in_entry = max(max_lines_in_entry, lines)
                total_lines += lines
    return max_entries_in_cell >= 2 or max_lines_in_entry >= 5 or total_lines >= 55


def _lab_group_label(lab):
    total_batches = int(getattr(lab, "total_batches", 1) or 1)
    if total_batches <= 1:
        return ""
    batch = int(getattr(lab, "batch", 1) or 1)
    return f"Group {batch}"


def _section_pdf_block(table):
    """Build a PDF block (grid of cells with text entries) for a section table."""
    rows = []
    for row in table["rows"]:
        cells = []
        for cell in row["cells"]:
            ctype = cell.get("type")
            if ctype == "skip":
                cells.append({"type": "skip"})
                continue
            base = {"type": ctype, "colspan": cell.get("colspan", 1), "width": _pdf_cell_width(cell)}
            entries = []
            if ctype == "lab":
                for lab in cell.get("labs", []):
                    group_label = _lab_group_label(lab)
                    title = f"{lab.subject.subject_name} Lab"
                    if group_label:
                        title = f"{title} ({group_label})"
                    lines = [title]
                    lines.append(lab.instructor.name if lab.instructor else "")
                    if getattr(lab, "second_instructor", None):
                        lines.append(lab.second_instructor.name)
                    for co in getattr(lab, "co_instructors", []) or []:
                        lines.append(f"+ {co.name}")
                    if getattr(lab, "room", None):
                        lines.append(f"Room: {lab.room.r_number}")
                    entries.append({"lines": [l for l in lines if l]})
            elif ctype == "class":
                for cls in cell.get("classes", []):
                    lines = [cls.subject.subject_name]
                    lines.append(cls.instructor.name if cls.instructor else "")
                    for co in getattr(cls, "co_instructors", []) or []:
                        lines.append(f"+ {co.name}")
                    if getattr(cls, "room", None):
                        lines.append(f"Room: {cls.room.r_number}")
                    entries.append({"lines": [l for l in lines if l]})
            base["entries"] = entries
            cells.append(base)
        rows.append({"day": row["day"], "cells": cells})
    subtitle = table.get("section_department_name", "")
    program = getattr(table.get("section"), "program_name", "")
    if program:
        subtitle = f"{subtitle} · {program}" if subtitle else program
    return {
        "title": str(table["section"].section_id),
        "subtitle": subtitle,
        "rows": rows,
        "compact": _pdf_block_is_compact(rows),
    }


def _room_pdf_block(table):
    rows = []
    for row in table["rows"]:
        cells = []
        for cell in row["cells"]:
            ctype = cell.get("type")
            base = {"type": ctype, "colspan": cell.get("colspan", 1), "width": _pdf_cell_width(cell)}
            entries = []
            if ctype == "lab":
                for lab in cell.get("labs", []):
                    lines = [f"{lab.subject.subject_name} Lab"]
                    grp = getattr(lab, "group", "")
                    lines.append(f"{lab.section} ({grp})" if grp else str(lab.section))
                    if lab.instructor:
                        lines.append(lab.instructor.name)
                    if getattr(lab, "second_instructor", None):
                        lines.append(lab.second_instructor.name)
                    entries.append({"lines": [l for l in lines if l]})
            elif ctype == "class":
                for cls in cell.get("classes", []):
                    lines = [cls.subject.subject_name, str(cls.section)]
                    if cls.instructor:
                        lines.append(cls.instructor.name)
                    entries.append({"lines": [l for l in lines if l]})
            base["entries"] = entries
            cells.append(base)
        rows.append({"day": row["day"], "cells": cells})
    room = table["room"]
    subtitle = ", ".join(table.get("dept_names", [])) or ""
    util = table.get("optimization_percentage", "")
    if util != "":
        subtitle = f"{subtitle} · Utilization {util}%" if subtitle else f"Utilization {util}%"
    return {
        "title": f"Room {room.r_number} ({room.room_type})",
        "subtitle": subtitle,
        "rows": rows,
        "compact": _pdf_block_is_compact(rows),
    }


def _teacher_pdf_block(table):
    rows = []
    for row in table["rows"]:
        cells = []
        for cell in row["cells"]:
            ctype = cell.get("type")
            base = {"type": ctype, "colspan": cell.get("colspan", 1), "width": _pdf_cell_width(cell)}
            entries = []
            if ctype == "lab":
                for lab in cell.get("labs", []):
                    lines = [f"{lab.subject.subject_name} Lab"]
                    grp = getattr(lab, "group", "")
                    lines.append(f"{lab.section} ({grp})" if grp else str(lab.section))
                    if getattr(lab, "room", None):
                        lines.append(f"Room: {lab.room.r_number}")
                    entries.append({"lines": [l for l in lines if l]})
            elif ctype == "class":
                for cls in cell.get("classes", []):
                    lines = [cls.subject.subject_name, str(cls.section)]
                    if getattr(cls, "room", None):
                        lines.append(f"Room: {cls.room.r_number}")
                    entries.append({"lines": [l for l in lines if l]})
            base["entries"] = entries
            cells.append(base)
        rows.append({"day": row["day"], "cells": cells})
    teacher = table["teacher"]
    wl = table.get("workload", {})
    subtitle = f"UID: {getattr(teacher, 'uid', '')} · Total Load: {wl.get('total', '')}"
    return {
        "title": teacher.name,
        "subtitle": subtitle,
        "rows": rows,
        "compact": _pdf_block_is_compact(rows),
    }


@login_required
def timetable_download_center(request, index):
    """Standalone page to choose departments and download/print/share."""
    selected, _tables, departments = _generated_tables_for_download(request, index)
    if selected is None:
        messages.info(request, "Session expired. View your saved timetables below.")
        return redirect("saved_timetable_list")

    return render(request, "download_center.html", {
        "index": index,
        "departments": departments,
        "college_name": COLLEGE_NAME,
    })


@login_required
def download_generated_timetable_pdf(request, index, view_type='section'):
    try:
        from xhtml2pdf import pisa
    except ImportError:
        return HttpResponse(
            "PDF generation dependencies are not installed on this machine yet.",
            status=503,
        )

    view_type = (view_type or "section").lower()
    if view_type not in {"section", "room", "teacher", "workload"}:
        view_type = "section"

    idx = int(index)
    state = _get_user_state(request.user.id)
    schedules = state.get("schedules") or GLOBAL_GENERATED_SCHEDULES or []
    if not schedules or idx < 1 or idx > len(schedules):
        raise Http404("Invalid timetable")
    selected = schedules[idx - 1]
    classes = list(selected.get("classes", []))
    labs = list(selected.get("labs", []))

    dept_filter = _parse_dept_filter(request)

    def _dept_code_of(section):
        try:
            return getattr(section.department, "code", "") or ""
        except Exception:
            return ""

    blocks = []
    is_workload = False
    workload_rows = []
    view_label = "Section"

    if view_type == "section":
        view_label = "Section"
        _sel, tables, _depts = _generated_tables_for_download(request, index, dept_filter)
        blocks = [_section_pdf_block(t) for t in tables]
    elif view_type == "room":
        view_label = "Room"
        room_tables = build_room_tables(classes, labs, user=request.user)
        if dept_filter:
            room_tables = [
                rt for rt in room_tables
                if any(code in dept_filter for code in rt.get("dept_codes", []))
            ]
        blocks = [_room_pdf_block(t) for t in room_tables]
    elif view_type == "teacher":
        view_label = "Teacher"
        teacher_tables = build_teacher_tables(classes, labs, user=request.user)
        if dept_filter:
            fb = _teacher_home_dept_fallback(classes, labs)
            teacher_tables = [
                t for t in teacher_tables
                if (_teacher_home_dept_code(t.get("teacher"), fb) or "") in dept_filter
            ]
        blocks = [_teacher_pdf_block(t) for t in teacher_tables]
    else:  # workload
        view_label = "Workload"
        is_workload = True
        workloads = _compute_teacher_workloads(classes, labs)
        if dept_filter:
            fb = _teacher_home_dept_fallback(classes, labs)
            workloads = {
                teacher: data for teacher, data in workloads.items()
                if (_teacher_home_dept_code(teacher, fb) or "") in dept_filter
            }
        for teacher, data in workloads.items():
            workload_rows.append({
                "name": teacher.name,
                "uid": getattr(teacher, "uid", ""),
                "departments": data.get("departments", "-"),
                "lectures": data.get("lectures", 0),
                "labs": data.get("labs", 0),
                "shared_labs": data.get("shared_labs", 0),
                "total": data.get("total", 0),
            })
        workload_rows.sort(key=lambda w: str(w["name"]).lower())

    html = render_to_string("generated_timetable_pdf.html", {
        "blocks": blocks,
        "is_workload": is_workload,
        "workload_rows": workload_rows,
        "view_label": view_label,
        "slot_labels": SLOT_LABELS,
        "college_name": COLLEGE_NAME,
        "college_logo": settings.STATIC_URL + "img/college_logo.png",
        "brand_logo": settings.STATIC_URL + "img/logo_email.png",
    })

    response = HttpResponse(content_type='application/pdf')
    inline = request.GET.get("inline") == "1"
    disposition = "inline" if inline else "attachment"
    response['Content-Disposition'] = f'{disposition}; filename="{view_type}_timetable_{idx}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response, link_callback=_pdf_link_callback)
    if pisa_status.err:
        return HttpResponse("PDF creation failed.", status=500)
    return response



# UNIFIED CSV CONVERTER (PDF/Excel → CSV) — All Entity Types
# ============================================================
ENTITY_CONFIGS = {
    "instructors": {
        "label": "Instructors",
        "columns": ["uid", "name", "designation", "max_workload", "email", "contact_number", "department_code"],
        "filename": "instructors.csv",
        "required": ["uid", "name", "email", "contact_number"],
        "keywords": {
            "uid": ["uid", "id", "teacher id", "teacherid", "faculty id", "code", "teacher_id"],
            "name": ["name", "teacher name", "teachername", "faculty name", "instructor", "faculty"],
            "designation": ["designation", "position", "role", "rank", "post"],
            "max_workload": ["max_workload", "workload", "max workload", "load", "hours"],
            "email": ["email", "e-mail", "mail", "email id", "email_id"],
            "contact_number": ["contact_number", "contact", "phone", "mobile", "phone number", "contact no"],
            "department_code": ["department_code", "department", "dept", "dept_code", "department code"],
        },
    },
    "subjects": {
        "label": "Subjects",
        "columns": ["department_code", "subject_number", "subject_name", "room_required", "required_lab_category", "classes_per_week", "duration"],
        "filename": "subjects.csv",
        "required": ["subject_number", "subject_name"],
        "keywords": {
            "department_code": ["department", "dept", "department_code", "dept_code"],
            "subject_number": ["subject_number", "subject no", "subject_id", "subject id", "code", "number", "course_number", "course no", "course_id"],
            "subject_name": ["subject_name", "subject name", "name", "title", "course_name", "course name"],
            "room_required": ["room_required", "room required", "room type", "room"],
            "required_lab_category": ["lab_category", "lab category", "required_lab", "lab"],
            "classes_per_week": ["classes_per_week", "classes per week", "classes", "per week", "frequency", "weekly"],
            "duration": ["duration", "hours", "duration_hours", "slot_duration", "slots"],
        },
    },
    "rooms": {
        "label": "Rooms",
        "columns": ["r_number", "department", "seating_capacity", "room_type", "lab_category"],
        "filename": "rooms.csv",
        "required": ["r_number"],
        "keywords": {
            "r_number": ["r_number", "room number", "room_number", "room no", "room id", "room_id", "number"],
            "department": ["department", "dept", "department_code", "dept_code"],
            "seating_capacity": ["seating_capacity", "seating capacity", "capacity", "seats"],
            "room_type": ["room_type", "room type", "type"],
            "lab_category": ["lab_category", "lab category", "lab", "category"],
        },
    },
    "timings": {
        "label": "Timings",
        "columns": ["pid", "time", "day"],
        "filename": "timings.csv",
        "required": ["pid"],
        "keywords": {
            "pid": ["pid", "period id", "period_id", "period", "slot id"],
            "time": ["time", "slot", "period number", "slot number"],
            "day": ["day", "weekday", "day_number"],
        },
    },
    "departments": {
        "label": "Departments",
        "columns": ["name", "code"],
        "filename": "departments.csv",
        "required": ["name", "code"],
        "keywords": {
            "name": ["name", "dept_name", "department_name", "department name", "department"],
            "code": ["code", "dept_code", "department_code", "department code"],
        },
    },
    "sections": {
        "label": "Sections",
        "columns": ["section_id", "program_name", "department", "student_strength"],
        "filename": "sections.csv",
        "required": ["section_id"],
        "keywords": {
            "section_id": ["section_id", "section id", "section", "id", "batch"],
            "program_name": ["program_name", "program", "degree", "course_type", "program type"],
            "department": ["department", "dept", "department_code", "dept_code"],
            "student_strength": ["student_strength", "student strength", "strength", "students", "size"],
        },
    },
    "section_subjects": {
        "label": "Section-Subject Mapping",
        "columns": ["section_id", "subject_number"],
        "filename": "section_subjects.csv",
        "required": ["section_id", "subject_number"],
        "keywords": {
            "section_id": ["section_id", "section id", "section", "batch"],
            "subject_number": ["subject_number", "subject no", "subject_id", "subject id", "subject", "course_number", "course no", "course"],
        },
    },
    "teacher_subjects": {
        "label": "Teacher-Subject Mapping",
        "columns": ["instructor", "subject_number"],
        "filename": "teacher_subjects.csv",
        "required": ["instructor", "subject_number"],
        "keywords": {
            "instructor": ["instructor", "teacher", "uid", "name", "faculty", "teacher name", "instructor_name"],
            "subject_number": ["subject_number", "subject no", "subject_id", "subject id", "subject", "course_number", "course no", "course"],
        },
    },
}


def _extract_rows_from_file(uploaded_file):
    """Extract rows from PDF or Excel file. Returns (rows, error_msg)."""
    filename = uploaded_file.name.lower()
    rows = []

    if filename.endswith((".xlsx", ".xls")):
        import openpyxl
        wb = openpyxl.load_workbook(uploaded_file, read_only=True)
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if any(cells):
                rows.append(cells)

    elif filename.endswith(".pdf"):
        import pdfplumber
        import tempfile
        import re as _re
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            for chunk in uploaded_file.chunks():
                tmp.write(chunk)
            tmp_path = tmp.name

        with pdfplumber.open(tmp_path) as pdf:
            for page in pdf.pages:
                table = page.extract_table()
                if table and len(table) > 1 and len(table[0]) >= 2:
                    for row in table:
                        cells = [str(c).strip() if c else "" for c in row]
                        if any(cells):
                            rows.append(cells)
                    continue

                table = page.extract_table({
                    "vertical_strategy": "text",
                    "horizontal_strategy": "text",
                })
                if table and len(table) > 1 and len(table[0]) >= 2:
                    for row in table:
                        cells = [str(c).strip() if c else "" for c in row]
                        if any(cells):
                            rows.append(cells)
                    continue

                text = page.extract_text(layout=True)
                if text:
                    for line in text.split("\n"):
                        line = line.strip()
                        if not line:
                            continue
                        parts = _re.split(r'\s{2,}', line)
                        if len(parts) >= 2:
                            rows.append(parts)
                    if rows:
                        continue

                words = page.extract_words(x_tolerance=3, y_tolerance=3, keep_blank_chars=True)
                if words:
                    y_groups = {}
                    for w in words:
                        y_key = round(w["top"] / 5) * 5
                        y_groups.setdefault(y_key, []).append(w)
                    for y_key in sorted(y_groups.keys()):
                        line_words = sorted(y_groups[y_key], key=lambda w: w["x0"])
                        cells = []
                        current = line_words[0]["text"]
                        prev_x1 = line_words[0]["x1"]
                        for w in line_words[1:]:
                            gap = w["x0"] - prev_x1
                            if gap > 15:
                                cells.append(current.strip())
                                current = w["text"]
                            else:
                                current += " " + w["text"]
                            prev_x1 = w["x1"]
                        cells.append(current.strip())
                        if len(cells) >= 2:
                            rows.append(cells)

        os.unlink(tmp_path)
    else:
        return [], "Unsupported file type. Please upload PDF or Excel (.xlsx/.xls)."

    return rows, None


@login_required(login_url='/accounts/login/')
def convert_csv(request):
    import json as _json
    entity_configs_json = _json.dumps({
        k: {"label": v["label"], "columns": v["columns"]}
        for k, v in ENTITY_CONFIGS.items()
    })

    if request.method == "POST":
        entity_type = request.POST.get("entity_type", "instructors")
        config = ENTITY_CONFIGS.get(entity_type)
        if not config:
            messages.error(request, "Invalid entity type.")
            return redirect("convert_csv")

        uploaded_file = request.FILES.get("file")
        if not uploaded_file:
            messages.error(request, "Please upload a file.")
            return redirect("convert_csv")

        try:
            rows, err = _extract_rows_from_file(uploaded_file)
        except Exception as e:
            messages.error(request, f"Error reading file: {e}")
            return redirect("convert_csv")

        if err:
            messages.error(request, err)
            return redirect("convert_csv")

        rows = [r for r in rows if len(r) >= 2]
        if not rows:
            messages.error(request, "No data found in the uploaded file.")
            return redirect("convert_csv")

        columns = config["columns"]
        keywords = config["keywords"]

        col_map = {}
        data_start = 0
        for row_idx in range(min(5, len(rows))):
            row_lower = [c.lower() for c in rows[row_idx]]
            temp_map = {}
            for field, kws in keywords.items():
                for i, cell in enumerate(row_lower):
                    if any(kw in cell for kw in kws):
                        temp_map[field] = i
                        break
            if len(temp_map) >= 2:
                col_map = temp_map
                data_start = row_idx + 1
                break

        if not col_map:
            for i, col in enumerate(columns):
                if i < len(rows[0]):
                    col_map[col] = i

        import io
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)

        required = config["required"]
        for row in rows[data_start:]:
            vals = {}
            for col in columns:
                vals[col] = row[col_map[col]] if col in col_map and col_map[col] < len(row) else ""
            if all(vals.get(r) for r in required):
                writer.writerow([vals[col] for col in columns])

        response = HttpResponse(output.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = f'attachment; filename="{config["filename"]}"'
        return response

    return render(request, "convert_csv.html", {
        "entity_configs_json": entity_configs_json,
    })


# ═══════════════════════════════════════════════════════════════
# TEACHER PREFERENCE VIEWS
# ═══════════════════════════════════════════════════════════════
import csv as _csv, io as _io, re as _re, json as _json
from django.http import HttpResponse as _HR
from django.core.mail import send_mail as _sm
from django.conf import settings as _cfg

_ERE = _re.compile(r'[\w\.\+\-]+@[\w\.\-]+\.[a-zA-Z]{2,}')

def teacher_pref_form(request):
    return render(request, 'teacher_pref_form.html')

def send_preferences_page(request):
    return render(request, 'send_preferences.html')

def teacher_responses_page(request):
    from .models import TeacherPreference
    subs = list(TeacherPreference.objects.all().values(
        'id','name','email','designation','subjects','classes','years','submitted_at'))
    for s in subs:
        s['submitted_at'] = _format_local_datetime(s['submitted_at'], '%d %b %Y, %I:%M %p')
    return render(request, 'teacher_responses.html', {
        'submissions_json': _json.dumps(subs), 'total': len(subs)})

@csrf_exempt
def teacher_pref_submit(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)
    try:
        from .models import TeacherPreference
        d = _json.loads(request.body)
        name=d.get('name','').strip(); email=d.get('email','').strip()
        desg=d.get('designation','').strip()
        subj=d.get('subjects',[]); cls=d.get('classes',[]); yrs=d.get('years',[])
        err = []
        if not name: err.append('Name required.')
        if not email or '@' not in email: err.append('Valid email required.')
        if not desg: err.append('Designation required.')
        if not subj: err.append('Select at least one subject.')
        if len(subj) > 3: err.append('Max 3 subjects.')
        if not cls: err.append('Select at least one class.')
        if not yrs: err.append('Select at least one year.')
        if err: return JsonResponse({'ok': False, 'errors': err}, status=400)
        sub = TeacherPreference.objects.create(
            name=name, email=email, designation=desg,
            subjects=subj, classes=cls, years=yrs)
        try:
            _sm(subject=f'SmartScheduler — Preferences Received: {name}',
                message=f'Name: {name}\nEmail: {email}\nDesignation: {desg}\nSubjects: {", ".join(subj)}\nClasses: {", ".join(cls)}\nYears: {", ".join(yrs)}',
                from_email=_brand_from_email(),
                recipient_list=[email, _cfg.EMAIL_HOST_USER], fail_silently=True)
        except: pass
        return JsonResponse({'ok': True, 'id': sub.id})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@csrf_exempt
def send_pref_links_smtp(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)
    try:
        d = _json.loads(request.body); emails = d.get('emails', [])
        base = request.build_absolute_uri('/teacher-pref-form/')
        sent, failed = [], []
        for email in emails:
            email = email.strip().lower()
            if not _ERE.match(email): failed.append(email); continue
            try:
                _sm(subject='SmartScheduler — Fill Your Teaching Preferences',
                    message=f'Dear Teacher,\n\nSmartScheduler by the SIH Winner Innovation Team\n\nFill your preferences:\n{base}?email={email}\n\nThank you,\nSmartScheduler Team\n{SIH_WINNER_TEXT}',
                    from_email=_brand_from_email(),
                    recipient_list=[email], fail_silently=False)
                sent.append(email)
            except: failed.append(email)
        return JsonResponse({'ok': True, 'sent': sent, 'failed': failed})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

@csrf_exempt
def parse_emails_view(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'POST required'}, status=405)
    try:
        f = request.FILES.get('file')
        if not f: return JsonResponse({'ok': False, 'error': 'No file.'}, status=400)
        n = f.name.lower(); raw = ''
        if n.endswith('.csv'):
            content = f.read()
            try: t = content.decode('utf-8')
            except: t = content.decode('latin-1', errors='replace')
            raw = ' '.join(cell for row in _csv.reader(_io.StringIO(t)) for cell in row)
        elif n.endswith('.txt'):
            content = f.read()
            try: raw = content.decode('utf-8')
            except: raw = content.decode('latin-1', errors='replace')
        elif n.endswith('.pdf'):
            try:
                from pypdf import PdfReader
                raw = '\n'.join(p.extract_text() or '' for p in PdfReader(_io.BytesIO(f.read())).pages)
            except Exception as e:
                return JsonResponse({'ok': False, 'error': f'PDF error: {e}'}, status=400)
        else:
            return JsonResponse({'ok': False, 'error': 'Upload CSV, TXT, or PDF.'}, status=400)
        found = _ERE.findall(raw); seen = set(); emails = []
        for e in found:
            e = e.lower().strip('.')
            if e not in seen: seen.add(e); emails.append(e)
        if not emails:
            return JsonResponse({'ok': False, 'error': 'No emails found.'}, status=400)
        return JsonResponse({'ok': True, 'emails': emails, 'count': len(emails)})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=500)

def export_preferences_csv(request):
    from .models import TeacherPreference
    response = _HR(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="teacher_preferences.csv"'
    w = _csv.writer(response)
    w.writerow(['Name','Email','Designation','Subjects','Classes','Years','Submitted'])
    for s in TeacherPreference.objects.all():
        w.writerow([s.name, s.email, s.designation,
            ', '.join(s.subjects), ', '.join(s.classes), ', '.join(s.years),
            _format_local_datetime(s.submitted_at, '%d %b %Y %H:%M')])
    return response
